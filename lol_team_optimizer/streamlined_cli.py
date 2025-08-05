"""
Streamlined CLI for League of Legends Team Optimizer.

This module provides a simplified 4-option interface that consolidates
all functionality into intuitive workflows.
"""

import sys
import logging
import json
from datetime import datetime
from typing import List, Optional, Dict, Any

from .core_engine import CoreEngine
from .analytics_help_system import analytics_help, show_contextual_help, show_help_menu, show_help_topic


class StreamlinedCLI:
    """
    Streamlined command-line interface with 4 main options:
    1. Quick Optimize - Integrated optimization workflow
    2. Manage Players - Consolidated player management
    3. View Analysis - Comprehensive analysis dashboard
    4. Settings - Configuration and maintenance
    """
    
    def __init__(self):
        """Initialize the streamlined CLI."""
        self.logger = logging.getLogger(__name__)
        
        try:
            self.engine = CoreEngine()
            self.logger.info("Streamlined CLI initialized successfully")
        except Exception as e:
            print(f"Failed to initialize system: {e}")
            print("Please check your configuration and try again.")
            sys.exit(1)
    
    def main(self) -> None:
        """Main entry point for the streamlined CLI."""
        print("=" * 60)
        print("    League of Legends Team Optimizer")
        print("    Streamlined Interface v2.0")
        print("=" * 60)
        
        # Show welcome message for first-time users
        if self.engine.system_status.get('player_count', 0) == 0:
            print("\n🎉 Welcome to the streamlined interface!")
            print("This new interface consolidates all functionality into 4 simple options:")
            print("• Quick Optimize: One-click team optimization with smart defaults")
            print("• Manage Players: All player operations in one place")
            print("• View Analysis: Comprehensive insights and comparisons")
            print("• Settings: System configuration and maintenance")
            print("\nType 'h' anytime for detailed help!")
        
        # Display system status
        self._display_system_status()
        
        while True:
            try:
                self._display_main_menu()
                choice = input("\nEnter your choice (1-6, or 'h' for help): ").strip().lower()
                
                if choice == "1":
                    self._quick_optimize()
                elif choice == "2":
                    self._manage_players()
                elif choice == "3":
                    self._view_analysis()
                elif choice == "4":
                    self._settings()
                elif choice == "5":
                    self._historical_match_browser()
                elif choice == "6":
                    print("\nThank you for using League of Legends Team Optimizer!")
                    self.logger.info("Application shutdown requested by user")
                    break
                elif choice == "h" or choice == "help":
                    self._show_help()
                elif choice == "a":
                    self._quick_add_player()
                elif choice == "o":
                    self._quick_optimize()
                else:
                    print("\nInvalid choice. Please enter a number between 1-6, or use quick actions (a/o/h).")
                
                input("\nPress Enter to continue...")
                
            except KeyboardInterrupt:
                print("\n\nExiting application...")
                self.logger.info("Application interrupted by user")
                break
            except Exception as e:
                self.logger.error(f"Unexpected error in main loop: {e}", exc_info=True)
                print(f"\nAn unexpected error occurred: {e}")
                print("The application will continue, but please report this issue.")
                input("\nPress Enter to continue...")
    
    def _display_system_status(self) -> None:
        """Display current system status."""
        status = self.engine.system_status
        
        print(f"\n📊 System Status:")
        print(f"   Players: {status.get('player_count', 0)} registered")
        print(f"   API: {'✅ Available' if status.get('api_available') else '⚠️ Offline'}")
        print(f"   Ready for optimization: {'✅ Yes' if status.get('ready_for_optimization') else '❌ No'}")
        
        # Show data quality indicators
        if status.get('player_count', 0) > 0:
            players_with_data = status.get('players_with_api_data', 0)
            players_with_prefs = status.get('players_with_preferences', 0)
            print(f"   Data Quality: {players_with_data}/{status.get('player_count', 0)} with API data, {players_with_prefs}/{status.get('player_count', 0)} with custom preferences")
        
        if not status.get('api_available'):
            print("   ⚠️ Running in offline mode - limited functionality")
        
        if status.get('player_count', 0) < 5:
            print(f"   ℹ️ Add {5 - status.get('player_count', 0)} more players for full team optimization")
        
        # Show champion data status
        champion_count = len(self.engine.champion_data_manager.champions)
        if champion_count > 0:
            print(f"   Champions: {champion_count} loaded")
        else:
            print("   ⚠️ No champion data loaded - some features may be limited")
        
        print()
    
    def _show_help(self) -> None:
        """Display help information for the streamlined interface."""
        print("\n" + "=" * 60)
        print("🆘 HELP - Streamlined Interface Guide")
        print("=" * 60)
        
        print("\n🎯 Quick Optimize:")
        print("   • Automatically finds the best team composition")
        print("   • Handles player selection and data fetching")
        print("   • Shows comprehensive results with champion recommendations")
        print("   • Offers follow-up actions like alternatives and analysis")
        
        print("\n👥 Manage Players:")
        print("   • Add new players with automatic API data fetching")
        print("   • Edit existing players and their preferences")
        print("   • Remove players or perform bulk operations")
        print("   • View detailed player information and statistics")
        
        print("\n📊 View Analysis:")
        print("   • Team overview and optimization readiness")
        print("   • Individual player analysis and comparisons")
        print("   • Role coverage and champion pool analysis")
        print("   • Team synergy analysis with visual representation")
        print("   • Performance trends and recommendations")
        
        print("\n⚙️ Settings:")
        print("   • System diagnostics and health checks")
        print("   • Cache management and data updates")
        print("   • API configuration and testing")
        print("   • Export logs and system information")
        
        print("\n💡 Tips:")
        print("   • Start by adding at least 5 players for full optimization")
        print("   • Set role preferences for better optimization results")
        print("   • Use Quick Optimize for the fastest workflow")
        print("   • Check system status regularly for API and data health")
        
        print("\n🔗 Quick Workflow:")
        print("   1. Add players (option 2)")
        print("   2. Run optimization (option 1)")
        print("   3. View detailed analysis (option 3)")
        print("   4. Adjust and re-optimize as needed")
    
    def _quick_add_player(self) -> None:
        """Quick player addition from main menu."""
        print("\n⚡ Quick Add Player")
        print("-" * 20)
        
        name = input("Player name: ").strip()
        if not name:
            print("❌ Player name is required.")
            return
        
        riot_id = input("Riot ID (gameName#tagLine): ").strip()
        if not riot_id:
            print("❌ Riot ID is required.")
            return
        
        print("🔄 Adding player...")
        success, message, player = self.engine.add_player_with_data(name, riot_id)
        
        if success:
            print(f"✅ {message}")
            # Update system status
            self.engine.system_status = self.engine._get_system_status()
        else:
            print(f"❌ {message}")
    
    def _show_quick_player_overview(self) -> None:
        """Show a quick overview of all players for optimization context."""
        print("\n📋 Quick Player Overview")
        print("-" * 30)
        
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("No players found.")
            return
        
        for i, player in enumerate(players, 1):
            # Data indicators
            indicators = []
            if player.performance_cache:
                indicators.append("📊")
            if player.champion_masteries:
                indicators.append("🏆")
            if any(pref != 3 for pref in player.role_preferences.values()):
                indicators.append("⭐")
            
            status_str = "".join(indicators) if indicators else "📝"
            
            # Top preferred roles
            top_roles = []
            for role, pref in player.role_preferences.items():
                if pref >= 4:
                    top_roles.append(role)
            
            roles_str = f" | Prefers: {', '.join(top_roles[:3])}" if top_roles else " | No strong preferences"
            
            print(f"{i:2}. {player.name:15} {status_str}{roles_str}")
        
        print("\nLegend: 📊 Performance | 🏆 Champions | ⭐ Preferences | 📝 Basic")
    
    def _check_players_needing_data(self, player_selection: Optional[List[str]]) -> List[str]:
        """Check which players need data enhancement."""
        players = self.engine.data_manager.load_player_data()
        
        if player_selection:
            # Filter to selected players
            players = [p for p in players if p.name in player_selection]
        
        players_needing_data = []
        for player in players:
            if not player.performance_cache or not player.champion_masteries:
                players_needing_data.append(player.name)
        
        return players_needing_data
    
    def _enhance_player_data(self, player_names: List[str]) -> None:
        """Enhance data for specified players."""
        operations = [{'action': 'refresh_data', 'name': name} for name in player_names]
        results = self.engine.bulk_player_operations(operations)
        
        success_count = sum(1 for success, _ in results if success)
        print(f"✅ Enhanced data for {success_count}/{len(player_names)} players")
    
    def _check_players_needing_preferences(self, player_selection: Optional[List[str]]) -> List[str]:
        """Check which players need custom preferences set."""
        players = self.engine.data_manager.load_player_data()
        
        if player_selection:
            # Filter to selected players
            players = [p for p in players if p.name in player_selection]
        
        players_needing_preferences = []
        for player in players:
            # Check if player has default preferences (all 3s) or missing preferences
            if not player.role_preferences or all(pref == 3 for pref in player.role_preferences.values()):
                players_needing_preferences.append(player.name)
        
        return players_needing_preferences
    
    def _prompt_for_missing_preferences(self, player_names: List[str]) -> None:
        """Prompt user to set preferences for players with missing data."""
        print("\n⭐ Setting Custom Role Preferences")
        print("-" * 35)
        print("Rate each role from 1-5 (1=Hate, 2=Dislike, 3=Neutral, 4=Like, 5=Love)")
        print("Press Enter to keep current preference or skip player")
        
        roles = ["top", "jungle", "middle", "support", "bottom"]
        
        for player_name in player_names:
            player = self.engine.data_manager.get_player_by_name(player_name)
            if not player:
                continue
            
            print(f"\n👤 Setting preferences for {player_name}:")
            
            # Show current preferences
            current_prefs = player.role_preferences or {role: 3 for role in roles}
            print("Current preferences:", end=" ")
            for role in roles:
                pref = current_prefs.get(role, 3)
                print(f"{role}:{pref}", end=" ")
            print()
            
            # Ask if user wants to set preferences for this player
            set_for_player = input(f"Set custom preferences for {player_name}? (Y/n/s=skip all): ").strip().lower()
            if set_for_player == 's':
                print("⏭️ Skipping preference setting for remaining players")
                break
            elif set_for_player == 'n':
                continue
            
            new_preferences = {}
            preferences_changed = False
            
            for role in roles:
                current_pref = current_prefs.get(role, 3)
                
                while True:
                    pref_input = input(f"  {role.capitalize():8} (current: {current_pref}): ").strip()
                    
                    if not pref_input:  # Keep current
                        new_preferences[role] = current_pref
                        break
                    
                    try:
                        pref_value = int(pref_input)
                        if 1 <= pref_value <= 5:
                            new_preferences[role] = pref_value
                            if pref_value != current_pref:
                                preferences_changed = True
                            break
                        else:
                            print("    Please enter a number between 1-5")
                    except ValueError:
                        print("    Please enter a valid number")
            
            # Save preferences if changed
            if preferences_changed:
                operations = [{'action': 'update_preferences', 'name': player_name, 'preferences': new_preferences}]
                results = self.engine.bulk_player_operations(operations)
                
                if results and results[0][0]:  # Check if successful
                    print(f"✅ Preferences updated for {player_name}")
                else:
                    print(f"❌ Failed to update preferences for {player_name}")
            else:
                print(f"ℹ️ No changes made for {player_name}")
        
        print("\n✅ Preference setting complete!")
    
    def _handle_optimization_failure(self, message: str) -> None:
        """Handle optimization failure with helpful suggestions."""
        print("\n🔧 Troubleshooting Suggestions:")
        
        if "insufficient" in message.lower() or "not enough" in message.lower():
            print("• Add more players to the system")
            print("• Check that players have valid Riot IDs")
            print("• Ensure at least 2 players are available")
        elif "api" in message.lower():
            print("• Check your internet connection")
            print("• Verify your RIOT_API_KEY is valid")
            print("• Try again in a few minutes (rate limiting)")
        elif "data" in message.lower():
            print("• Update player preferences in Manage Players")
            print("• Refresh player data from the API")
            print("• Check system diagnostics in Settings")
        else:
            print("• Try with different players")
            print("• Check system status in Settings")
            print("• Restart the application if issues persist")
        
        retry = input("\nWould you like to try optimization again? (y/n): ").strip().lower()
        if retry == 'y':
            self._quick_optimize()
    
    def _display_main_menu(self) -> None:
        """Display the streamlined main menu."""
        print("\n" + "=" * 50)
        print("MAIN MENU")
        print("=" * 50)
        print("1. 🎯 Quick Optimize")
        print("   └─ Smart team optimization with automatic setup")
        print("2. 👥 Manage Players") 
        print("   └─ Add, edit, remove players and preferences")
        print("3. 📊 View Analysis")
        print("   └─ Comprehensive player and team analysis")
        print("4. ⚙️ Settings")
        print("   └─ Configuration, diagnostics, and maintenance")
        print("5. 📜 Historical Match Browser")
        print("   └─ View and analyze historical match data")
        print("6. 🚪 Exit")
        
        # Show quick tips based on system status
        status = self.engine.system_status
        if status.get('player_count', 0) == 0:
            print("\n💡 Tip: Start by adding players in option 2")
        elif status.get('player_count', 0) < 5:
            print(f"\n💡 Tip: Add {5 - status.get('player_count', 0)} more players for optimal results")
        elif status.get('ready_for_optimization'):
            print("\n💡 Tip: You're ready for optimization! Try option 1")
        
        # Show quick actions
        print("\n⚡ Quick Actions: 'a' = Add Player, 'o' = Quick Optimize, 'h' = Help")
    
    def _quick_optimize(self) -> None:
        """Quick optimization workflow with smart defaults and inline management."""
        print("\n" + "=" * 50)
        print("🎯 QUICK OPTIMIZE - Smart Team Optimization")
        print("=" * 50)
        
        # Check if we have enough players
        status = self.engine.system_status
        player_count = status.get('player_count', 0)
        
        if player_count == 0:
            print("🚀 Welcome to Quick Optimize!")
            print("This streamlined workflow will guide you through team optimization.")
            print("\nNo players found. Let's add some players first!")
            self._add_players_for_optimization()
            
            # Check again after adding players
            status = self.engine.system_status = self.engine._get_system_status()
            player_count = status.get('player_count', 0)
            
            if player_count == 0:
                print("\n❌ No players added. Cannot proceed with optimization.")
                return
        
        # Smart recommendations based on player count
        if player_count < 5:
            print(f"\n📊 Current Status: {player_count} players registered")
            print(f"💡 Recommendation: Add {5 - player_count} more players for optimal results")
            print("\nOptions:")
            print("1. 🔄 Add more players (recommended)")
            print("2. ⚡ Optimize with current players (some roles may be suboptimal)")
            print("3. 📊 View current players first")
            print("4. ❌ Cancel")
            
            choice = input("\nEnter your choice (1-4): ").strip()
            if choice == "1":
                self._add_players_for_optimization()
                # Refresh status and continue
                status = self.engine.system_status = self.engine._get_system_status()
                player_count = status.get('player_count', 0)
                if player_count == 0:
                    return
            elif choice == "3":
                self._show_quick_player_overview()
                return
            elif choice == "4":
                return
            # Continue with option 2
        
        # Smart player selection with data quality assessment
        print(f"\n🎯 Player Selection ({player_count} available):")
        
        # Show data quality summary
        players_with_data = status.get('players_with_api_data', 0)
        players_with_prefs = status.get('players_with_preferences', 0)
        
        print(f"📊 Data Quality: {players_with_data}/{player_count} with API data, {players_with_prefs}/{player_count} with preferences")
        
        print("\nSelection Options:")
        print("1. 🤖 Auto-select best players (recommended)")
        print("2. 👤 Select specific players")
        print("3. 🌐 Use all players")
        print("4. 📋 View player details first")
        
        selection_choice = input("\nEnter your choice (1-4): ").strip()
        
        if selection_choice == "4":
            self._show_quick_player_overview()
            # Ask again after showing overview
            selection_choice = input("\nNow choose selection method (1-3): ").strip()
        
        player_selection = None
        auto_select = True
        
        if selection_choice == "2":
            player_selection = self._select_specific_players()
            auto_select = False
            if not player_selection:
                return
        elif selection_choice == "3":
            auto_select = False
        
        # Show what will happen
        if auto_select and not player_selection:
            print(f"\n🤖 Auto-selecting best players based on data completeness and performance...")
        elif player_selection:
            print(f"\n👤 Using {len(player_selection)} selected players...")
        else:
            print(f"\n🌐 Using all {player_count} players...")
        
        # Pre-optimization checks and data enhancement
        print("\n🔍 Pre-optimization Analysis...")
        
        # Check for missing data and offer to enhance
        if self.engine.api_available:
            players_needing_data = self._check_players_needing_data(player_selection)
            if players_needing_data:
                print(f"⚠️ Found {len(players_needing_data)} players with incomplete data")
                enhance = input("Fetch missing data before optimization? (Y/n): ").strip().lower()
                if enhance != 'n':
                    print("🔄 Enhancing player data...")
                    self._enhance_player_data(players_needing_data)
        
        # Check for missing preferences and prompt for them
        players_needing_preferences = self._check_players_needing_preferences(player_selection)
        if players_needing_preferences:
            print(f"\n⭐ Found {len(players_needing_preferences)} players with default preferences")
            print("Setting custom preferences will improve optimization accuracy.")
            set_prefs = input("Set custom preferences now? (Y/n): ").strip().lower()
            if set_prefs != 'n':
                self._prompt_for_missing_preferences(players_needing_preferences)
        
        # Run optimization with progress indication
        print("\n🎯 Running optimization...")
        if not status.get('api_available'):
            print("⚠️ Running in offline mode - using cached data and preferences")
        else:
            print("✅ Using API data for enhanced accuracy")
        
        success, message, result = self.engine.optimize_team_smart(player_selection, auto_select)
        
        if not success:
            print(f"\n❌ Optimization failed: {message}")
            self._handle_optimization_failure(message)
            return
        
        print(f"\n🎉 {message}")
        
        # Display comprehensive results
        self._display_optimization_results(result)
        
        # Intelligent follow-up actions
        self._optimization_follow_up(result)
    
    def _add_players_for_optimization(self) -> None:
        """Add players specifically for optimization."""
        print("\n📝 Adding Players for Optimization")
        print("-" * 35)
        
        while True:
            print(f"\nCurrent players: {self.engine.system_status.get('player_count', 0)}")
            print("Add a new player:")
            
            name = input("Player name: ").strip()
            if not name:
                break
            
            riot_id = input("Riot ID (gameName#tagLine): ").strip()
            if not riot_id:
                break
            
            print("🔄 Adding player...")
            success, message, player = self.engine.add_player_with_data(name, riot_id)
            
            if success:
                print(f"✅ {message}")
                
                # Update system status
                self.engine.system_status = self.engine._get_system_status()
                
                if self.engine.system_status.get('player_count', 0) >= 5:
                    print("\n🎉 You now have enough players for optimization!")
                    break
                
                continue_adding = input("\nAdd another player? (y/n): ").strip().lower()
                if continue_adding != 'y':
                    break
            else:
                print(f"❌ {message}")
                retry = input("Try again? (y/n): ").strip().lower()
                if retry != 'y':
                    break
    
    def _select_specific_players(self) -> Optional[List[str]]:
        """Allow user to select specific players for optimization."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("No players available.")
            return None
        
        print("\nAvailable players:")
        for i, player in enumerate(players, 1):
            status_indicators = []
            if player.performance_cache:
                status_indicators.append("📊")
            if player.champion_masteries:
                status_indicators.append("🏆")
            if any(pref != 3 for pref in player.role_preferences.values()):
                status_indicators.append("⭐")
            
            status_str = "".join(status_indicators) if status_indicators else "📝"
            print(f"{i:2}. {player.name} ({player.summoner_name}) {status_str}")
        
        print("\nLegend: 📊 Performance Data | 🏆 Champion Data | ⭐ Custom Preferences | 📝 Basic Data")
        
        selection_input = input("\nEnter player numbers (e.g., 1,3,5) or 'all': ").strip()
        
        if selection_input.lower() == 'all':
            return [p.name for p in players]
        
        try:
            indices = [int(x.strip()) - 1 for x in selection_input.split(',')]
            selected_players = []
            
            for idx in indices:
                if 0 <= idx < len(players):
                    selected_players.append(players[idx].name)
                else:
                    print(f"Invalid player number: {idx + 1}")
                    return None
            
            return selected_players
            
        except ValueError:
            print("Invalid input format. Please use numbers separated by commas.")
            return None
    
    def _display_optimization_results(self, result) -> None:
        """Display comprehensive optimization results with enhanced information."""
        if not result or not result.assignments:
            print("No optimization results to display.")
            return
        
        best = result.best_assignment
        
        print("\n" + "=" * 70)
        print("🏆 OPTIMIZATION RESULTS - BEST TEAM COMPOSITION")
        print("=" * 70)
        
        # Summary statistics
        team_size = len(best.assignments)
        avg_individual_score = sum(best.individual_scores.values()) / len(best.individual_scores) if best.individual_scores else 0
        
        print(f"📊 Summary:")
        print(f"   Total Score: {best.total_score:.2f} | Team Size: {team_size}/5 | Avg Individual: {avg_individual_score:.2f}")
        print(f"   Optimization Time: {result.optimization_time:.2f}s | Alternatives Found: {len(result.assignments)}")
        print()
        
        # Role assignments with comprehensive information
        print("🎯 Role Assignments & Champion Recommendations:")
        print("-" * 60)
        
        roles = ["top", "jungle", "middle", "support", "bottom"]
        for role in roles:
            player_name = best.assignments.get(role, "Unassigned")
            individual_score = best.individual_scores.get(player_name, 0)
            
            # Get player preference for this role if available
            try:
                players = self.engine.data_manager.load_player_data()
                player_obj = next((p for p in players if p.name == player_name), None)
                role_pref = player_obj.role_preferences.get(role, 0) if player_obj else 0
                pref_indicator = f"(Pref: {role_pref}/5)" if role_pref > 0 else ""
            except:
                pref_indicator = ""
            
            # Score indicator
            score_indicator = "🟢" if individual_score >= 4 else "🟡" if individual_score >= 2 else "🔴"
            
            print(f"{role.upper():8} │ {player_name:18} │ {score_indicator} {individual_score:.2f} {pref_indicator}")
            
            # Champion recommendations with enhanced display
            if role in best.champion_recommendations and best.champion_recommendations[role]:
                print(f"         │ 🏆 Top Champions:")
                for i, rec in enumerate(best.champion_recommendations[role][:3], 1):
                    suitability = "Perfect" if rec.role_suitability >= 0.9 else "Excellent" if rec.role_suitability >= 0.7 else "Good"
                    mastery_indicator = "🥇" if rec.mastery_level == 7 else "🥈" if rec.mastery_level >= 5 else "🥉"
                    print(f"         │   {i}. {rec.champion_name:12} {mastery_indicator} L{rec.mastery_level} ({rec.mastery_points:,} pts) - {suitability}")
            else:
                print(f"         │ ⚠️ No champion data available")
            print()
        
        # Team synergies with enhanced display
        if best.synergy_scores:
            print("🤝 Team Synergies:")
            print("-" * 20)
            sorted_synergies = sorted(best.synergy_scores.items(), key=lambda x: x[1], reverse=True)
            
            for i, ((player1, player2), synergy) in enumerate(sorted_synergies[:5], 1):
                synergy_indicator = "🔥" if synergy >= 4 else "✨" if synergy >= 2 else "⚡"
                print(f"  {i}. {player1:12} + {player2:12} {synergy_indicator} {synergy:.2f}")
            
            if len(sorted_synergies) > 5:
                print(f"     ... and {len(sorted_synergies) - 5} more synergy pairs")
            print()
        
        # Team strength analysis
        self._display_team_strength_summary(best)
        
        # Explanation
        if hasattr(best, 'explanation') and best.explanation:
            print("🧠 Optimization Reasoning:")
            print("-" * 25)
            print(best.explanation)
            print()
    
    def _optimization_follow_up(self, result) -> None:
        """Intelligent follow-up actions after optimization."""
        print("\n🎯 What would you like to do next?")
        print("-" * 35)
        
        # Smart suggestions based on results
        alternatives_available = len(result.assignments) > 1
        team_size = len(result.best_assignment.assignments)
        
        print("1. 🔄 Re-optimize with different settings")
        if alternatives_available:
            print("2. 🔀 View alternative team compositions")
        else:
            print("2. 📊 Analyze current team composition")
        print("3. 👥 Analyze individual players")
        print("4. 💾 Export results")
        print("5. ➕ Add more players and re-optimize")
        print("6. 🏠 Return to main menu")
        
        # Show smart recommendations
        if team_size < 5:
            print(f"\n💡 Recommendation: Add {5 - team_size} more players for complete team")
        elif not alternatives_available:
            print("\n💡 Recommendation: Try different player selection for alternatives")
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == "1":
            print("\n🔄 Starting new optimization...")
            self._quick_optimize()
        elif choice == "2":
            if alternatives_available:
                self._show_alternative_compositions(result)
            else:
                self._analyze_optimization_players(result)
        elif choice == "3":
            self._analyze_optimization_players(result)
        elif choice == "4":
            self._export_results(result)
        elif choice == "5":
            print("\n➕ Adding more players...")
            self._add_players_for_optimization()
            # Offer to re-optimize
            reopt = input("\nRe-optimize with new players? (Y/n): ").strip().lower()
            if reopt != 'n':
                self._quick_optimize()
        # Choice 6 or invalid returns to main menu
    
    def _show_alternative_compositions(self, result) -> None:
        """Show alternative team compositions from optimization results."""
        if not result or len(result.assignments) <= 1:
            print("\n❌ No alternative compositions available")
            return
        
        print("\n" + "=" * 60)
        print("🔀 ALTERNATIVE TEAM COMPOSITIONS")
        print("=" * 60)
        
        # Show top 3 alternatives (excluding the best one already shown)
        alternatives = result.assignments[1:4]  # Skip the best (first) one
        
        for i, assignment in enumerate(alternatives, 2):  # Start from 2 since best is #1
            print(f"\n🏅 Alternative #{i} (Score: {assignment.total_score:.2f})")
            print("-" * 40)
            
            roles = ["top", "jungle", "middle", "support", "bottom"]
            for role in roles:
                player_name = assignment.assignments.get(role, "Unassigned")
                individual_score = assignment.individual_scores.get(player_name, 0)
                score_indicator = "🟢" if individual_score >= 4 else "🟡" if individual_score >= 2 else "🔴"
                
                print(f"{role.upper():8} │ {player_name:18} │ {score_indicator} {individual_score:.2f}")
            
            # Show key differences from best composition
            best = result.best_assignment
            differences = []
            for role in roles:
                if assignment.assignments.get(role) != best.assignments.get(role):
                    old_player = best.assignments.get(role, "None")
                    new_player = assignment.assignments.get(role, "None")
                    differences.append(f"{role}: {old_player} → {new_player}")
            
            if differences:
                print(f"\n📝 Key Changes: {', '.join(differences[:2])}")
                if len(differences) > 2:
                    print(f"    ... and {len(differences) - 2} more changes")
        
        # Ask if user wants to see more details or select an alternative
        print(f"\nShowing top 3 alternatives out of {len(result.assignments)} total compositions")
        
        choice = input("\nWould you like to (d)etails on specific alternative, (a)ll alternatives, or (r)eturn? ").strip().lower()
        
        if choice == 'd':
            try:
                alt_num = int(input("Enter alternative number for details: ").strip())
                if 2 <= alt_num <= len(result.assignments):
                    selected_alt = result.assignments[alt_num - 1]
                    print(f"\n📋 Detailed view of Alternative #{alt_num}:")
                    self._display_single_composition_details(selected_alt)
                else:
                    print("Invalid alternative number")
            except ValueError:
                print("Please enter a valid number")
        elif choice == 'a':
            self._show_all_alternatives(result)
    
    def _display_single_composition_details(self, assignment) -> None:
        """Display detailed information for a single team composition."""
        print("\n" + "-" * 50)
        
        roles = ["top", "jungle", "middle", "support", "bottom"]
        for role in roles:
            player_name = assignment.assignments.get(role, "Unassigned")
            individual_score = assignment.individual_scores.get(player_name, 0)
            
            print(f"\n{role.upper()} - {player_name} (Score: {individual_score:.2f})")
            
            # Show champion recommendations if available
            if role in assignment.champion_recommendations and assignment.champion_recommendations[role]:
                print("  🏆 Recommended Champions:")
                for i, rec in enumerate(assignment.champion_recommendations[role][:3], 1):
                    print(f"    {i}. {rec.champion_name} (L{rec.mastery_level}, {rec.mastery_points:,} pts)")
        
        # Show team synergies
        if assignment.synergy_scores:
            print("\n🤝 Team Synergies:")
            sorted_synergies = sorted(assignment.synergy_scores.items(), key=lambda x: x[1], reverse=True)
            for (player1, player2), synergy in sorted_synergies[:3]:
                print(f"  {player1} + {player2}: {synergy:.2f}")
    
    def _show_all_alternatives(self, result) -> None:
        """Show a summary of all alternative compositions."""
        print("\n" + "=" * 60)
        print("📊 ALL ALTERNATIVE COMPOSITIONS SUMMARY")
        print("=" * 60)
        
        for i, assignment in enumerate(result.assignments, 1):
            team_size = len(assignment.assignments)
            avg_score = sum(assignment.individual_scores.values()) / len(assignment.individual_scores) if assignment.individual_scores else 0
            
            status = "👑 BEST" if i == 1 else f"#{i}"
            print(f"{status:8} │ Score: {assignment.total_score:5.2f} │ Avg: {avg_score:4.2f} │ Size: {team_size}/5")
            
            # Show role assignments in compact format
            roles_str = " | ".join([
                f"{role[:3].upper()}:{assignment.assignments.get(role, 'None')[:8]}"
                for role in ["top", "jungle", "middle", "support", "bottom"]
            ])
            print(f"         │ {roles_str}")
            print()
    
    def _analyze_optimization_players(self, result) -> None:
        """Analyze the players involved in the optimization."""
        if not result or not result.best_assignment:
            print("\n❌ No optimization results to analyze")
            return
        
        print("\n" + "=" * 60)
        print("👥 PLAYER ANALYSIS FROM OPTIMIZATION")
        print("=" * 60)
        
        # Get all players involved in the optimization
        all_players_in_result = set()
        for assignment in result.assignments:
            all_players_in_result.update(assignment.assignments.values())
        
        # Load full player data
        all_players = self.engine.data_manager.load_player_data()
        involved_players = [p for p in all_players if p.name in all_players_in_result]
        
        for player in involved_players:
            print(f"\n👤 {player.name} ({player.summoner_name})")
            print("-" * 30)
            
            # Show role performance in this optimization
            best_assignment = result.best_assignment
            assigned_role = None
            for role, assigned_player in best_assignment.assignments.items():
                if assigned_player == player.name:
                    assigned_role = role
                    break
            
            if assigned_role:
                individual_score = best_assignment.individual_scores.get(player.name, 0)
                role_pref = player.role_preferences.get(assigned_role, 3)
                print(f"  🎯 Assigned Role: {assigned_role.upper()} (Score: {individual_score:.2f}, Preference: {role_pref}/5)")
            else:
                print("  ⚠️ Not assigned in best composition")
            
            # Show role preferences
            print("  ⭐ Role Preferences:")
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                pref = player.role_preferences.get(role, 3)
                pref_indicator = "🔥" if pref == 5 else "👍" if pref == 4 else "😐" if pref == 3 else "👎" if pref == 2 else "❌"
                print(f"    {role.capitalize():8}: {pref}/5 {pref_indicator}")
            
            # Show data completeness
            data_indicators = []
            if player.performance_cache:
                data_indicators.append("📊 Performance")
            if player.champion_masteries:
                data_indicators.append("🏆 Champions")
            if any(pref != 3 for pref in player.role_preferences.values()):
                data_indicators.append("⭐ Custom Prefs")
            
            print(f"  📋 Data: {' | '.join(data_indicators) if data_indicators else 'Basic only'}")
            
            # Show top champions if available
            if player.champion_masteries:
                top_champions = sorted(player.champion_masteries.values(), 
                                     key=lambda x: x.mastery_points, reverse=True)[:3]
                print("  🏆 Top Champions:")
                for champ in top_champions:
                    print(f"    {champ.champion_name} (L{champ.mastery_level}, {champ.mastery_points:,} pts)")
        
        print(f"\n📊 Summary: Analyzed {len(involved_players)} players from optimization results")
    
    def _export_results(self, result) -> None:
        """Export optimization results to a file."""
        if not result:
            print("\n❌ No results to export")
            return
        
        try:
            from datetime import datetime
            import json
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"optimization_results_{timestamp}.json"
            
            # Prepare export data
            export_data = {
                'timestamp': datetime.now().isoformat(),
                'optimization_time': getattr(result, 'optimization_time', 0),
                'total_alternatives': len(result.assignments),
                'best_composition': {
                    'total_score': result.best_assignment.total_score,
                    'assignments': result.best_assignment.assignments,
                    'individual_scores': result.best_assignment.individual_scores,
                    'explanation': getattr(result.best_assignment, 'explanation', '')
                },
                'all_alternatives': []
            }
            
            # Add all alternatives
            for i, assignment in enumerate(result.assignments):
                alt_data = {
                    'rank': i + 1,
                    'total_score': assignment.total_score,
                    'assignments': assignment.assignments,
                    'individual_scores': assignment.individual_scores
                }
                export_data['all_alternatives'].append(alt_data)
            
            # Write to file
            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)
            
            print(f"\n✅ Results exported to {filename}")
            print(f"📊 Exported {len(result.assignments)} team compositions")
            
        except Exception as e:
            print(f"\n❌ Export failed: {e}")
            self.logger.error(f"Failed to export results: {e}")
    
    def _display_team_strength_summary(self, assignment) -> None:
        """Display a summary of team strengths and potential concerns."""
        print("💪 Team Strength Analysis:")
        print("-" * 25)
        
        # Analyze individual scores
        scores = list(assignment.individual_scores.values())
        if scores:
            avg_score = sum(scores) / len(scores)
            min_score = min(scores)
            max_score = max(scores)
            
            # Strength indicators
            strengths = []
            concerns = []
            
            if avg_score >= 4:
                strengths.append("High overall team performance")
            elif avg_score >= 3:
                strengths.append("Solid team performance")
            else:
                concerns.append("Below average team performance")
            
            if max_score - min_score <= 1:
                strengths.append("Well-balanced role assignments")
            elif max_score - min_score >= 3:
                concerns.append("Significant performance gaps between roles")
            
            # Check for very low scores
            weak_roles = [role for role, player in assignment.assignments.items() 
                         if assignment.individual_scores.get(player, 0) < 2]
            if weak_roles:
                concerns.append(f"Weak performance in: {', '.join(weak_roles)}")
            
            # Display strengths
            if strengths:
                for strength in strengths:
                    print(f"  ✅ {strength}")
            
            # Display concerns
            if concerns:
                for concern in concerns:
                    print(f"  ⚠️ {concern}")
            
            # Show score distribution
            print(f"  📈 Score Range: {min_score:.2f} - {max_score:.2f} (avg: {avg_score:.2f})")
            
            # Show role-specific insights
            role_insights = []
            for role, player in assignment.assignments.items():
                score = assignment.individual_scores.get(player, 0)
                if score >= 4.5:
                    role_insights.append(f"{role}: Excellent fit")
                elif score < 2:
                    role_insights.append(f"{role}: Needs attention")
            
            if role_insights:
                print("  🎯 Role Insights:")
                for insight in role_insights[:3]:  # Show top 3 insights
                    print(f"    • {insight}")
            
            if not strengths and not concerns:
                print("  📊 Standard team composition")
        
        print()
    
    def _show_alternative_compositions(self, result) -> None:
        """Show alternative team compositions."""
        if len(result.assignments) <= 1:
            print("\nNo alternative compositions available.")
            return
        
        print("\n" + "=" * 50)
        print("🔄 ALTERNATIVE COMPOSITIONS")
        print("=" * 50)
        
        for i, assignment in enumerate(result.assignments[1:4], 2):  # Show top 3 alternatives
            print(f"\nOption {i} (Score: {assignment.total_score:.2f}):")
            print("-" * 30)
            
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                player_name = assignment.assignments.get(role, "Unassigned")
                score = assignment.individual_scores.get(player_name, 0)
                print(f"{role.capitalize():8}: {player_name} ({score:.2f})")
    
    def _analyze_optimization_players(self, result) -> None:
        """Analyze players from the optimization result."""
        if not result.best_assignment.assignments:
            print("\nNo players to analyze.")
            return
        
        player_names = list(result.best_assignment.assignments.values())
        analysis = self.engine.get_comprehensive_analysis(player_names)
        
        self._display_analysis_results(analysis)
    
    def _export_results(self, result) -> None:
        """Export optimization results."""
        try:
            import json
            from datetime import datetime
            
            # Create export data
            export_data = {
                'timestamp': datetime.now().isoformat(),
                'optimization_results': {
                    'total_score': result.best_assignment.total_score,
                    'optimization_time': result.optimization_time,
                    'assignments': result.best_assignment.assignments,
                    'individual_scores': result.best_assignment.individual_scores,
                    'synergy_scores': {f"{k[0]}+{k[1]}": v for k, v in result.best_assignment.synergy_scores.items()},
                }
            }
            
            # Save to file
            filename = f"optimization_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)
            
            print(f"\n✅ Results exported to {filename}")
            
        except Exception as e:
            print(f"\n❌ Export failed: {e}")
    
    def _manage_players(self) -> None:
        """Consolidated player management interface with enhanced overview."""
        while True:
            print("\n" + "=" * 60)
            print("👥 PLAYER MANAGEMENT - Consolidated Interface")
            print("=" * 60)
            
            # Enhanced player overview
            players = self.engine.data_manager.load_player_data()
            status = self.engine.system_status
            
            print(f"📊 Overview: {len(players)} players registered")
            
            if players:
                # Data quality summary
                players_with_data = status.get('players_with_api_data', 0)
                players_with_prefs = status.get('players_with_preferences', 0)
                
                print(f"   Data Quality: {players_with_data}/{len(players)} with API data, {players_with_prefs}/{len(players)} with preferences")
                print(f"   Optimization Ready: {'✅ Yes' if status.get('ready_for_optimization') else '❌ No'}")
                
                # Quick player list with enhanced status
                print("\n📋 Player List:")
                for i, player in enumerate(players[:8], 1):  # Show first 8
                    # Status indicators
                    indicators = []
                    if player.performance_cache:
                        indicators.append("📊")
                    if player.champion_masteries:
                        indicators.append("🏆")
                    if any(pref != 3 for pref in player.role_preferences.values()):
                        indicators.append("⭐")
                    
                    status_str = "".join(indicators) if indicators else "📝"
                    
                    # Show last updated
                    last_updated = player.last_updated.strftime('%m/%d') if player.last_updated else "Never"
                    
                    print(f"  {i:2}. {player.name:15} {status_str} | Updated: {last_updated}")
                
                if len(players) > 8:
                    print(f"     ... and {len(players) - 8} more players")
                
                print("\nLegend: 📊 Performance | 🏆 Champions | ⭐ Preferences | 📝 Basic")
            else:
                print("   No players found. Add some players to get started!")
            
            print("\n🛠️ Management Options:")
            print("1. ➕ Add new player")
            print("2. 👤 View/edit player details")
            print("3. 🗑️ Remove player")
            print("4. 🔄 Bulk operations")
            print("5. 📊 Player comparison")
            print("6. 📜 Historical match extraction")
            print("7. 📈 View extraction status")
            print("8. 🏠 Back to main menu")
            
            # Smart suggestions
            if len(players) == 0:
                print("\n💡 Tip: Start by adding players with option 1")
            elif len(players) < 5:
                print(f"\n💡 Tip: Add {5 - len(players)} more players for full team optimization")
            elif players_with_data < len(players) // 2:
                print("\n💡 Tip: Use bulk operations to refresh player data")
            
            choice = input("\nEnter your choice (1-8): ").strip()
            
            if choice == "1":
                self._add_single_player()
            elif choice == "2":
                self._view_edit_player()
            elif choice == "3":
                self._remove_player()
            elif choice == "4":
                self._bulk_operations()
            elif choice == "5":
                self._player_comparison()
            elif choice == "6":
                self._historical_match_extraction()
            elif choice == "7":
                self._view_extraction_status()
            elif choice == "8":
                break
            else:
                print("Invalid choice. Please enter a number between 1-8.")
    
    def _add_single_player(self) -> None:
        """Add a single player with comprehensive setup and validation."""
        print("\n" + "=" * 40)
        print("➕ ADD NEW PLAYER - Comprehensive Setup")
        print("=" * 40)
        
        # Step 1: Basic Information
        print("📝 Step 1: Basic Information")
        print("-" * 25)
        
        name = input("Player display name: ").strip()
        if not name:
            print("❌ Player name is required.")
            return
        
        # Check for duplicate names
        existing_player = self.engine.data_manager.get_player_by_name(name)
        if existing_player:
            print(f"❌ Player '{name}' already exists.")
            return
        
        riot_id = input("Riot ID (gameName#tagLine): ").strip()
        if not riot_id:
            print("❌ Riot ID is required.")
            return
        
        if '#' not in riot_id:
            print("❌ Invalid format. Use: gameName#tagLine")
            return
        
        # Step 2: Data Fetching Options
        print(f"\n📊 Step 2: Data Options")
        print("-" * 20)
        
        auto_fetch = True
        if self.engine.api_available:
            print("API is available. Data fetching options:")
            print("1. Fetch all data (recommended)")
            print("2. Fetch basic data only")
            print("3. Skip data fetching")
            
            fetch_choice = input("Choose option (1-3) [1]: ").strip()
            if fetch_choice == "3":
                auto_fetch = False
            elif fetch_choice == "2":
                auto_fetch = "basic"
        else:
            print("⚠️ API not available - will add player without data fetching")
            auto_fetch = False
        
        # Step 3: Add Player with Smart Defaults
        print(f"\n🔄 Step 3: Adding Player with Smart Defaults")
        print("-" * 40)
        print(f"Adding '{name}' with Riot ID '{riot_id}'...")
        
        if self.engine.api_available:
            print("🔍 Fetching API data and calculating intelligent preferences...")
        else:
            print("⚠️ API unavailable - using intelligent default preferences...")
        
        success, message, player = self.engine.add_player_with_data(name, riot_id, auto_fetch)
        
        if not success:
            print(f"❌ {message}")
            
            # Offer troubleshooting
            if "not found" in message.lower():
                print("\n🔧 Troubleshooting:")
                print("• Check the spelling of the Riot ID")
                print("• Ensure the tag is correct (visible in League client)")
                print("• Make sure the account has played League of Legends")
                
                retry = input("\nTry with a different Riot ID? (y/n): ").strip().lower()
                if retry == 'y':
                    self._add_single_player()
            return
        
        print(f"✅ {message}")
        
        # Step 4: Smart Setup and Analysis
        if player:
            print(f"\n🧠 Step 4: Smart Analysis Results")
            print("-" * 32)
            
            # Show data quality assessment
            freshness_info = self.engine.get_player_data_freshness(player.name)
            quality_score = freshness_info.get('data_quality_score', 0)
            
            print(f"📊 Data Quality Score: {quality_score}/100")
            
            if quality_score >= 80:
                print("🌟 Excellent data quality - ready for optimization!")
            elif quality_score >= 60:
                print("✅ Good data quality - optimization ready")
            elif quality_score >= 40:
                print("⚠️ Fair data quality - consider refreshing data")
            else:
                print("❌ Limited data quality - manual preferences recommended")
            
            # Show intelligent preferences if calculated
            has_intelligent_prefs = any(pref != 3 for pref in player.role_preferences.values())
            if has_intelligent_prefs and (player.performance_cache or player.champion_masteries):
                print("\n🎯 Intelligent Role Preferences (calculated from data):")
                role_names = {"top": "Top", "jungle": "Jungle", "middle": "Mid", "support": "Support", "bottom": "ADC"}
                for role in ["top", "jungle", "middle", "support", "bottom"]:
                    pref = player.role_preferences.get(role, 3)
                    stars = "⭐" * pref
                    print(f"   {role_names[role]:8}: {stars} ({pref}/5)")
            else:
                print("\n🎯 Default Role Preferences (no performance data available):")
                role_names = {"top": "Top", "jungle": "Jungle", "middle": "Mid", "support": "Support", "bottom": "ADC"}
                for role in ["top", "jungle", "middle", "support", "bottom"]:
                    pref = player.role_preferences.get(role, 3)
                    stars = "⭐" * pref
                    print(f"   {role_names[role]:8}: {stars} ({pref}/5)")
            
            # Show player summary with enhanced information
            print(f"\n📋 Player Summary:")
            print(f"   Name: {player.name}")
            print(f"   Riot ID: {player.summoner_name}")
            print(f"   Performance Data: {'✅' if player.performance_cache else '❌'}")
            print(f"   Champion Data: {'✅' if player.champion_masteries else '❌'}")
            print(f"   Intelligent Preferences: {'✅' if has_intelligent_prefs else '❌'}")
            
            # Show recommendations
            recommendations = freshness_info.get('recommendations', [])
            if recommendations:
                print(f"\n💡 Recommendations:")
                for rec in recommendations:
                    print(f"   • {rec}")
            
            # Update system status
            self.engine.system_status = self.engine._get_system_status()
            
            # Step 5: Optional Customization
            print(f"\n⚙️ Step 5: Optional Customization")
            print("-" * 30)
            
            # Offer to customize preferences if they want
            if has_intelligent_prefs:
                customize = input("Customize the calculated preferences? (y/N): ").strip().lower()
            else:
                customize = input("Set custom role preferences? (Y/n): ").strip().lower()
                if customize != 'n':
                    customize = 'y'
            
            if customize == 'y':
                self._set_player_preferences(player)
            
            # Offer to refresh data if quality is low
            if quality_score < 60 and self.engine.api_available and player.puuid:
                refresh = input("Try to fetch more recent data? (Y/n): ").strip().lower()
                if refresh != 'n':
                    print("🔄 Refreshing player data...")
                    refresh_success, refresh_msg = self.engine.refresh_player_data(player.name, force_api_fetch=True)
                    if refresh_success:
                        print(f"✅ {refresh_msg}")
                        # Recalculate freshness info
                        freshness_info = self.engine.get_player_data_freshness(player.name)
                        new_quality = freshness_info.get('data_quality_score', 0)
                        print(f"📊 Updated Data Quality Score: {new_quality}/100")
                    else:
                        print(f"❌ {refresh_msg}")
            
            # Offer next actions
            print(f"\n🎯 Next Actions:")
            if self.engine.system_status.get('player_count', 0) >= 5:
                print("• You now have enough players for optimization!")
                print("• Try Quick Optimize from the main menu")
            else:
                needed = 5 - self.engine.system_status.get('player_count', 0)
                print(f"• Add {needed} more players for full team optimization")
            
            view_analysis = input("\nView detailed player analysis? (y/n): ").strip().lower()
            if view_analysis == 'y':
                self._view_player_analysis(player)
    
    def _view_edit_player(self) -> None:
        """View and edit player details."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("\nNo players found.")
            return
        
        # Select player
        print("\nSelect player:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name} ({player.summoner_name})")
        
        try:
            choice = int(input(f"\nEnter player number (1-{len(players)}): ")) - 1
            if 0 <= choice < len(players):
                player = players[choice]
                self._player_detail_menu(player)
            else:
                print("Invalid selection.")
        except ValueError:
            print("Please enter a valid number.")
    
    def _player_detail_menu(self, player) -> None:
        """Detailed player management menu."""
        while True:
            print(f"\n" + "=" * 40)
            print(f"👤 {player.name.upper()}")
            print("=" * 40)
            
            print(f"Summoner: {player.summoner_name}")
            print(f"Last Updated: {player.last_updated.strftime('%Y-%m-%d %H:%M') if player.last_updated else 'Never'}")
            print(f"Performance Data: {'✅' if player.performance_cache else '❌'}")
            print(f"Champion Data: {'✅' if player.champion_masteries else '❌'}")
            
            # Show preferences
            print("\nRole Preferences:")
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                pref = player.role_preferences.get(role, 3)
                stars = "★" * pref + "☆" * (5 - pref)
                print(f"  {role.capitalize():8}: {stars} ({pref}/5)")
            
            print("\nOptions:")
            print("1. Update role preferences")
            print("2. Refresh API data")
            print("3. View detailed analysis")
            print("4. Update Riot ID")
            print("5. Back to player list")
            
            choice = input("\nEnter your choice (1-5): ").strip()
            
            if choice == "1":
                self._set_player_preferences(player)
            elif choice == "2":
                self._refresh_player_data(player)
            elif choice == "3":
                self._view_player_analysis(player)
            elif choice == "4":
                self._update_riot_id(player)
            elif choice == "5":
                break
            else:
                print("Invalid choice.")
    
    def _set_player_preferences(self, player) -> None:
        """Set role preferences for a player."""
        print(f"\n⭐ Set Role Preferences for {player.name}")
        print("-" * 40)
        print("Rate each role from 1-5 (1=hate, 3=neutral, 5=love)")
        print("Press Enter to keep current value.")
        
        new_preferences = {}
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            current = player.role_preferences.get(role, 3)
            
            while True:
                pref_input = input(f"{role.capitalize()} [{current}]: ").strip()
                
                if not pref_input:
                    new_preferences[role] = current
                    break
                
                try:
                    pref_value = int(pref_input)
                    if 1 <= pref_value <= 5:
                        new_preferences[role] = pref_value
                        break
                    else:
                        print("Please enter a number between 1-5.")
                except ValueError:
                    print("Please enter a valid number.")
        
        # Save preferences
        try:
            self.engine.data_manager.update_preferences(player.name, new_preferences)
            player.role_preferences = new_preferences
            print("✅ Preferences updated successfully!")
        except Exception as e:
            print(f"❌ Failed to update preferences: {e}")
    
    def _refresh_player_data(self, player) -> None:
        """Refresh API data for a player."""
        if not self.engine.api_available:
            print("❌ API not available - cannot refresh data.")
            return
        
        print(f"🔄 Refreshing data for {player.name}...")
        success = self.engine._fetch_player_api_data(player)
        
        if success:
            print("✅ Data refreshed successfully!")
        else:
            print("⚠️ Some data could not be fetched.")
    
    def _view_player_analysis(self, player) -> None:
        """View detailed analysis for a single player."""
        analysis = self.engine.get_comprehensive_analysis([player.name])
        
        if 'error' in analysis:
            print(f"❌ Analysis failed: {analysis['error']}")
            return
        
        if not analysis['players']:
            print("❌ No analysis data available.")
            return
        
        player_analysis = analysis['players'][0]
        
        print(f"\n" + "=" * 50)
        print(f"📊 DETAILED ANALYSIS: {player.name.upper()}")
        print("=" * 50)
        
        # Data completeness
        completeness = player_analysis['data_completeness']
        print(f"Data Completeness: {completeness['completeness_score']:.1%}")
        print(f"  Preferences: {'✅' if completeness['has_preferences'] else '❌'}")
        print(f"  Performance: {'✅' if completeness['has_performance_data'] else '❌'}")
        print(f"  Champions: {'✅' if completeness['has_champion_data'] else '❌'}")
        
        # Role analysis
        print("\nRole Suitability:")
        for role, data in player_analysis['role_analysis'].items():
            print(f"  {role.capitalize():8}: {data['suitability']:10} (Pref: {data['preference']}/5)")
            if data.get('performance_score', 0) > 0:
                print(f"             Performance: {data['performance_score']:.1f}% WR")
            if data.get('champion_pool_size', 0) > 0:
                print(f"             Champions: {data['champion_pool_size']} available")
        
        # Strengths and weaknesses
        if player_analysis['strengths']:
            print("\nStrengths:")
            for strength in player_analysis['strengths']:
                print(f"  ✅ {strength}")
        
        if player_analysis['weaknesses']:
            print("\nWeaknesses:")
            for weakness in player_analysis['weaknesses']:
                print(f"  ⚠️ {weakness}")
    
    def _update_riot_id(self, player) -> None:
        """Update a player's Riot ID."""
        print(f"\nCurrent Riot ID: {player.summoner_name}")
        new_riot_id = input("Enter new Riot ID (gameName#tagLine): ").strip()
        
        if not new_riot_id or '#' not in new_riot_id:
            print("Invalid Riot ID format.")
            return
        
        # Validate if API is available
        if self.engine.api_available:
            print("🔄 Validating new Riot ID...")
            try:
                summoner_data = self.engine.riot_client.get_summoner_data(new_riot_id)
                if summoner_data:
                    player.summoner_name = new_riot_id
                    player.puuid = summoner_data.get('puuid', '')
                    # Clear cached data since it's for a different account
                    player.performance_cache = {}
                    player.champion_masteries = {}
                    player.last_updated = None
                    
                    self.engine.data_manager.update_player(player)
                    print("✅ Riot ID updated successfully!")
                    
                    # Offer to fetch new data
                    fetch_data = input("Fetch data for new account? (y/n): ").strip().lower()
                    if fetch_data == 'y':
                        self._refresh_player_data(player)
                else:
                    print("❌ Riot ID not found.")
            except Exception as e:
                print(f"❌ Validation failed: {e}")
        else:
            # Update without validation
            player.summoner_name = new_riot_id
            player.puuid = ''
            self.engine.data_manager.update_player(player)
            print("✅ Riot ID updated (not validated - API offline).")
    
    def _player_comparison(self) -> None:
        """Compare multiple players side by side."""
        players = self.engine.data_manager.load_player_data()
        
        if len(players) < 2:
            print("\nNeed at least 2 players for comparison.")
            return
        
        print("\n" + "=" * 50)
        print("👥 PLAYER COMPARISON")
        print("=" * 50)
        
        print("Select players to compare (2-5 players recommended):")
        for i, player in enumerate(players, 1):
            # Status indicators
            indicators = []
            if player.performance_cache:
                indicators.append("📊")
            if player.champion_masteries:
                indicators.append("🏆")
            if any(pref != 3 for pref in player.role_preferences.values()):
                indicators.append("⭐")
            
            status_str = "".join(indicators) if indicators else "📝"
            print(f"{i:2}. {player.name:15} {status_str}")
        
        print("\nLegend: 📊 Performance | 🏆 Champions | ⭐ Preferences | 📝 Basic")
        
        try:
            selection = input("\nEnter player numbers (e.g., 1,3,5): ").strip()
            indices = [int(x.strip()) - 1 for x in selection.split(',')]
            
            selected_players = []
            for idx in indices:
                if 0 <= idx < len(players):
                    selected_players.append(players[idx].name)
                else:
                    print(f"Invalid player number: {idx + 1}")
                    return
            
            if len(selected_players) < 2:
                print("Please select at least 2 players.")
                return
            
            # Get analysis for selected players
            analysis = self.engine.get_comprehensive_analysis(selected_players)
            
            if 'error' in analysis:
                print(f"❌ Analysis failed: {analysis['error']}")
                return
            
            # Display comparison
            self._display_player_comparison(analysis['players'])
            
        except ValueError:
            print("Invalid input format. Please use numbers separated by commas.")
    
    def _display_player_comparison(self, players_analysis) -> None:
        """Display side-by-side player comparison with enhanced metrics."""
        print("\n" + "=" * 80)
        print("👥 DETAILED PLAYER COMPARISON")
        print("=" * 80)
        
        # Header
        print(f"{'Metric':<25}", end="")
        for player in players_analysis:
            print(f"{player['name'][:12]:>15}", end="")
        print()
        print("-" * 80)
        
        # Data completeness
        print(f"{'Data Completeness':<25}", end="")
        for player in players_analysis:
            completeness = player['data_completeness']['completeness_score']
            print(f"{completeness:>14.0%}", end="")
        print()
        
        # Role preferences comparison
        roles = ["top", "jungle", "middle", "support", "bottom"]
        for role in roles:
            print(f"{role.capitalize() + ' Preference':<25}", end="")
            for player in players_analysis:
                pref = player['role_analysis'][role]['preference']
                suitability = player['role_analysis'][role]['suitability'][:4]  # First 4 chars
                print(f"{pref}/5 ({suitability}):>10", end="")
            print()
        
        # Champion data
        print(f"{'Total Champions':<25}", end="")
        for player in players_analysis:
            champ_data = player.get('champion_analysis', {})
            total = champ_data.get('total_champions', 0)
            print(f"{total:>15}", end="")
        print()
        
        print(f"{'Mastery 7 Champions':<25}", end="")
        for player in players_analysis:
            champ_data = player.get('champion_analysis', {})
            m7 = champ_data.get('mastery_7_champions', 0)
            print(f"{m7:>15}", end="")
        print()
        
        # Strengths summary
        print(f"\n{'Player Strengths:':<25}")
        max_strengths = max(len(p.get('strengths', [])) for p in players_analysis)
        for i in range(max_strengths):
            print(f"{'':25}", end="")
            for player in players_analysis:
                strengths = player.get('strengths', [])
                strength = strengths[i] if i < len(strengths) else ""
                print(f"{strength[:14]:>15}", end="")
            print()
        
        print("\n" + "=" * 80)
    
    def _remove_player(self) -> None:
        """Remove a player from the system."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("\nNo players found.")
            return
        
        print("\nSelect player to remove:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name} ({player.summoner_name})")
        
        try:
            choice = int(input(f"\nEnter player number (1-{len(players)}) or 0 to cancel: ")) - 1
            
            if choice == -1:  # User entered 0
                return
            
            if 0 <= choice < len(players):
                player = players[choice]
                confirm = input(f"Are you sure you want to remove '{player.name}'? (y/n): ").strip().lower()
                
                if confirm == 'y':
                    success = self.engine.data_manager.delete_player(player.name)
                    if success:
                        print(f"✅ Player '{player.name}' removed successfully!")
                        # Update system status
                        self.engine.system_status = self.engine._get_system_status()
                    else:
                        print(f"❌ Failed to remove player '{player.name}'.")
            else:
                print("Invalid selection.")
                
        except ValueError:
            print("Please enter a valid number.")
    
    def _bulk_operations(self) -> None:
        """Perform bulk operations on players with enhanced options."""
        while True:
            print("\n" + "=" * 50)
            print("🔄 BULK OPERATIONS - Batch Player Management")
            print("=" * 50)
            
            # Show current status
            players = self.engine.data_manager.load_player_data()
            status = self.engine.system_status
            
            print(f"📊 Current Status: {len(players)} players")
            if players:
                players_with_data = status.get('players_with_api_data', 0)
                players_with_prefs = status.get('players_with_preferences', 0)
                print(f"   {players_with_data}/{len(players)} with API data")
                print(f"   {players_with_prefs}/{len(players)} with custom preferences")
            
            print("\n🛠️ Bulk Operations:")
            print("1. ➕ Add multiple players from list")
            print("2. 🔄 Refresh all player data")
            print("3. ⭐ Set default preferences for all")
            print("4. 🧹 Clean up incomplete players")
            print("5. 💾 Export player data")
            print("6. 📥 Import player data")
            print("7. 📊 Generate player report")
            print("8. 🏠 Back to player management")
            
            # Smart recommendations
            if len(players) > 0:
                if players_with_data < len(players) // 2:
                    print("\n💡 Recommendation: Refresh player data (option 2)")
                elif players_with_prefs < len(players) // 2:
                    print("\n💡 Recommendation: Set preferences for better optimization (option 3)")
            
            choice = input("\nEnter your choice (1-8): ").strip()
            
            if choice == "1":
                self._bulk_add_players()
            elif choice == "2":
                self._bulk_refresh_data()
            elif choice == "3":
                self._bulk_set_preferences()
            elif choice == "4":
                self._bulk_cleanup_players()
            elif choice == "5":
                self._export_player_data()
            elif choice == "6":
                self._import_player_data()
            elif choice == "7":
                self._generate_player_report()
            elif choice == "8":
                break
            else:
                print("Invalid choice. Please enter a number between 1-8.")
    
    def _bulk_add_players(self) -> None:
        """Add multiple players from input."""
        print("\n📝 Bulk Add Players")
        print("-" * 20)
        print("Enter players one per line in format: PlayerName,gameName#tagLine")
        print("Press Enter on empty line to finish.")
        
        operations = []
        while True:
            line = input("Player: ").strip()
            if not line:
                break
            
            try:
                name, riot_id = line.split(',', 1)
                operations.append({
                    'action': 'add',
                    'name': name.strip(),
                    'riot_id': riot_id.strip(),
                    'auto_fetch': True
                })
            except ValueError:
                print("Invalid format. Use: PlayerName,gameName#tagLine")
        
        if operations:
            print(f"\n🔄 Adding {len(operations)} players...")
            results = self.engine.bulk_player_operations(operations)
            
            success_count = sum(1 for success, _ in results if success)
            print(f"\n✅ Successfully added {success_count}/{len(operations)} players")
            
            # Show any failures
            for i, (success, message) in enumerate(results):
                if not success:
                    print(f"❌ Player {i+1}: {message}")
    
    def _bulk_refresh_data(self) -> None:
        """Refresh API data for all players with smart caching."""
        if not self.engine.api_available:
            print("❌ API not available - cannot refresh data.")
            return
        
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players found.")
            return
        
        # Show data freshness overview
        print(f"\n📊 Data Freshness Overview:")
        stale_players = []
        fresh_players = []
        
        for player in players:
            freshness_info = self.engine.get_player_data_freshness(player.name)
            age_days = freshness_info.get('data_age_days')
            quality_score = freshness_info.get('data_quality_score', 0)
            
            if age_days is None or age_days > 7 or quality_score < 60:
                stale_players.append(player.name)
            else:
                fresh_players.append(player.name)
        
        print(f"   🔴 Stale data: {len(stale_players)} players")
        print(f"   🟢 Fresh data: {len(fresh_players)} players")
        
        if not stale_players:
            print("\n✅ All player data is fresh! No refresh needed.")
            return
        
        print(f"\n🔄 Refresh Options:")
        print(f"1. Refresh only stale data ({len(stale_players)} players)")
        print(f"2. Force refresh all players ({len(players)} players)")
        print(f"3. Select specific players")
        print(f"4. Cancel")
        
        choice = input("\nEnter your choice (1-4): ").strip()
        
        players_to_refresh = []
        if choice == "1":
            players_to_refresh = stale_players
        elif choice == "2":
            players_to_refresh = [p.name for p in players]
        elif choice == "3":
            print("\nSelect players to refresh:")
            for i, player in enumerate(players, 1):
                freshness_info = self.engine.get_player_data_freshness(player.name)
                age_days = freshness_info.get('data_age_days')
                quality_score = freshness_info.get('data_quality_score', 0)
                status = "🔴" if player.name in stale_players else "🟢"
                age_str = f"{age_days}d" if age_days else "Never"
                print(f"  {i:2}. {status} {player.name:15} | Age: {age_str:6} | Quality: {quality_score}/100")
            
            selection = input("\nEnter player numbers (comma-separated): ").strip()
            try:
                indices = [int(x.strip()) - 1 for x in selection.split(',')]
                players_to_refresh = [players[i].name for i in indices if 0 <= i < len(players)]
            except (ValueError, IndexError):
                print("Invalid selection.")
                return
        elif choice == "4":
            return
        else:
            print("Invalid choice.")
            return
        
        if not players_to_refresh:
            print("No players selected.")
            return
        
        # Confirm refresh
        force_refresh = choice == "2"
        refresh_type = "force refresh" if force_refresh else "smart refresh"
        confirm = input(f"\n{refresh_type.title()} data for {len(players_to_refresh)} players? (y/n): ").strip().lower()
        if confirm != 'y':
            return
        
        print(f"\n🔄 Performing {refresh_type} for {len(players_to_refresh)} players...")
        print("This may take a moment due to API rate limiting...")
        
        # Use the new bulk refresh method
        results = self.engine.bulk_refresh_player_data(players_to_refresh, max_concurrent=3)
        
        # Process results
        success_count = 0
        failed_players = []
        
        for player_name, (success, message) in results.items():
            if player_name == "error":
                print(f"❌ {message}")
                return
            
            if success:
                success_count += 1
                print(f"✅ {player_name}: {message}")
            else:
                failed_players.append(player_name)
                print(f"❌ {player_name}: {message}")
        
        print(f"\n📊 Refresh Summary:")
        print(f"   ✅ Successful: {success_count}/{len(players_to_refresh)}")
        if failed_players:
            print(f"   ❌ Failed: {len(failed_players)}")
            print(f"   Failed players: {', '.join(failed_players)}")
        
        # Show updated data quality
        if success_count > 0:
            print(f"\n🔄 Recalculating intelligent preferences for updated players...")
            updated_prefs_count = 0
            for player_name in players_to_refresh:
                if (player_name, True) in [(k, v[0]) for k, v in results.items()]:
                    player = self.engine.data_manager.get_player_by_name(player_name)
                    if player and (player.performance_cache or player.champion_masteries):
                        updated_prefs_count += 1
            
            if updated_prefs_count > 0:
                print(f"✅ Updated intelligent preferences for {updated_prefs_count} players")
        
        # Update system status
        self.engine.system_status = self.engine._get_system_status()
    
    def _bulk_set_preferences(self) -> None:
        """Set default preferences for all players."""
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players found.")
            return
        
        print("\nSet default role preferences for all players:")
        print("This will overwrite existing preferences.")
        
        confirm = input("Continue? (y/n): ").strip().lower()
        if confirm != 'y':
            return
        
        # Get default preferences
        default_prefs = {}
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            while True:
                pref_input = input(f"{role.capitalize()} preference (1-5) [3]: ").strip()
                
                if not pref_input:
                    default_prefs[role] = 3
                    break
                
                try:
                    pref_value = int(pref_input)
                    if 1 <= pref_value <= 5:
                        default_prefs[role] = pref_value
                        break
                    else:
                        print("Please enter a number between 1-5.")
                except ValueError:
                    print("Please enter a valid number.")
        
        # Apply to all players
        operations = [
            {'action': 'update_preferences', 'name': p.name, 'preferences': default_prefs}
            for p in players
        ]
        
        results = self.engine.bulk_player_operations(operations)
        success_count = sum(1 for success, _ in results if success)
        print(f"✅ Updated preferences for {success_count}/{len(players)} players")
    
    def _export_player_data(self) -> None:
        """Export player data to file."""
        try:
            import json
            from datetime import datetime
            
            players = self.engine.data_manager.load_player_data()
            if not players:
                print("No players to export.")
                return
            
            # Convert to exportable format
            export_data = {
                'export_timestamp': datetime.now().isoformat(),
                'player_count': len(players),
                'players': [
                    {
                        'name': p.name,
                        'summoner_name': p.summoner_name,
                        'puuid': p.puuid,
                        'role_preferences': p.role_preferences,
                        'last_updated': p.last_updated.isoformat() if p.last_updated else None,
                        'has_performance_data': bool(p.performance_cache),
                        'has_champion_data': bool(p.champion_masteries)
                    }
                    for p in players
                ]
            }
            
            filename = f"player_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)
            
            print(f"✅ Player data exported to {filename}")
            
        except Exception as e:
            print(f"❌ Export failed: {e}")
    
    def _import_player_data(self) -> None:
        """Import player data from file."""
        filename = input("Enter filename to import: ").strip()
        if not filename:
            return
        
        try:
            import json
            
            with open(filename, 'r') as f:
                import_data = json.load(f)
            
            if 'players' not in import_data:
                print("❌ Invalid import file format.")
                return
            
            players_to_import = import_data['players']
            print(f"Found {len(players_to_import)} players to import.")
            
            confirm = input("Import these players? (y/n): ").strip().lower()
            if confirm != 'y':
                return
            
            # Create operations for import
            operations = []
            for player_data in players_to_import:
                operations.append({
                    'action': 'add',
                    'name': player_data['name'],
                    'riot_id': player_data['summoner_name'],
                    'auto_fetch': False  # Don't auto-fetch during import
                })
            
            results = self.engine.bulk_player_operations(operations)
            success_count = sum(1 for success, _ in results if success)
            
            print(f"✅ Successfully imported {success_count}/{len(operations)} players")
            
        except FileNotFoundError:
            print(f"❌ File '{filename}' not found.")
        except json.JSONDecodeError:
            print("❌ Invalid JSON file format.")
        except Exception as e:
            print(f"❌ Import failed: {e}")
    
    def _bulk_cleanup_players(self) -> None:
        """Clean up players with incomplete or outdated data."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("No players to clean up.")
            return
        
        print("\n🧹 Player Cleanup Analysis")
        print("-" * 30)
        
        # Identify players needing cleanup
        outdated_players = []
        incomplete_players = []
        
        from datetime import datetime, timedelta
        cutoff_date = datetime.now() - timedelta(days=30)
        
        for player in players:
            # Check for outdated data
            if not player.last_updated or player.last_updated < cutoff_date:
                outdated_players.append(player.name)
            
            # Check for incomplete data
            if not player.performance_cache and not player.champion_masteries:
                incomplete_players.append(player.name)
        
        print(f"Found {len(outdated_players)} players with outdated data (>30 days)")
        print(f"Found {len(incomplete_players)} players with no API data")
        
        if not outdated_players and not incomplete_players:
            print("✅ All players have recent and complete data!")
            return
        
        print("\nCleanup Options:")
        print("1. Refresh outdated player data")
        print("2. Remove players with no API data")
        print("3. Show detailed cleanup report")
        print("4. Cancel")
        
        choice = input("\nEnter your choice (1-4): ").strip()
        
        if choice == "1" and outdated_players:
            print(f"🔄 Refreshing data for {len(outdated_players)} players...")
            operations = [{'action': 'refresh_data', 'name': name} for name in outdated_players]
            results = self.engine.bulk_player_operations(operations)
            success_count = sum(1 for success, _ in results if success)
            print(f"✅ Refreshed {success_count}/{len(outdated_players)} players")
        
        elif choice == "2" and incomplete_players:
            print(f"Players to remove: {', '.join(incomplete_players)}")
            confirm = input(f"Remove {len(incomplete_players)} players with no data? (y/n): ").strip().lower()
            if confirm == 'y':
                operations = [{'action': 'remove', 'name': name} for name in incomplete_players]
                results = self.engine.bulk_player_operations(operations)
                success_count = sum(1 for success, _ in results if success)
                print(f"✅ Removed {success_count}/{len(incomplete_players)} players")
        
        elif choice == "3":
            print("\n📊 Detailed Cleanup Report:")
            print("-" * 30)
            if outdated_players:
                print(f"Outdated players: {', '.join(outdated_players)}")
            if incomplete_players:
                print(f"Incomplete players: {', '.join(incomplete_players)}")
    
    def _generate_player_report(self) -> None:
        """Generate a comprehensive player report."""
        print("\n📊 Generating Player Report...")
        
        analysis = self.engine.get_comprehensive_analysis()
        
        if 'error' in analysis:
            print(f"❌ Report generation failed: {analysis['error']}")
            return
        
        print("\n" + "=" * 60)
        print("📋 COMPREHENSIVE PLAYER REPORT")
        print("=" * 60)
        
        # System overview
        status = analysis.get('system_status', {})
        print(f"📊 System Overview:")
        print(f"   Total Players: {status.get('player_count', 0)}")
        print(f"   API Status: {'Available' if status.get('api_available') else 'Offline'}")
        print(f"   Optimization Ready: {'Yes' if status.get('ready_for_optimization') else 'No'}")
        
        # Player summary
        if 'players' in analysis:
            players_data = analysis['players']
            print(f"\n👥 Player Summary ({len(players_data)} players):")
            print("-" * 40)
            
            # Data quality metrics
            complete_data = sum(1 for p in players_data if p['data_completeness']['completeness_score'] >= 0.8)
            custom_prefs = sum(1 for p in players_data if p['data_completeness']['has_preferences'])
            
            print(f"   Complete Data: {complete_data}/{len(players_data)} players ({complete_data/len(players_data)*100:.0f}%)")
            print(f"   Custom Preferences: {custom_prefs}/{len(players_data)} players ({custom_prefs/len(players_data)*100:.0f}%)")
            
            # Role coverage
            print(f"\n🎯 Role Coverage Analysis:")
            roles = ["top", "jungle", "middle", "support", "bottom"]
            for role in roles:
                suitable_count = sum(1 for p in players_data 
                                   if p['role_analysis'][role]['suitability'] in ['Excellent', 'Good'])
                print(f"   {role.capitalize():8}: {suitable_count} suitable players")
            
            # Top performers
            print(f"\n🏆 Top Performers by Role:")
            for role in roles:
                role_players = [(p['name'], p['role_analysis'][role]['preference']) 
                              for p in players_data]
                role_players.sort(key=lambda x: x[1], reverse=True)
                if role_players:
                    top_player = role_players[0]
                    print(f"   {role.capitalize():8}: {top_player[0]} (Preference: {top_player[1]}/5)")
        
        # Team readiness
        if 'team_analysis' in analysis:
            team = analysis['team_analysis']
            print(f"\n🎯 Team Readiness: {team.get('optimization_readiness', 'Unknown')}")
        
        # Recommendations
        if analysis.get('recommendations'):
            print(f"\n💡 Recommendations:")
            for rec in analysis['recommendations']:
                print(f"   • {rec}")
        
        # Offer to export
        export_report = input("\nExport this report to file? (y/n): ").strip().lower()
        if export_report == 'y':
            try:
                import json
                from datetime import datetime
                
                filename = f"player_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                with open(filename, 'w') as f:
                    json.dump(analysis, f, indent=2, default=str)
                
                print(f"✅ Report exported to {filename}")
            except Exception as e:
                print(f"❌ Export failed: {e}")
    
    def _view_analysis(self) -> None:
        """Comprehensive analysis dashboard with consolidated views."""
        while True:
            print("\n" + "=" * 60)
            print("📊 ANALYSIS DASHBOARD - Comprehensive Insights")
            print("=" * 60)
            
            # Quick system overview
            status = self.engine.system_status
            players = self.engine.data_manager.load_player_data()
            
            print(f"📋 Quick Overview:")
            print(f"   Players: {len(players)} | API: {'✅' if status.get('api_available') else '❌'} | Ready: {'✅' if status.get('ready_for_optimization') else '❌'}")
            
            if players:
                players_with_data = status.get('players_with_api_data', 0)
                print(f"   Data Quality: {players_with_data}/{len(players)} with complete data")
            
            print("\n🔍 Analysis Options:")
            print("1. 🏠 Team Overview & Readiness Assessment")
            print("2. 👤 Individual Player Deep Dive")
            print("3. ⚖️ Player Comparison Matrix")
            print("4. 🎯 Role Coverage & Gap Analysis")
            print("5. 🏆 Champion Pool Analysis")
            print("6. 🤝 Team Synergy Analysis")
            print("7. 📈 Performance Trends & Insights")
            print("8. 🔮 Optimization Predictions")
            print("9. 🏅 Champion Performance Analytics")
            print("10. 🏛️ Team Composition Analysis")
            print("11. 📊 Interactive Analytics Dashboard")
            print("12. 🏠 Back to main menu")
            
            # Smart recommendations
            if len(players) == 0:
                print("\n💡 Tip: Add players first to see analysis options")
            elif len(players) < 5:
                print(f"\n💡 Tip: Add {5 - len(players)} more players for complete team analysis")
            elif players_with_data < len(players) // 2:
                print("\n💡 Tip: Refresh player data for more accurate analysis")
            
            choice = input("\nEnter your choice (1-11): ").strip()
            
            if choice == "1":
                self._team_overview_analysis()
            elif choice == "2":
                self._individual_analysis_menu()
            elif choice == "3":
                self._player_comparison_analysis()
            elif choice == "4":
                self._role_coverage_analysis()
            elif choice == "5":
                self._champion_pool_analysis()
            elif choice == "6":
                self._team_synergy_analysis()
            elif choice == "7":
                self._performance_trends_analysis()
            elif choice == "8":
                self._optimization_predictions()
            elif choice == "9":
                self._champion_performance_analytics()
            elif choice == "10":
                self._team_composition_analysis()
            elif choice == "11":
                self._interactive_analytics_dashboard()
            elif choice == "12":
                break
            else:
                print("Invalid choice. Please enter a number between 1-12.")
    
    def _team_overview_analysis(self) -> None:
        """Display comprehensive team overview with enhanced metrics."""
        print("\n🔄 Generating comprehensive team analysis...")
        
        analysis = self.engine.get_comprehensive_analysis()
        
        if 'error' in analysis:
            print(f"❌ Analysis failed: {analysis['error']}")
            return
        
        print("\n" + "=" * 70)
        print("🏠 TEAM OVERVIEW & READINESS ASSESSMENT")
        print("=" * 70)
        
        # Enhanced system status
        status = analysis.get('system_status', {})
        print(f"🔧 System Health:")
        print(f"   Total Players: {status.get('player_count', 0)}")
        print(f"   API Status: {'🟢 Online' if status.get('api_available') else '🔴 Offline'}")
        print(f"   Champions Loaded: {len(self.engine.champion_data_manager.champions)}")
        print(f"   Optimization Ready: {'🟢 Yes' if status.get('ready_for_optimization') else '🔴 No'}")
        
        # Data quality assessment
        if 'players' in analysis and analysis['players']:
            players_data = analysis['players']
            
            print(f"\n📊 Data Quality Assessment:")
            complete_data = sum(1 for p in players_data if p['data_completeness']['completeness_score'] >= 0.8)
            partial_data = sum(1 for p in players_data if 0.3 <= p['data_completeness']['completeness_score'] < 0.8)
            minimal_data = len(players_data) - complete_data - partial_data
            
            print(f"   🟢 Complete Data: {complete_data} players ({complete_data/len(players_data)*100:.0f}%)")
            print(f"   🟡 Partial Data: {partial_data} players ({partial_data/len(players_data)*100:.0f}%)")
            print(f"   🔴 Minimal Data: {minimal_data} players ({minimal_data/len(players_data)*100:.0f}%)")
            
            # Team composition readiness
            print(f"\n🎯 Team Composition Readiness:")
            if 'team_analysis' in analysis and analysis['team_analysis']:
                team = analysis['team_analysis']
                readiness = team.get('optimization_readiness', 'Unknown')
                
                if readiness == 'Ready':
                    print(f"   Status: 🟢 {readiness} - All roles have suitable players")
                elif readiness == 'Partial':
                    print(f"   Status: 🟡 {readiness} - Some roles may be suboptimal")
                else:
                    print(f"   Status: 🔴 {readiness} - Insufficient player coverage")
                
                # Role coverage breakdown
                print(f"\n🎭 Role Coverage Breakdown:")
                role_coverage = team.get('role_coverage', {})
                for role in ["top", "jungle", "middle", "support", "bottom"]:
                    if role in role_coverage:
                        coverage = role_coverage[role]
                        quality = coverage['coverage_quality']
                        suitable_count = len(coverage['suitable_players'])
                        
                        if quality == 'Good':
                            indicator = "🟢"
                        elif quality == 'Limited':
                            indicator = "🟡"
                        else:
                            indicator = "🔴"
                        
                        print(f"   {role.capitalize():8}: {indicator} {quality:8} ({suitable_count} suitable players)")
                        
                        # Show top candidates
                        if coverage['suitable_players']:
                            top_candidates = coverage['suitable_players'][:3]
                            print(f"            Top candidates: {', '.join(top_candidates)}")
            
            # Team strengths and weaknesses
            self._analyze_team_strengths_weaknesses(players_data)
        
        # Recommendations with priorities
        if analysis.get('recommendations'):
            print(f"\n💡 Priority Recommendations:")
            for i, rec in enumerate(analysis['recommendations'][:5], 1):
                priority = "🔴 High" if i <= 2 else "🟡 Medium"
                print(f"   {priority}: {rec}")
        
        # Quick actions
        print(f"\n⚡ Quick Actions:")
        print("   • Press 'o' to run Quick Optimize")
        print("   • Press 'p' to manage players")
        print("   • Press 'r' to refresh all data")
        
        action = input("\nTake quick action (o/p/r) or press Enter to continue: ").strip().lower()
        if action == 'o':
            self._quick_optimize()
        elif action == 'p':
            self._manage_players()
        elif action == 'r':
            self._bulk_refresh_all_data()
    
    def _analyze_team_strengths_weaknesses(self, players_data) -> None:
        """Analyze overall team strengths and weaknesses."""
        print(f"\n💪 Team Strengths & Weaknesses:")
        
        # Analyze role distribution
        role_preferences = {}
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            role_preferences[role] = []
            for player in players_data:
                pref = player['role_analysis'][role]['preference']
                if pref >= 4:
                    role_preferences[role].append((player['name'], pref))
        
        # Identify strengths
        strengths = []
        weaknesses = []
        
        # Role coverage strengths/weaknesses
        for role, players_list in role_preferences.items():
            if len(players_list) >= 3:
                strengths.append(f"Strong {role} player pool ({len(players_list)} candidates)")
            elif len(players_list) == 0:
                weaknesses.append(f"No dedicated {role} players")
            elif len(players_list) == 1:
                weaknesses.append(f"Limited {role} options (only {players_list[0][0]})")
        
        # Champion diversity
        total_champions = sum(p.get('champion_analysis', {}).get('total_champions', 0) for p in players_data)
        avg_champions = total_champions / len(players_data) if players_data else 0
        
        if avg_champions >= 30:
            strengths.append(f"High champion diversity (avg {avg_champions:.0f} per player)")
        elif avg_champions < 15:
            weaknesses.append(f"Limited champion pools (avg {avg_champions:.0f} per player)")
        
        # Data completeness
        complete_players = sum(1 for p in players_data if p['data_completeness']['completeness_score'] >= 0.8)
        if complete_players >= len(players_data) * 0.8:
            strengths.append("Excellent data quality for optimization")
        elif complete_players < len(players_data) * 0.5:
            weaknesses.append("Insufficient data for accurate optimization")
        
        # Display results
        if strengths:
            print("   ✅ Strengths:")
            for strength in strengths:
                print(f"      • {strength}")
        
        if weaknesses:
            print("   ⚠️ Areas for Improvement:")
            for weakness in weaknesses:
                print(f"      • {weakness}")
        
        if not strengths and not weaknesses:
            print("   📊 Balanced team composition with standard coverage")
    
    def _bulk_refresh_all_data(self) -> None:
        """Refresh data for all players."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("No players to refresh.")
            return
        
        if not self.engine.api_available:
            print("❌ API not available - cannot refresh data.")
            return
        
        print(f"🔄 Refreshing data for all {len(players)} players...")
        operations = [{'action': 'refresh_data', 'name': p.name} for p in players]
        results = self.engine.bulk_player_operations(operations)
        
        success_count = sum(1 for success, _ in results if success)
        print(f"✅ Refreshed data for {success_count}/{len(players)} players")
    
    def _performance_trends_analysis(self) -> None:
        """Analyze performance trends and insights."""
        print("\n" + "=" * 60)
        print("📈 PERFORMANCE TRENDS & INSIGHTS")
        print("=" * 60)
        
        analysis = self.engine.get_comprehensive_analysis()
        
        if 'error' in analysis:
            print(f"❌ Analysis failed: {analysis['error']}")
            return
        
        if not analysis.get('players'):
            print("No player data available for trend analysis.")
            return
        
        players_data = analysis['players']
        
        # Performance distribution analysis
        print("🎯 Performance Distribution:")
        
        # Analyze role performance across all players
        role_performance = {}
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            performances = []
            for player in players_data:
                role_data = player['role_analysis'][role]
                if role_data.get('performance_score', 0) > 0:
                    performances.append(role_data['performance_score'])
            
            if performances:
                avg_perf = sum(performances) / len(performances)
                role_performance[role] = {
                    'average': avg_perf,
                    'count': len(performances),
                    'max': max(performances),
                    'min': min(performances)
                }
        
        # Display role performance trends
        if role_performance:
            print("\n📊 Role Performance Averages:")
            sorted_roles = sorted(role_performance.items(), key=lambda x: x[1]['average'], reverse=True)
            
            for role, perf in sorted_roles:
                avg = perf['average']
                count = perf['count']
                indicator = "🟢" if avg >= 60 else "🟡" if avg >= 45 else "🔴"
                print(f"   {role.capitalize():8}: {indicator} {avg:.1f}% avg WR ({count} players with data)")
        
        # Champion mastery trends
        print("\n🏆 Champion Mastery Trends:")
        
        mastery_levels = []
        total_champions = []
        
        for player in players_data:
            champ_data = player.get('champion_analysis', {})
            if champ_data:
                mastery_levels.append(champ_data.get('mastery_7_champions', 0))
                total_champions.append(champ_data.get('total_champions', 0))
        
        if mastery_levels and total_champions:
            avg_mastery_7 = sum(mastery_levels) / len(mastery_levels)
            avg_total_champs = sum(total_champions) / len(total_champions)
            
            print(f"   Average Mastery 7 Champions: {avg_mastery_7:.1f} per player")
            print(f"   Average Total Champions: {avg_total_champs:.1f} per player")
            
            # Identify champion specialists vs generalists
            specialists = [p['name'] for p in players_data 
                          if p.get('champion_analysis', {}).get('mastery_7_champions', 0) >= 5]
            generalists = [p['name'] for p in players_data 
                          if p.get('champion_analysis', {}).get('total_champions', 0) >= 40]
            
            if specialists:
                print(f"   🎯 Champion Specialists: {', '.join(specialists)}")
            if generalists:
                print(f"   🌐 Champion Generalists: {', '.join(generalists)}")
        
        # Data freshness analysis
        print("\n📅 Data Freshness Analysis:")
        
        from datetime import datetime, timedelta
        now = datetime.now()
        
        fresh_data = 0  # < 7 days
        recent_data = 0  # 7-30 days
        stale_data = 0  # > 30 days
        no_data = 0
        
        players = self.engine.data_manager.load_player_data()
        for player in players:
            if not player.last_updated:
                no_data += 1
            else:
                days_old = (now - player.last_updated).days
                if days_old < 7:
                    fresh_data += 1
                elif days_old < 30:
                    recent_data += 1
                else:
                    stale_data += 1
        
        total = len(players)
        if total > 0:
            print(f"   🟢 Fresh (< 7 days): {fresh_data} players ({fresh_data/total*100:.0f}%)")
            print(f"   🟡 Recent (7-30 days): {recent_data} players ({recent_data/total*100:.0f}%)")
            print(f"   🔴 Stale (> 30 days): {stale_data} players ({stale_data/total*100:.0f}%)")
            print(f"   ⚫ No data: {no_data} players ({no_data/total*100:.0f}%)")
        
        # Recommendations based on trends
        print("\n💡 Trend-Based Recommendations:")
        
        if stale_data > total * 0.3:
            print("   • Consider refreshing player data for more accurate analysis")
        
        if role_performance:
            weak_roles = [role for role, perf in role_performance.items() if perf['average'] < 45]
            if weak_roles:
                print(f"   • Focus on improving performance in: {', '.join(weak_roles)}")
        
        if avg_mastery_7 < 2:
            print("   • Encourage players to achieve higher champion mastery")
    
    def _optimization_predictions(self) -> None:
        """Predict optimization outcomes and provide insights."""
        print("\n" + "=" * 60)
        print("🔮 OPTIMIZATION PREDICTIONS & INSIGHTS")
        print("=" * 60)
        
        players = self.engine.data_manager.load_player_data()
        
        if len(players) < 2:
            print("Need at least 2 players for optimization predictions.")
            return
        
        print("🔄 Analyzing potential team compositions...")
        
        # Get comprehensive analysis
        analysis = self.engine.get_comprehensive_analysis()
        
        if 'error' in analysis:
            print(f"❌ Analysis failed: {analysis['error']}")
            return
        
        # Predict optimization outcomes
        print("\n🎯 Optimization Predictions:")
        
        if len(players) >= 5:
            print("   ✅ Full team optimization possible")
            print("   📊 Expected outcome: Complete role assignments")
            
            # Predict likely assignments based on preferences
            players_data = analysis.get('players', [])
            if players_data:
                print("\n🔮 Likely Role Assignments:")
                
                for role in ["top", "jungle", "middle", "support", "bottom"]:
                    # Find players with highest preference for this role
                    role_candidates = []
                    for player in players_data:
                        pref = player['role_analysis'][role]['preference']
                        suitability = player['role_analysis'][role]['suitability']
                        if pref >= 4 or suitability in ['Excellent', 'Good']:
                            role_candidates.append((player['name'], pref, suitability))
                    
                    role_candidates.sort(key=lambda x: x[1], reverse=True)
                    
                    if role_candidates:
                        top_candidate = role_candidates[0]
                        confidence = "High" if top_candidate[1] >= 4 else "Medium"
                        print(f"   {role.capitalize():8}: {top_candidate[0]} (Confidence: {confidence})")
                    else:
                        print(f"   {role.capitalize():8}: No clear preference (will be assigned)")
        
        else:
            missing = 5 - len(players)
            print(f"   ⚠️ Partial optimization only ({missing} roles may be unassigned)")
            print("   📊 Expected outcome: Some roles left empty")
        
        # Predict potential issues
        print("\n⚠️ Potential Optimization Issues:")
        
        issues_found = False
        
        # Check for role conflicts
        if analysis.get('players'):
            role_conflicts = {}
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                high_pref_players = []
                for player in analysis['players']:
                    if player['role_analysis'][role]['preference'] >= 4:
                        high_pref_players.append(player['name'])
                
                if len(high_pref_players) > 1:
                    role_conflicts[role] = high_pref_players
                    issues_found = True
            
            if role_conflicts:
                print("   🔄 Role Preference Conflicts:")
                for role, players_list in role_conflicts.items():
                    print(f"      {role.capitalize()}: {', '.join(players_list)} all prefer this role")
            
            # Check for gaps
            role_gaps = []
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                suitable_players = sum(1 for player in analysis['players'] 
                                     if player['role_analysis'][role]['preference'] >= 3)
                if suitable_players == 0:
                    role_gaps.append(role)
                    issues_found = True
            
            if role_gaps:
                print("   ❌ Role Coverage Gaps:")
                print(f"      No players prefer: {', '.join(role_gaps)}")
        
        if not issues_found:
            print("   ✅ No major issues predicted - optimization should work well")
        
        # Success probability
        print("\n📈 Optimization Success Probability:")
        
        success_factors = 0
        total_factors = 5
        
        if len(players) >= 5:
            success_factors += 1
        if analysis.get('system_status', {}).get('api_available'):
            success_factors += 1
        if analysis.get('system_status', {}).get('players_with_api_data', 0) >= len(players) * 0.6:
            success_factors += 1
        if not role_conflicts:
            success_factors += 1
        if not role_gaps:
            success_factors += 1
        
        success_rate = (success_factors / total_factors) * 100
        
        if success_rate >= 80:
            indicator = "🟢"
            rating = "Excellent"
        elif success_rate >= 60:
            indicator = "🟡"
            rating = "Good"
        else:
            indicator = "🔴"
            rating = "Needs Improvement"
        
        print(f"   {indicator} Success Probability: {success_rate:.0f}% ({rating})")
        
        # Actionable recommendations
        print("\n🎯 Pre-Optimization Recommendations:")
        
        if len(players) < 5:
            print(f"   • Add {5 - len(players)} more players for complete team")
        
        if role_gaps:
            print(f"   • Find players who prefer: {', '.join(role_gaps)}")
        
        if analysis.get('system_status', {}).get('players_with_api_data', 0) < len(players) * 0.6:
            print("   • Refresh player data for better accuracy")
        
        if role_conflicts:
            print("   • Consider adjusting role preferences to reduce conflicts")
        
        # Offer to run optimization
        if len(players) >= 2:
            run_opt = input("\nRun optimization now to test predictions? (y/n): ").strip().lower()
            if run_opt == 'y':
                self._quick_optimize()
    
    def _display_analysis_results(self, analysis: Dict[str, Any]) -> None:
        """Display comprehensive analysis results."""
        print("\n" + "=" * 60)
        print("📊 COMPREHENSIVE TEAM ANALYSIS")
        print("=" * 60)
        
        # System status
        status = analysis.get('system_status', {})
        print(f"System Status:")
        print(f"  Players: {status.get('player_count', 0)}")
        print(f"  API: {'Available' if status.get('api_available') else 'Offline'}")
        print(f"  Ready for optimization: {'Yes' if status.get('ready_for_optimization') else 'No'}")
        
        # Team analysis
        if 'team_analysis' in analysis and analysis['team_analysis']:
            team = analysis['team_analysis']
            print(f"\nTeam Readiness: {team.get('optimization_readiness', 'Unknown')}")
            
            print("\nRole Coverage:")
            for role, coverage in team.get('role_coverage', {}).items():
                quality = coverage['coverage_quality']
                players = coverage['suitable_players']
                print(f"  {role.capitalize():8}: {quality:8} ({len(players)} suitable players)")
        
        # Individual players summary
        if 'players' in analysis:
            print(f"\nPlayer Summary ({len(analysis['players'])} players):")
            print("-" * 50)
            
            for player in analysis['players']:
                completeness = player['data_completeness']['completeness_score']
                print(f"{player['name']:20} │ Data: {completeness:.0%} │ ", end="")
                
                # Show top roles
                role_scores = []
                for role, data in player['role_analysis'].items():
                    if data['suitability'] in ['Excellent', 'Good']:
                        role_scores.append(f"{role}({data['preference']})")
                
                if role_scores:
                    print(f"Strong: {', '.join(role_scores[:3])}")
                else:
                    print("No strong roles identified")
        
        # Recommendations
        if 'recommendations' in analysis and analysis['recommendations']:
            print("\nRecommendations:")
            for rec in analysis['recommendations']:
                print(f"  💡 {rec}")
    
    def _individual_analysis_menu(self) -> None:
        """Menu for individual player analysis."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("\nNo players found.")
            return
        
        print("\nSelect player for detailed analysis:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name}")
        
        try:
            choice = int(input(f"\nEnter player number (1-{len(players)}): ")) - 1
            if 0 <= choice < len(players):
                self._view_player_analysis(players[choice])
            else:
                print("Invalid selection.")
        except ValueError:
            print("Please enter a valid number.")
    
    def _player_comparison_analysis(self) -> None:
        """Compare multiple players side by side."""
        players = self.engine.data_manager.load_player_data()
        
        if len(players) < 2:
            print("\nNeed at least 2 players for comparison.")
            return
        
        print("\nSelect players to compare (enter numbers separated by commas):")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name}")
        
        try:
            selection = input("\nPlayer numbers: ").strip()
            indices = [int(x.strip()) - 1 for x in selection.split(',')]
            
            selected_players = []
            for idx in indices:
                if 0 <= idx < len(players):
                    selected_players.append(players[idx].name)
                else:
                    print(f"Invalid player number: {idx + 1}")
                    return
            
            if len(selected_players) < 2:
                print("Please select at least 2 players.")
                return
            
            # Get analysis for selected players
            analysis = self.engine.get_comprehensive_analysis(selected_players)
            
            if 'error' in analysis:
                print(f"❌ Analysis failed: {analysis['error']}")
                return
            
            # Display comparison
            self._display_player_comparison(analysis['players'])
            
        except ValueError:
            print("Invalid input format.")
    
    def _display_player_comparison(self, players_analysis: List[Dict[str, Any]]) -> None:
        """Display side-by-side player comparison."""
        print("\n" + "=" * 80)
        print("👥 PLAYER COMPARISON")
        print("=" * 80)
        
        # Header
        print(f"{'Metric':<20}", end="")
        for player in players_analysis:
            print(f"{player['name']:<15}", end="")
        print()
        print("-" * 80)
        
        # Data completeness
        print(f"{'Data Completeness':<20}", end="")
        for player in players_analysis:
            completeness = player['data_completeness']['completeness_score']
            print(f"{completeness:.0%}{'':11}", end="")
        print()
        
        # Role preferences comparison
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            print(f"{role.capitalize() + ' Preference':<20}", end="")
            for player in players_analysis:
                pref = player['role_analysis'][role]['preference']
                suitability = player['role_analysis'][role]['suitability']
                print(f"{pref}/5 ({suitability[:4]}){'':3}", end="")
            print()
        
        # Champion data
        print(f"{'Total Champions':<20}", end="")
        for player in players_analysis:
            champ_data = player.get('champion_analysis', {})
            total = champ_data.get('total_champions', 0)
            print(f"{total}{'':12}", end="")
        print()
        
        print(f"{'Mastery 7 Champs':<20}", end="")
        for player in players_analysis:
            champ_data = player.get('champion_analysis', {})
            m7 = champ_data.get('mastery_7_champions', 0)
            print(f"{m7}{'':12}", end="")
        print()
    
    def _role_coverage_analysis(self) -> None:
        """Comprehensive role coverage and gap analysis."""
        analysis = self.engine.get_comprehensive_analysis()
        
        if 'error' in analysis:
            print(f"❌ Analysis failed: {analysis['error']}")
            return
        
        print("\n" + "=" * 70)
        print("🎯 ROLE COVERAGE & GAP ANALYSIS")
        print("=" * 70)
        
        if not analysis.get('players'):
            print("No player data available for role analysis.")
            return
        
        # Overall coverage summary
        print("📊 Coverage Summary:")
        
        role_stats = {}
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            excellent_players = []
            good_players = []
            fair_players = []
            poor_players = []
            
            for player in analysis['players']:
                role_data = player['role_analysis'][role]
                suitability = role_data['suitability']
                preference = role_data['preference']
                
                if suitability == 'Excellent' or preference >= 5:
                    excellent_players.append(player['name'])
                elif suitability == 'Good' or preference >= 4:
                    good_players.append(player['name'])
                elif suitability == 'Fair' or preference >= 3:
                    fair_players.append(player['name'])
                else:
                    poor_players.append(player['name'])
            
            role_stats[role] = {
                'excellent': excellent_players,
                'good': good_players,
                'fair': fair_players,
                'poor': poor_players
            }
            
            # Coverage indicator
            total_suitable = len(excellent_players) + len(good_players)
            if total_suitable >= 3:
                indicator = "🟢 Strong"
            elif total_suitable >= 2:
                indicator = "🟡 Adequate"
            elif total_suitable >= 1:
                indicator = "🟠 Limited"
            else:
                indicator = "🔴 Weak"
            
            print(f"   {role.capitalize():8}: {indicator:12} ({total_suitable} strong candidates)")
        
        # Detailed role breakdown
        print(f"\n🔍 Detailed Role Breakdown:")
        
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            print(f"\n🎭 {role.upper()} ROLE ANALYSIS:")
            print("-" * 30)
            
            stats = role_stats[role]
            
            # Show excellent candidates
            if stats['excellent']:
                print(f"   🟢 Excellent ({len(stats['excellent'])}): {', '.join(stats['excellent'])}")
            
            # Show good candidates
            if stats['good']:
                print(f"   🟡 Good ({len(stats['good'])}): {', '.join(stats['good'])}")
            
            # Show concerns
            if not stats['excellent'] and not stats['good']:
                print(f"   🔴 No strong candidates available")
                if stats['fair']:
                    print(f"   ⚠️ Fair options: {', '.join(stats['fair'])}")
            
            # Detailed candidate analysis
            role_candidates = []
            for player in analysis['players']:
                role_data = player['role_analysis'][role]
                role_candidates.append({
                    'name': player['name'],
                    'preference': role_data['preference'],
                    'suitability': role_data['suitability'],
                    'performance': role_data.get('performance_score', 0),
                    'champions': role_data.get('champion_pool_size', 0)
                })
            
            # Sort by overall suitability
            suitability_order = {'Excellent': 4, 'Good': 3, 'Fair': 2, 'Poor': 1, 'Unknown': 0}
            role_candidates.sort(key=lambda x: (suitability_order.get(x['suitability'], 0), x['preference']), reverse=True)
            
            print(f"   📋 Top Candidates:")
            for i, candidate in enumerate(role_candidates[:3], 1):
                # Suitability indicator
                suit_indicator = {
                    'Excellent': '🟢',
                    'Good': '🟡', 
                    'Fair': '🟠',
                    'Poor': '🔴',
                    'Unknown': '⚫'
                }.get(candidate['suitability'], '⚫')
                
                print(f"      {i}. {candidate['name']:15} {suit_indicator} {candidate['suitability']:10} (Pref: {candidate['preference']}/5)", end="")
                
                if candidate['performance'] > 0:
                    print(f" | WR: {candidate['performance']:.0f}%", end="")
                if candidate['champions'] > 0:
                    print(f" | {candidate['champions']} champs", end="")
                print()
        
        # Gap analysis and recommendations
        print(f"\n⚠️ Gap Analysis & Recommendations:")
        
        critical_gaps = []
        moderate_gaps = []
        
        for role, stats in role_stats.items():
            total_suitable = len(stats['excellent']) + len(stats['good'])
            
            if total_suitable == 0:
                critical_gaps.append(role)
            elif total_suitable == 1:
                moderate_gaps.append(role)
        
        if critical_gaps:
            print(f"   🔴 Critical Gaps: {', '.join(critical_gaps)}")
            print(f"      • Urgently need players who prefer these roles")
            print(f"      • Consider role preference adjustments")
        
        if moderate_gaps:
            print(f"   🟡 Limited Options: {', '.join(moderate_gaps)}")
            print(f"      • Could benefit from additional players")
            print(f"      • Single point of failure risk")
        
        if not critical_gaps and not moderate_gaps:
            print(f"   ✅ Good coverage across all roles")
            print(f"   💡 Consider optimizing for best performance matches")
        
        # Flexibility analysis
        print(f"\n🔄 Player Flexibility Analysis:")
        
        flexible_players = []
        specialists = []
        
        for player in analysis['players']:
            strong_roles = []
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                role_data = player['role_analysis'][role]
                if role_data['suitability'] in ['Excellent', 'Good'] or role_data['preference'] >= 4:
                    strong_roles.append(role)
            
            if len(strong_roles) >= 3:
                flexible_players.append((player['name'], strong_roles))
            elif len(strong_roles) == 1:
                specialists.append((player['name'], strong_roles[0]))
        
        if flexible_players:
            print(f"   🌐 Flexible Players:")
            for name, roles in flexible_players:
                print(f"      • {name}: {', '.join(roles)}")
        
        if specialists:
            print(f"   🎯 Role Specialists:")
            for name, role in specialists:
                print(f"      • {name}: {role} specialist")
        
        # Quick action suggestions
        print(f"\n⚡ Quick Actions:")
        if critical_gaps:
            print(f"   • Add players for critical gaps: {', '.join(critical_gaps)}")
        if moderate_gaps:
            print(f"   • Consider backup options for: {', '.join(moderate_gaps)}")
        
        action = input("\nView detailed player analysis for a specific role? (role name or Enter to continue): ").strip().lower()
        if action in ["top", "jungle", "middle", "support", "bottom"]:
            self._detailed_role_analysis(action, analysis)
    
    def _detailed_role_analysis(self, role: str, analysis: Dict[str, Any]) -> None:
        """Detailed analysis for a specific role."""
        print(f"\n" + "=" * 50)
        print(f"🎭 DETAILED {role.upper()} ROLE ANALYSIS")
        print("=" * 50)
        
        role_players = []
        for player in analysis['players']:
            role_data = player['role_analysis'][role]
            
            # Get champion data for this role
            champion_info = ""
            if 'top_champions' in role_data and role_data['top_champions']:
                top_champ = role_data['top_champions'][0]
                champion_info = f"{top_champ['name']} (L{top_champ['mastery_level']})"
            
            role_players.append({
                'name': player['name'],
                'preference': role_data['preference'],
                'suitability': role_data['suitability'],
                'performance': role_data.get('performance_score', 0),
                'champions': role_data.get('champion_pool_size', 0),
                'top_champion': champion_info,
                'data_quality': player['data_completeness']['completeness_score']
            })
        
        # Sort by overall suitability
        suitability_order = {'Excellent': 4, 'Good': 3, 'Fair': 2, 'Poor': 1, 'Unknown': 0}
        role_players.sort(key=lambda x: (suitability_order.get(x['suitability'], 0), x['preference']), reverse=True)
        
        print(f"📊 All Players for {role.capitalize()} Role:")
        print("-" * 50)
        
        for i, player_data in enumerate(role_players, 1):
            # Status indicator
            suit_indicator = {
                'Excellent': '🟢',
                'Good': '🟡', 
                'Fair': '🟠',
                'Poor': '🔴',
                'Unknown': '⚫'
            }.get(player_data['suitability'], '⚫')
            
            print(f"{i:2}. {player_data['name']:15} {suit_indicator} {player_data['suitability']:10}")
            print(f"     Preference: {player_data['preference']}/5 | Data Quality: {player_data['data_quality']:.0%}")
            
            if player_data['performance'] > 0:
                print(f"     Performance: {player_data['performance']:.1f}% WR", end="")
            if player_data['champions'] > 0:
                print(f" | Champions: {player_data['champions']}", end="")
            if player_data['top_champion']:
                print(f" | Top: {player_data['top_champion']}", end="")
            print()
            print()
        
        # Role-specific recommendations
        print(f"💡 {role.capitalize()} Role Recommendations:")
        
        excellent_players = [p for p in role_players if p['suitability'] == 'Excellent']
        good_players = [p for p in role_players if p['suitability'] == 'Good']
        
        if excellent_players:
            print(f"   🎯 Primary choice: {excellent_players[0]['name']}")
            if len(excellent_players) > 1:
                print(f"   🔄 Alternatives: {', '.join(p['name'] for p in excellent_players[1:3])}")
        elif good_players:
            print(f"   🎯 Best available: {good_players[0]['name']}")
            if len(good_players) > 1:
                print(f"   🔄 Alternatives: {', '.join(p['name'] for p in good_players[1:3])}")
        else:
            print(f"   ⚠️ No strong candidates - consider adding {role} specialists")
    
    def _champion_pool_analysis(self) -> None:
        """Comprehensive champion pool analysis across all players."""
        players = self.engine.data_manager.load_player_data()
        
        if not players:
            print("\nNo players found.")
            return
        
        print("\n" + "=" * 70)
        print("🏆 CHAMPION POOL ANALYSIS - Team Champion Coverage")
        print("=" * 70)
        
        # Collect all champion data
        all_champions = {}
        total_mastery_points = 0
        players_with_data = 0
        
        for player in players:
            if player.champion_masteries:
                players_with_data += 1
                for champ_id, mastery in player.champion_masteries.items():
                    champ_name = mastery.champion_name
                    if champ_name not in all_champions:
                        all_champions[champ_name] = {
                            'players': [],
                            'total_points': 0,
                            'max_level': 0,
                            'roles': set()
                        }
                    
                    all_champions[champ_name]['players'].append({
                        'name': player.name,
                        'level': mastery.mastery_level,
                        'points': mastery.mastery_points
                    })
                    all_champions[champ_name]['total_points'] += mastery.mastery_points
                    all_champions[champ_name]['max_level'] = max(all_champions[champ_name]['max_level'], mastery.mastery_level)
                    all_champions[champ_name]['roles'].update(mastery.primary_roles)
                    
                    total_mastery_points += mastery.mastery_points
        
        if not all_champions:
            print("No champion data available. Refresh player data to see champion analysis.")
            return
        
        # Overview statistics
        print(f"📊 Champion Pool Overview:")
        print(f"   Total Unique Champions: {len(all_champions)}")
        print(f"   Players with Data: {players_with_data}/{len(players)}")
        print(f"   Total Mastery Points: {total_mastery_points:,}")
        print(f"   Average per Player: {total_mastery_points // players_with_data:,}" if players_with_data > 0 else "")
        
        # Champion coverage by role
        print(f"\n🎭 Champion Coverage by Role:")
        
        role_champions = {
            'top': set(),
            'jungle': set(), 
            'middle': set(),
            'support': set(),
            'bottom': set()
        }
        
        for champ_name, data in all_champions.items():
            for role in data['roles']:
                if role in role_champions:
                    role_champions[role].add(champ_name)
        
        for role, champions in role_champions.items():
            coverage_indicator = "🟢" if len(champions) >= 15 else "🟡" if len(champions) >= 8 else "🔴"
            print(f"   {role.capitalize():8}: {coverage_indicator} {len(champions):3} champions")
        
        # Most popular champions
        print(f"\n🌟 Most Popular Champions (by player count):")
        popular_champs = sorted(all_champions.items(), key=lambda x: len(x[1]['players']), reverse=True)
        
        for i, (champ_name, data) in enumerate(popular_champs[:10], 1):
            player_count = len(data['players'])
            max_level = data['max_level']
            roles = ', '.join(sorted(data['roles'])) if data['roles'] else 'Flexible'
            
            popularity_indicator = "🔥" if player_count >= 3 else "✨" if player_count >= 2 else "⭐"
            
            print(f"   {i:2}. {champ_name:15} {popularity_indicator} {player_count} players | Max L{max_level} | {roles}")
        
        # Highest mastery champions
        print(f"\n🏅 Highest Mastery Champions (by total points):")
        high_mastery = sorted(all_champions.items(), key=lambda x: x[1]['total_points'], reverse=True)
        
        for i, (champ_name, data) in enumerate(high_mastery[:10], 1):
            total_points = data['total_points']
            player_count = len(data['players'])
            max_level = data['max_level']
            
            mastery_indicator = "🥇" if max_level == 7 else "🥈" if max_level >= 5 else "🥉"
            
            print(f"   {i:2}. {champ_name:15} {mastery_indicator} {total_points:,} pts | {player_count} players | Max L{max_level}")
        
        # Champion diversity analysis
        print(f"\n🌈 Champion Diversity Analysis:")
        
        # Calculate diversity metrics
        mastery_7_champions = sum(1 for data in all_champions.values() if data['max_level'] == 7)
        mastery_6_plus = sum(1 for data in all_champions.values() if data['max_level'] >= 6)
        
        print(f"   Mastery Level 7: {mastery_7_champions} champions")
        print(f"   Mastery Level 6+: {mastery_6_plus} champions")
        
        # Role distribution
        role_distribution = {}
        for role in ['top', 'jungle', 'middle', 'support', 'bottom']:
            role_distribution[role] = len(role_champions[role])
        
        most_covered = max(role_distribution.items(), key=lambda x: x[1])
        least_covered = min(role_distribution.items(), key=lambda x: x[1])
        
        print(f"   Most Covered Role: {most_covered[0]} ({most_covered[1]} champions)")
        print(f"   Least Covered Role: {least_covered[0]} ({least_covered[1]} champions)")
        
        # Player specialization analysis
        print(f"\n👤 Player Specialization Analysis:")
        
        specialists = []  # Players with many high-level champions
        generalists = []  # Players with many different champions
        
        for player in players:
            if player.champion_masteries:
                total_champs = len(player.champion_masteries)
                mastery_7_count = sum(1 for m in player.champion_masteries.values() if m.mastery_level == 7)
                
                if mastery_7_count >= 5:
                    specialists.append((player.name, mastery_7_count, total_champs))
                elif total_champs >= 40:
                    generalists.append((player.name, total_champs, mastery_7_count))
        
        if specialists:
            print(f"   🎯 Champion Specialists (5+ Level 7):")
            for name, m7_count, total in specialists:
                print(f"      • {name}: {m7_count} Level 7 champions ({total} total)")
        
        if generalists:
            print(f"   🌐 Champion Generalists (40+ champions):")
            for name, total, m7_count in generalists:
                print(f"      • {name}: {total} champions ({m7_count} Level 7)")
        
        # Gap analysis
        print(f"\n⚠️ Champion Pool Gaps & Recommendations:")
        
        gaps_found = False
        
        # Check for role gaps
        for role, champion_count in role_distribution.items():
            if champion_count < 5:
                print(f"   🔴 Limited {role} champion pool ({champion_count} champions)")
                gaps_found = True
        
        # Check for popular champion coverage
        meta_champions = ['Jinx', 'Caitlyn', 'Graves', 'Lee Sin', 'Yasuo', 'Zed', 'Ahri', 'LeBlanc', 
                         'Thresh', 'Leona', 'Garen', 'Darius']  # Example meta champions
        
        missing_meta = []
        for champ in meta_champions:
            if champ not in all_champions:
                missing_meta.append(champ)
        
        if missing_meta and len(missing_meta) > len(meta_champions) * 0.5:
            print(f"   🟡 Limited meta champion coverage")
            gaps_found = True
        
        if not gaps_found:
            print(f"   ✅ Good champion pool coverage across all roles")
        
        # Interactive options
        print(f"\n🔍 Detailed Analysis Options:")
        print("   • Type a champion name to see who plays it")
        print("   • Type a role name to see champion coverage")
        print("   • Type 'export' to save champion data")
        
        action = input("\nEnter option or press Enter to continue: ").strip()
        
        if action.lower() in all_champions:
            self._champion_detail_analysis(action, all_champions[action])
        elif action.lower() in ['top', 'jungle', 'middle', 'support', 'bottom']:
            self._role_champion_analysis(action.lower(), role_champions[action.lower()], all_champions)
        elif action.lower() == 'export':
            self._export_champion_data(all_champions)
    
    def _champion_detail_analysis(self, champion_name: str, champion_data: Dict[str, Any]) -> None:
        """Detailed analysis for a specific champion."""
        print(f"\n🏆 {champion_name.upper()} - Detailed Analysis")
        print("-" * 40)
        
        print(f"Players who play {champion_name}:")
        
        # Sort players by mastery level and points
        players_sorted = sorted(champion_data['players'], 
                              key=lambda x: (x['level'], x['points']), reverse=True)
        
        for i, player_data in enumerate(players_sorted, 1):
            level_indicator = "🥇" if player_data['level'] == 7 else "🥈" if player_data['level'] >= 5 else "🥉"
            print(f"   {i}. {player_data['name']:15} {level_indicator} Level {player_data['level']} ({player_data['points']:,} points)")
        
        roles_str = ', '.join(sorted(champion_data['roles'])) if champion_data['roles'] else 'Flexible'
        print(f"\nPrimary Roles: {roles_str}")
        print(f"Highest Mastery: Level {champion_data['max_level']}")
        print(f"Total Team Points: {champion_data['total_points']:,}")
    
    def _role_champion_analysis(self, role: str, role_champions: set, all_champions: Dict[str, Any]) -> None:
        """Detailed champion analysis for a specific role."""
        print(f"\n🎭 {role.upper()} ROLE - Champion Analysis")
        print("-" * 40)
        
        if not role_champions:
            print(f"No champions available for {role} role.")
            return
        
        # Get champion data for this role
        role_champion_data = []
        for champ_name in role_champions:
            if champ_name in all_champions:
                data = all_champions[champ_name]
                role_champion_data.append((champ_name, data))
        
        # Sort by popularity (player count) and mastery
        role_champion_data.sort(key=lambda x: (len(x[1]['players']), x[1]['max_level']), reverse=True)
        
        print(f"Available {role} champions ({len(role_champion_data)}):")
        
        for i, (champ_name, data) in enumerate(role_champion_data[:15], 1):  # Show top 15
            player_count = len(data['players'])
            max_level = data['max_level']
            
            popularity_indicator = "🔥" if player_count >= 3 else "✨" if player_count >= 2 else "⭐"
            mastery_indicator = "🥇" if max_level == 7 else "🥈" if max_level >= 5 else "🥉"
            
            print(f"   {i:2}. {champ_name:15} {popularity_indicator} {player_count} players {mastery_indicator} Max L{max_level}")
        
        if len(role_champion_data) > 15:
            print(f"   ... and {len(role_champion_data) - 15} more champions")
    
    def _export_champion_data(self, all_champions: Dict[str, Any]) -> None:
        """Export champion data to file."""
        try:
            import json
            from datetime import datetime
            
            # Prepare export data
            export_data = {
                'export_timestamp': datetime.now().isoformat(),
                'total_champions': len(all_champions),
                'champions': {}
            }
            
            for champ_name, data in all_champions.items():
                export_data['champions'][champ_name] = {
                    'players': data['players'],
                    'total_points': data['total_points'],
                    'max_level': data['max_level'],
                    'roles': list(data['roles']),
                    'player_count': len(data['players'])
                }
            
            filename = f"champion_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(filename, 'w') as f:
                json.dump(export_data, f, indent=2)
            
            print(f"✅ Champion data exported to {filename}")
            
        except Exception as e:
            print(f"❌ Export failed: {e}")
    
    def _champion_performance_analytics(self) -> None:
        """Champion performance analytics interface with baseline comparisons and recommendations."""
        while True:
            print("\n" + "=" * 70)
            print("🏅 CHAMPION PERFORMANCE ANALYTICS")
            print("=" * 70)
            
            # Check if analytics engines are available
            if not hasattr(self.engine, 'historical_analytics_engine') or not self.engine.historical_analytics_engine:
                print("❌ Analytics engines not available. Please check system configuration.")
                return
            
            players = self.engine.data_manager.load_player_data()
            if not players:
                print("No players found. Add players first to see champion performance analytics.")
                return
            
            print(f"📊 Analytics Options:")
            print("1. 🎯 Champion Performance Analysis")
            print("2. 🏆 Champion Recommendations")
            print("3. 🤝 Champion Synergy Analysis")
            print("4. 📈 Performance Trend Visualization")
            print("5. 🔄 Refresh Analytics Data")
            print("6. 🏠 Back to Analysis Menu")
            
            choice = input("\nEnter your choice (1-6): ").strip()
            
            if choice == "1":
                self._champion_performance_analysis()
            elif choice == "2":
                self._champion_recommendations_interface()
            elif choice == "3":
                self._champion_synergy_interface()
            elif choice == "4":
                self._performance_trend_visualization()
            elif choice == "5":
                self._refresh_analytics_data()
            elif choice == "6":
                break
            else:
                print("Invalid choice. Please enter a number between 1-6.")
    
    def _champion_performance_analysis(self) -> None:
        """Display champion-specific performance reports with baseline comparisons."""
        print("\n🎯 Champion Performance Analysis")
        print("-" * 40)
        
        # Select player
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players available.")
            return
        
        print("Select a player:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name} ({player.summoner_name})")
        
        try:
            player_idx = int(input("\nEnter player number: ").strip()) - 1
            if not 0 <= player_idx < len(players):
                print("Invalid player selection.")
                return
            
            selected_player = players[player_idx]
        except ValueError:
            print("Invalid input. Please enter a number.")
            return
        
        # Get player's champion performance data
        try:
            print(f"\n🔄 Analyzing champion performance for {selected_player.name}...")
            
            # Use analytics engine to get champion performance
            from .analytics_models import AnalyticsFilters, DateRange
            from datetime import datetime, timedelta
            
            # Create filters for recent performance (last 3 months)
            end_date = datetime.now()
            start_date = end_date - timedelta(days=90)
            filters = AnalyticsFilters(
                date_range=DateRange(start_date=start_date, end_date=end_date),
                min_games=1
            )
            
            # Get player analytics
            player_analytics = self.engine.historical_analytics_engine.analyze_player_performance(
                selected_player.puuid, filters
            )
            
            # Display results
            self._display_champion_performance_results(selected_player, player_analytics)
            
        except Exception as e:
            self.logger.error(f"Champion performance analysis failed: {e}")
            print(f"❌ Analysis failed: {e}")
            
            # Fallback to basic champion mastery display
            print("\n📊 Fallback: Champion Mastery Data")
            self._display_basic_champion_mastery(selected_player)
    
    def _display_champion_performance_results(self, player, analytics) -> None:
        """Display comprehensive champion performance results."""
        print(f"\n" + "=" * 60)
        print(f"🏅 CHAMPION PERFORMANCE REPORT - {player.name}")
        print("=" * 60)
        
        if not analytics:
            print("No analytics data available.")
            return
        
        # Handle different analytics formats
        champion_data = None
        if hasattr(analytics, 'champion_performance') and analytics.champion_performance:
            champion_data = analytics.champion_performance
        elif hasattr(analytics, 'performance') and analytics.performance:
            # Single champion analysis
            champion_data = {analytics.champion_id: analytics}
        
        if not champion_data:
            print("No champion performance data available.")
            print("This may be due to insufficient recent match data.")
            return
        
        # Overall performance summary
        if hasattr(analytics, 'analysis_period') and analytics.analysis_period:
            print(f"📊 Analysis Period: {analytics.analysis_period.start_date.strftime('%Y-%m-%d')} to {analytics.analysis_period.end_date.strftime('%Y-%m-%d')}")
        
        total_games = 0
        if isinstance(champion_data, dict):
            # For champion_performance dict, access games_played through performance attribute
            for perf in champion_data.values():
                if hasattr(perf, 'performance') and perf.performance:
                    total_games += getattr(perf.performance, 'games_played', 0)
                else:
                    total_games += getattr(perf, 'games_played', 0)
        else:
            if hasattr(champion_data, 'performance') and champion_data.performance:
                total_games = getattr(champion_data.performance, 'games_played', 0)
            else:
                total_games = getattr(champion_data, 'games_played', 0)
            
        print(f"🎮 Total Games Analyzed: {total_games}")
        
        if total_games == 0:
            print("\n⚠️ No games found in the analysis period.")
            print("Try expanding the date range or check if match data is available.")
            return
        
        # Top performing champions
        print(f"\n🏆 Champion Performance Analysis:")
        print("-" * 50)
        
        # Handle both dict and single champion formats
        if isinstance(champion_data, dict):
            # Sort champions by performance
            try:
                def get_sort_key(item):
                    """Get sorting key for champion performance."""
                    champion_id, perf = item
                    
                    # Try to get performance delta from baseline
                    baseline_data = getattr(perf, 'performance_vs_baseline', None)
                    if baseline_data and isinstance(baseline_data, dict):
                        overall_delta = baseline_data.get('overall')
                        if overall_delta and hasattr(overall_delta, 'delta_percentage'):
                            return overall_delta.delta_percentage
                    
                    # Fallback to win rate from performance
                    performance_data = getattr(perf, 'performance', perf)
                    return getattr(performance_data, 'win_rate', 0) * 100  # Convert to percentage for comparison
                
                sorted_champions = sorted(
                    champion_data.items(),
                    key=get_sort_key,
                    reverse=True
                )
            except Exception as e:
                print(f"Warning: Could not sort champions ({e}), using original order")
                sorted_champions = list(champion_data.items())
            
            for i, (champion_id, perf) in enumerate(sorted_champions[:10], 1):
                self._display_single_champion_performance(i, perf)
        else:
            # Single champion
            self._display_single_champion_performance(1, champion_data)
        
        # Role-specific performance (only for multi-champion analysis)
        if isinstance(champion_data, dict) and len(champion_data) > 1:
            print(f"\n🎭 Performance by Role:")
            print("-" * 30)
            
            role_performance = {}
            for champion_id, perf in champion_data.items():
                role = getattr(perf, 'role', 'Unknown')
                if role not in role_performance:
                    role_performance[role] = []
                role_performance[role].append(perf)
            
            for role, champions in role_performance.items():
                if champions:
                    # Access win_rate and games_played through the performance attribute
                    avg_wr = sum(getattr(getattr(c, 'performance', c), 'win_rate', 0) for c in champions) / len(champions)
                    total_games = sum(getattr(getattr(c, 'performance', c), 'games_played', 0) for c in champions)
                    best_champ = max(champions, key=lambda x: getattr(getattr(x, 'performance', x), 'win_rate', 0))
                    best_name = getattr(best_champ, 'champion_name', 'Unknown')
                    
                    print(f"{role.capitalize():8}: {avg_wr*100:.1f}% avg WR | {total_games:3} games | Best: {best_name}")
        
        # Show data quality information
        print(f"\n📊 Data Quality Information:")
        if hasattr(analytics, 'confidence_scores') and analytics.confidence_scores:
            avg_confidence = sum(analytics.confidence_scores.values()) / len(analytics.confidence_scores)
            print(f"   Average Confidence: {avg_confidence:.1%}")
        
        print(f"   Sample Size: {total_games} games")
        if total_games < 10:
            print("   ⚠️ Small sample size - results may not be statistically significant")
        elif total_games < 30:
            print("   ⚠️ Moderate sample size - consider more games for better accuracy")
        else:
            print("   ✅ Good sample size for reliable analysis")
    
    def _display_single_champion_performance(self, index, perf) -> None:
        """Display performance data for a single champion."""
        try:
            champion_name = getattr(perf, 'champion_name', 'Unknown Champion')
            
            # Access performance metrics correctly
            performance_data = getattr(perf, 'performance', perf)  # Fallback to perf itself if no performance attr
            
            win_rate = getattr(performance_data, 'win_rate', 0)
            games_played = getattr(performance_data, 'games_played', 0)
            
            # Get performance delta if available
            delta = getattr(perf, 'performance_vs_baseline', None)
            if delta and isinstance(delta, dict):
                # Handle dict format like {"overall": PerformanceDelta}
                overall_delta = delta.get('overall')
                if overall_delta and hasattr(overall_delta, 'delta_percentage'):
                    delta_str = f"+{overall_delta.delta_percentage:.1f}%" if overall_delta.delta_percentage > 0 else f"{overall_delta.delta_percentage:.1f}%"
                    
                    # Performance indicator
                    if overall_delta.delta_percentage > 10:
                        indicator = "🟢"
                    elif overall_delta.delta_percentage > 0:
                        indicator = "🟡"
                    else:
                        indicator = "🔴"
                else:
                    delta_str = "N/A"
                    indicator = "⚪"
            elif delta and hasattr(delta, 'delta_percentage'):
                # Handle direct PerformanceDelta format
                delta_str = f"+{delta.delta_percentage:.1f}%" if delta.delta_percentage > 0 else f"{delta.delta_percentage:.1f}%"
                
                # Performance indicator
                if delta.delta_percentage > 10:
                    indicator = "🟢"
                elif delta.delta_percentage > 0:
                    indicator = "🟡"
                else:
                    indicator = "🔴"
            else:
                delta_str = "N/A"
                indicator = "⚪"
            
            print(f"{index:2}. {champion_name:15} {indicator} {win_rate*100:.1f}% WR | {games_played:2} games | {delta_str} vs baseline")
            
            # Show detailed metrics for top 3
            if index <= 3:
                avg_kda = getattr(performance_data, 'avg_kda', 0)
                avg_cs = getattr(performance_data, 'avg_cs_per_min', 0)
                avg_vision = getattr(performance_data, 'avg_vision_score', 0)
                
                print(f"    📈 KDA: {avg_kda:.2f} | CS/min: {avg_cs:.1f} | Vision: {avg_vision:.1f}")
                
                # Recent form if available
                recent_form = getattr(perf, 'recent_form', None)
                if recent_form:
                    trend_direction = getattr(recent_form, 'trend_direction', 'stable')
                    recent_wr = getattr(recent_form, 'recent_win_rate', 0)
                    recent_games = getattr(recent_form, 'recent_games', 0)
                    
                    recent_trend = "📈" if trend_direction == 'improving' else "📉" if trend_direction == 'declining' else "➡️"
                    print(f"    {recent_trend} Recent Form: {recent_wr*100:.1f}% WR (last {recent_games} games)")
                    
        except Exception as e:
            print(f"{index:2}. Champion data error: {str(e)[:50]}...")
    
    def _display_basic_champion_mastery(self, player) -> None:
        """Display basic champion mastery as fallback."""
        if not player.champion_masteries:
            print("No champion mastery data available.")
            return
        
        print(f"\n🏆 Champion Mastery for {player.name}:")
        print("-" * 40)
        
        # Sort by mastery level and points
        sorted_masteries = sorted(
            player.champion_masteries.items(),
            key=lambda x: (x[1].mastery_level, x[1].mastery_points),
            reverse=True
        )
        
        for i, (champ_id, mastery) in enumerate(sorted_masteries[:15], 1):
            level_indicator = "🥇" if mastery.mastery_level == 7 else "🥈" if mastery.mastery_level >= 5 else "🥉"
            roles = ", ".join(mastery.primary_roles) if mastery.primary_roles else "Flexible"
            
            print(f"{i:2}. {mastery.champion_name:15} {level_indicator} L{mastery.mastery_level} | {mastery.mastery_points:,} pts | {roles}")
    
    def _champion_recommendations_interface(self) -> None:
        """Champion recommendation interface with detailed explanations."""
        print("\n🏆 Champion Recommendations")
        print("-" * 30)
        
        # Select player
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players available.")
            return
        
        print("Select a player:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name} ({player.summoner_name})")
        
        try:
            player_idx = int(input("\nEnter player number: ").strip()) - 1
            if not 0 <= player_idx < len(players):
                print("Invalid player selection.")
                return
            
            selected_player = players[player_idx]
        except ValueError:
            print("Invalid input. Please enter a number.")
            return
        
        # Select role
        print("\nSelect role:")
        roles = ["top", "jungle", "middle", "support", "bottom"]
        for i, role in enumerate(roles, 1):
            print(f"{i}. {role.capitalize()}")
        
        try:
            role_idx = int(input("\nEnter role number: ").strip()) - 1
            if not 0 <= role_idx < len(roles):
                print("Invalid role selection.")
                return
            
            selected_role = roles[role_idx]
        except ValueError:
            print("Invalid input. Please enter a number.")
            return
        
        # Get team context (simplified for now)
        from .analytics_models import TeamContext
        team_context = TeamContext(
            existing_picks={},
            target_role=selected_role,
            target_player_puuid=selected_player.puuid
        )
        
        try:
            print(f"\n🔄 Generating champion recommendations for {selected_player.name} ({selected_role})...")
            
            # Get recommendations
            recommendations = self.engine.champion_recommendation_engine.get_champion_recommendations(
                selected_player.puuid, selected_role, team_context
            )
            
            # Display recommendations
            self._display_champion_recommendations(selected_player, selected_role, recommendations)
            
        except Exception as e:
            self.logger.error(f"Champion recommendation failed: {e}")
            print(f"❌ Recommendation failed: {e}")
            
            # Fallback to mastery-based recommendations
            print("\n📊 Fallback: Mastery-Based Recommendations")
            self._display_mastery_recommendations(selected_player, selected_role)
    
    def _display_champion_recommendations(self, player, role, recommendations) -> None:
        """Display detailed champion recommendations."""
        print(f"\n" + "=" * 60)
        print(f"🏆 CHAMPION RECOMMENDATIONS - {player.name} ({role.upper()})")
        print("=" * 60)
        
        if not recommendations:
            print("No recommendations available.")
            print("This may be due to insufficient historical data or analysis engine unavailability.")
            return
        
        print(f"📊 Found {len(recommendations)} recommendations")
        print("-" * 50)
        
        for i, rec in enumerate(recommendations[:10], 1):
            try:
                champion_name = getattr(rec, 'champion_name', 'Unknown Champion')
                rec_score = getattr(rec, 'recommendation_score', 0)
                confidence = getattr(rec, 'confidence', 0)
                
                # Confidence indicator
                conf_indicator = "🟢" if confidence >= 0.8 else "🟡" if confidence >= 0.6 else "🔴"
                
                print(f"{i:2}. {champion_name:15} {conf_indicator} Score: {rec_score:.2f} | Confidence: {confidence:.1%}")
                
                # Show detailed breakdown for top 3
                if i <= 3:
                    self._display_recommendation_details(rec)
                    
            except Exception as e:
                print(f"{i:2}. Recommendation data error: {str(e)[:50]}...")
        
        # Show recommendation methodology
        print(f"\n📊 Recommendation Methodology:")
        print("   🎯 Individual Performance: Historical success rate with champion")
        print("   🤝 Team Synergy: Compatibility with team composition")
        print("   📈 Recent Form: Recent performance trends")
        print("   🏆 Meta Relevance: Current meta strength")
        print("   📊 Data Confidence: Quality and quantity of available data")
        
        # Show additional tips
        print(f"\n💡 Tips:")
        print("   • Higher confidence scores indicate more reliable recommendations")
        print("   • Consider both score and confidence when making decisions")
        print("   • Green indicators suggest strong recommendations")
    
    def _display_recommendation_details(self, rec) -> None:
        """Display detailed information for a single recommendation."""
        try:
            # Show reasoning if available
            reasoning = getattr(rec, 'reasoning', None)
            if reasoning:
                primary_reason = getattr(reasoning, 'primary_reason', None)
                if primary_reason:
                    print(f"    💡 {primary_reason}")
                
                # Show additional reasons if available
                if hasattr(reasoning, 'contributing_factors'):
                    factors = reasoning.contributing_factors[:2]  # Top 2 factors
                    for factor in factors:
                        print(f"    • {factor}")
            
            # Show expected performance
            expected_perf = getattr(rec, 'expected_performance', None)
            if expected_perf:
                exp_wr = getattr(expected_perf, 'win_rate', 0)
                exp_kda = getattr(expected_perf, 'avg_kda', 0)
                print(f"    📈 Expected: {exp_wr:.1%} WR | {exp_kda:.2f} KDA")
            
            # Show historical performance
            hist_perf = getattr(rec, 'historical_performance', None)
            if hist_perf:
                hist_wr = getattr(hist_perf, 'win_rate', 0)
                hist_games = getattr(hist_perf, 'games_played', 0)
                print(f"    📊 Historical: {hist_wr:.1%} WR | {hist_games} games")
            
            # Show synergy analysis if available
            synergy = getattr(rec, 'synergy_analysis', None)
            if synergy:
                synergy_score = getattr(synergy, 'overall_synergy_score', 0)
                if synergy_score != 0:
                    synergy_indicator = "🤝" if synergy_score > 0 else "⚠️"
                    print(f"    {synergy_indicator} Team Synergy: {synergy_score:+.2f}")
            
            print()
            
        except Exception as e:
            print(f"    ❌ Details unavailable: {str(e)[:30]}...")
            print()
    
    def _display_mastery_recommendations(self, player, role) -> None:
        """Display mastery-based recommendations as fallback."""
        if not player.champion_masteries:
            print("No champion mastery data available.")
            return
        
        # Filter champions by role
        role_champions = []
        for champ_id, mastery in player.champion_masteries.items():
            if role in mastery.primary_roles or not mastery.primary_roles:  # Include flexible champions
                role_champions.append(mastery)
        
        if not role_champions:
            print(f"No champions found for {role} role.")
            return
        
        # Sort by mastery level and points
        role_champions.sort(key=lambda x: (x.mastery_level, x.mastery_points), reverse=True)
        
        print(f"🏆 Top {role.capitalize()} Champions by Mastery:")
        for i, mastery in enumerate(role_champions[:5], 1):
            level_indicator = "🥇" if mastery.mastery_level == 7 else "🥈" if mastery.mastery_level >= 5 else "🥉"
            print(f"{i}. {mastery.champion_name:15} {level_indicator} L{mastery.mastery_level} | {mastery.mastery_points:,} pts")
    
    def _champion_synergy_interface(self) -> None:
        """Champion synergy analysis interface."""
        print("\n🤝 Champion Synergy Analysis")
        print("-" * 30)
        
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players available.")
            return
        
        print("Synergy Analysis Options:")
        print("1. 🎯 Champion Pair Synergies")
        print("2. 🏆 Team Composition Synergies")
        print("3. ⚠️ Anti-Synergy Warnings")
        print("4. 📊 Player Champion Synergy Matrix")
        print("5. 🏠 Back to Analytics Menu")
        
        choice = input("\nEnter your choice (1-5): ").strip()
        
        if choice == "1":
            self._champion_pair_synergies()
        elif choice == "2":
            self._team_composition_synergies()
        elif choice == "3":
            self._anti_synergy_warnings()
        elif choice == "4":
            self._player_champion_synergy_matrix()
        elif choice == "5":
            return
        else:
            print("Invalid choice. Please enter a number between 1-5.")
    
    def _champion_pair_synergies(self) -> None:
        """Display champion pair synergy analysis."""
        print("\n🎯 Champion Pair Synergies")
        print("-" * 25)
        
        try:
            # Check if champion synergy analyzer is available
            if not hasattr(self.engine, 'champion_synergy_analyzer') or not self.engine.champion_synergy_analyzer:
                print("⚠️ Champion synergy analyzer not available. Using fallback analysis.")
                self._fallback_champion_synergies()
                return
            
            # Get synergy data
            synergy_data = self.engine.champion_synergy_analyzer.get_champion_synergies()
            
            if not synergy_data:
                print("No synergy data available.")
                return
            
            # Display top synergistic pairs
            print("🏆 Top Champion Synergies:")
            print("-" * 40)
            
            sorted_synergies = sorted(synergy_data.items(), key=lambda x: x[1], reverse=True)
            
            for i, ((champ1_id, champ2_id), synergy_score) in enumerate(sorted_synergies[:10], 1):
                champ1_name = self.engine.champion_data_manager.get_champion_name(champ1_id)
                champ2_name = self.engine.champion_data_manager.get_champion_name(champ2_id)
                
                strength = "Strong" if synergy_score >= 0.7 else "Moderate" if synergy_score >= 0.4 else "Weak"
                indicator = "🟢" if synergy_score >= 0.7 else "🟡" if synergy_score >= 0.4 else "🔴"
                
                print(f"{i:2}. {champ1_name} + {champ2_name} {indicator} {synergy_score:.3f} ({strength})")
                
        except Exception as e:
            self.logger.error(f"Champion pair synergy analysis failed: {e}")
            print(f"❌ Analysis failed: {e}")
            self._fallback_champion_synergies()
    
    def _team_composition_synergies(self) -> None:
        """Display team composition synergy analysis."""
        print("\n🏆 Team Composition Synergies")
        print("-" * 30)
        
        try:
            # Check if team composition analyzer is available
            if not hasattr(self.engine, 'team_composition_analyzer') or not self.engine.team_composition_analyzer:
                print("⚠️ Team composition analyzer not available.")
                print("This feature requires historical match data and composition analysis.")
                return
            
            # Get recent successful compositions
            compositions = self.engine.team_composition_analyzer.get_top_compositions(limit=5)
            
            if not compositions:
                print("No composition data available.")
                return
            
            print("🏆 Top Performing Team Compositions:")
            print("-" * 45)
            
            for i, comp in enumerate(compositions, 1):
                print(f"\n{i}. Composition Score: {comp.performance_score:.2f}")
                print(f"   Win Rate: {comp.win_rate:.1%} ({comp.games_played} games)")
                
                # Display role assignments
                for role, champion_info in comp.composition.items():
                    print(f"   {role.capitalize():8}: {champion_info['champion_name']}")
                
                # Show synergy highlights
                if hasattr(comp, 'synergy_highlights') and comp.synergy_highlights:
                    print(f"   🤝 Key Synergies: {', '.join(comp.synergy_highlights[:2])}")
                
        except Exception as e:
            self.logger.error(f"Team composition synergy analysis failed: {e}")
            print(f"❌ Analysis failed: {e}")
    
    def _anti_synergy_warnings(self) -> None:
        """Display anti-synergy warnings."""
        print("\n⚠️ Anti-Synergy Warnings")
        print("-" * 25)
        
        try:
            # Check if synergy analyzer is available
            if not hasattr(self.engine, 'champion_synergy_analyzer') or not self.engine.champion_synergy_analyzer:
                print("⚠️ Champion synergy analyzer not available.")
                self._fallback_anti_synergies()
                return
            
            # Get negative synergies
            anti_synergies = self.engine.champion_synergy_analyzer.get_anti_synergies()
            
            if not anti_synergies:
                print("No anti-synergy data available.")
                return
            
            print("⚠️ Champion Combinations to Avoid:")
            print("-" * 40)
            
            for i, ((champ1_id, champ2_id), anti_synergy_score) in enumerate(anti_synergies[:10], 1):
                champ1_name = self.engine.champion_data_manager.get_champion_name(champ1_id)
                champ2_name = self.engine.champion_data_manager.get_champion_name(champ2_id)
                
                severity = "High" if anti_synergy_score <= -0.5 else "Moderate" if anti_synergy_score <= -0.2 else "Low"
                indicator = "🔴" if anti_synergy_score <= -0.5 else "🟡" if anti_synergy_score <= -0.2 else "🟠"
                
                print(f"{i:2}. {champ1_name} + {champ2_name} {indicator} {anti_synergy_score:.3f} ({severity} conflict)")
                
        except Exception as e:
            self.logger.error(f"Anti-synergy analysis failed: {e}")
            print(f"❌ Analysis failed: {e}")
            self._fallback_anti_synergies()
    
    def _player_champion_synergy_matrix(self) -> None:
        """Display player-specific champion synergy matrix."""
        print("\n📊 Player Champion Synergy Matrix")
        print("-" * 35)
        
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players available.")
            return
        
        # Select player
        print("Select a player:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name} ({player.summoner_name})")
        
        try:
            player_idx = int(input("\nEnter player number: ").strip()) - 1
            if not 0 <= player_idx < len(players):
                print("Invalid player selection.")
                return
            
            selected_player = players[player_idx]
        except ValueError:
            print("Invalid input. Please enter a number.")
            return
        
        try:
            print(f"\n🔄 Analyzing champion synergies for {selected_player.name}...")
            
            # Get player's champion pool
            if not selected_player.champion_masteries:
                print("No champion mastery data available for this player.")
                return
            
            # Display synergy matrix for player's top champions
            top_champions = sorted(
                selected_player.champion_masteries.items(),
                key=lambda x: (x[1].mastery_level, x[1].mastery_points),
                reverse=True
            )[:8]  # Top 8 champions
            
            print(f"\n🏆 Champion Synergy Matrix for {selected_player.name}:")
            print("-" * 50)
            
            # Header
            print("Champion".ljust(12), end="")
            for champ_id, mastery in top_champions:
                print(f"{mastery.champion_name[:8]:>8}", end="")
            print()
            
            # Matrix rows
            for i, (champ1_id, mastery1) in enumerate(top_champions):
                print(f"{mastery1.champion_name[:11]:11}", end=" ")
                
                for j, (champ2_id, mastery2) in enumerate(top_champions):
                    if i == j:
                        print("   -   ", end="")
                    else:
                        # Try to get synergy score
                        try:
                            if hasattr(self.engine, 'champion_synergy_analyzer') and self.engine.champion_synergy_analyzer:
                                synergy = self.engine.champion_synergy_analyzer.get_champion_pair_synergy(champ1_id, champ2_id)
                                if synergy is not None:
                                    print(f"{synergy:7.2f}", end="")
                                else:
                                    print("  N/A  ", end="")
                            else:
                                print("  N/A  ", end="")
                        except:
                            print("  N/A  ", end="")
                print()
            
            print("\nLegend: Positive values = Good synergy, Negative values = Poor synergy")
            
        except Exception as e:
            self.logger.error(f"Player champion synergy matrix failed: {e}")
            print(f"❌ Analysis failed: {e}")
    
    def _fallback_champion_synergies(self) -> None:
        """Fallback champion synergy display using basic data."""
        print("\n📊 Fallback: Basic Champion Synergies")
        print("-" * 35)
        
        # Show some common synergistic champion pairs based on roles
        synergy_examples = [
            ("Yasuo", "Malphite", "Ultimate combo synergy"),
            ("Orianna", "Malphite", "Ball delivery synergy"),
            ("Jinx", "Thresh", "ADC-Support synergy"),
            ("Graves", "Twisted Fate", "Jungle-Mid roam synergy"),
            ("Shen", "Twisted Fate", "Global presence synergy")
        ]
        
        print("🏆 Common Champion Synergies:")
        for i, (champ1, champ2, description) in enumerate(synergy_examples, 1):
            print(f"{i}. {champ1} + {champ2}: {description}")
    
    def _fallback_anti_synergies(self) -> None:
        """Fallback anti-synergy display using basic data."""
        print("\n📊 Fallback: Common Anti-Synergies")
        print("-" * 35)
        
        # Show some common problematic combinations
        anti_synergy_examples = [
            ("Multiple AD carries", "Lack of magic damage"),
            ("No crowd control", "Difficulty engaging fights"),
            ("All melee champions", "Vulnerable to poke compositions"),
            ("No tank/frontline", "Team too squishy"),
            ("Multiple scaling champions", "Weak early game")
        ]
        
        print("⚠️ Common Team Composition Issues:")
        for i, (problem, description) in enumerate(anti_synergy_examples, 1):
            print(f"{i}. {problem}: {description}")
    
    def _performance_trend_visualization(self) -> None:
        """Performance trend visualization in text format."""
        print("\n📈 Performance Trend Visualization")
        print("-" * 35)
        
        # Select player
        players = self.engine.data_manager.load_player_data()
        if not players:
            print("No players available.")
            return
        
        print("Select a player:")
        for i, player in enumerate(players, 1):
            print(f"{i}. {player.name} ({player.summoner_name})")
        
        try:
            player_idx = int(input("\nEnter player number: ").strip()) - 1
            if not 0 <= player_idx < len(players):
                print("Invalid player selection.")
                return
            
            selected_player = players[player_idx]
        except ValueError:
            print("Invalid input. Please enter a number.")
            return
        
        # Select trend type
        print("\nSelect trend analysis type:")
        print("1. 📊 Overall Performance Trend")
        print("2. 🏆 Champion-Specific Trends")
        print("3. 🎭 Role-Specific Trends")
        print("4. 📈 Recent Form Analysis")
        
        try:
            trend_choice = int(input("\nEnter choice (1-4): ").strip())
            if not 1 <= trend_choice <= 4:
                print("Invalid choice.")
                return
        except ValueError:
            print("Invalid input. Please enter a number.")
            return
        
        try:
            print(f"\n🔄 Analyzing performance trends for {selected_player.name}...")
            
            # Check if analytics engine is available
            if not hasattr(self.engine, 'historical_analytics_engine') or not self.engine.historical_analytics_engine:
                print("⚠️ Historical analytics engine not available. Using fallback analysis.")
                self._fallback_trend_analysis(selected_player, trend_choice)
                return
            
            # Get trend analysis based on choice
            if trend_choice == 1:
                self._overall_performance_trend(selected_player)
            elif trend_choice == 2:
                self._champion_specific_trends(selected_player)
            elif trend_choice == 3:
                self._role_specific_trends(selected_player)
            elif trend_choice == 4:
                self._recent_form_analysis(selected_player)
                
        except Exception as e:
            self.logger.error(f"Trend analysis failed: {e}")
            print(f"❌ Trend analysis failed: {e}")
            print("Using fallback analysis...")
            self._fallback_trend_analysis(selected_player, trend_choice)
    
    def _overall_performance_trend(self, player) -> None:
        """Display overall performance trend."""
        from datetime import datetime, timedelta
        
        # Analyze last 6 months with monthly windows
        end_date = datetime.now()
        start_date = end_date - timedelta(days=180)
        
        trend_analysis = self.engine.historical_analytics_engine.calculate_performance_trends(
            player.puuid, time_window=30  # 30-day windows
        )
        
        # Display trend visualization
        self._display_performance_trends(player, trend_analysis)
    
    def _champion_specific_trends(self, player) -> None:
        """Display champion-specific performance trends."""
        if not player.champion_masteries:
            print("No champion mastery data available.")
            return
        
        # Get top 3 champions
        top_champions = sorted(
            player.champion_masteries.items(),
            key=lambda x: (x[1].mastery_level, x[1].mastery_points),
            reverse=True
        )[:3]
        
        print(f"\n🏆 Champion-Specific Trends for {player.name}:")
        print("-" * 45)
        
        for champ_id, mastery in top_champions:
            try:
                # Get champion-specific analytics
                from .analytics_models import AnalyticsFilters, DateRange
                from datetime import datetime, timedelta
                
                end_date = datetime.now()
                start_date = end_date - timedelta(days=90)
                filters = AnalyticsFilters(
                    date_range=DateRange(start_date=start_date, end_date=end_date),
                    champions=[champ_id],
                    min_games=1
                )
                
                champion_analytics = self.engine.historical_analytics_engine.analyze_champion_performance(
                    player.puuid, champ_id, None
                )
                
                if champion_analytics:
                    print(f"\n📊 {mastery.champion_name}:")
                    print(f"   Win Rate: {champion_analytics.win_rate:.1%}")
                    print(f"   Games Played: {champion_analytics.games_played}")
                    print(f"   Average KDA: {champion_analytics.avg_kda:.2f}")
                    
                    # Show recent form if available
                    if hasattr(champion_analytics, 'recent_form') and champion_analytics.recent_form:
                        recent = champion_analytics.recent_form
                        trend_indicator = "📈" if recent.trend_direction > 0 else "📉" if recent.trend_direction < 0 else "➡️"
                        print(f"   Recent Form: {trend_indicator} {recent.recent_win_rate:.1%} (last {recent.recent_games} games)")
                else:
                    print(f"\n📊 {mastery.champion_name}: No recent match data")
                    
            except Exception as e:
                print(f"\n📊 {mastery.champion_name}: Analysis failed ({str(e)[:50]}...)")
    
    def _role_specific_trends(self, player) -> None:
        """Display role-specific performance trends."""
        print(f"\n🎭 Role-Specific Trends for {player.name}:")
        print("-" * 40)
        
        roles = ["top", "jungle", "middle", "support", "bottom"]
        
        for role in roles:
            try:
                # Get role-specific analytics
                from .analytics_models import AnalyticsFilters, DateRange
                from datetime import datetime, timedelta
                
                end_date = datetime.now()
                start_date = end_date - timedelta(days=90)
                filters = AnalyticsFilters(
                    date_range=DateRange(start_date=start_date, end_date=end_date),
                    roles=[role],
                    min_games=1
                )
                
                role_analytics = self.engine.historical_analytics_engine.analyze_player_performance(
                    player.puuid, filters
                )
                
                if role_analytics and role in role_analytics.role_performance:
                    role_perf = role_analytics.role_performance[role]
                    print(f"\n📊 {role.capitalize()}:")
                    print(f"   Win Rate: {role_perf.win_rate:.1%}")
                    print(f"   Games Played: {role_perf.games_played}")
                    print(f"   Average KDA: {role_perf.avg_kda:.2f}")
                    
                    # Show preference alignment
                    if hasattr(player, 'role_preferences') and player.role_preferences:
                        preference = player.role_preferences.get(role, 3)
                        pref_indicator = "⭐" if preference >= 4 else "👍" if preference >= 3 else "👎"
                        print(f"   Preference: {pref_indicator} {preference}/5")
                else:
                    print(f"\n📊 {role.capitalize()}: No recent match data")
                    
            except Exception as e:
                print(f"\n📊 {role.capitalize()}: Analysis failed")
    
    def _recent_form_analysis(self, player) -> None:
        """Display recent form analysis."""
        print(f"\n📈 Recent Form Analysis for {player.name}:")
        print("-" * 40)
        
        try:
            # Get recent performance data
            from .analytics_models import AnalyticsFilters, DateRange
            from datetime import datetime, timedelta
            
            # Last 30 days
            end_date = datetime.now()
            start_date = end_date - timedelta(days=30)
            recent_filters = AnalyticsFilters(
                date_range=DateRange(start_date=start_date, end_date=end_date),
                min_games=1
            )
            
            # Last 90 days for comparison
            comparison_start = end_date - timedelta(days=90)
            comparison_filters = AnalyticsFilters(
                date_range=DateRange(start_date=comparison_start, end_date=start_date),
                min_games=1
            )
            
            recent_analytics = self.engine.historical_analytics_engine.analyze_player_performance(
                player.puuid, recent_filters
            )
            
            comparison_analytics = self.engine.historical_analytics_engine.analyze_player_performance(
                player.puuid, comparison_filters
            )
            
            if recent_analytics and comparison_analytics:
                recent_wr = recent_analytics.overall_performance.win_rate
                comparison_wr = comparison_analytics.overall_performance.win_rate
                
                trend = recent_wr - comparison_wr
                trend_indicator = "📈" if trend > 0.05 else "📉" if trend < -0.05 else "➡️"
                
                print(f"📊 Performance Comparison:")
                print(f"   Last 30 days: {recent_wr:.1%} WR ({recent_analytics.overall_performance.games_played} games)")
                print(f"   Previous 60 days: {comparison_wr:.1%} WR ({comparison_analytics.overall_performance.games_played} games)")
                print(f"   Trend: {trend_indicator} {trend:+.1%} change")
                
                # Show recent champion performance
                if recent_analytics.champion_performance:
                    print(f"\n🏆 Recent Champion Performance:")
                    sorted_champs = sorted(
                        recent_analytics.champion_performance.items(),
                        key=lambda x: x[1].games_played,
                        reverse=True
                    )
                    
                    for champ_id, perf in sorted_champs[:5]:
                        print(f"   {perf.champion_name}: {perf.win_rate:.1%} WR ({perf.games_played} games)")
            else:
                print("Insufficient recent match data for form analysis.")
                
        except Exception as e:
            print(f"❌ Recent form analysis failed: {e}")
    
    def _fallback_trend_analysis(self, player, trend_choice) -> None:
        """Fallback trend analysis using basic data."""
        print(f"\n📊 Fallback: Basic Trend Analysis for {player.name}")
        print("-" * 45)
        
        if trend_choice == 1:  # Overall performance
            print("📈 Overall Performance Trend:")
            print("   • Based on champion mastery progression")
            if player.champion_masteries:
                total_mastery = sum(m.mastery_points for m in player.champion_masteries.values())
                high_level_champs = sum(1 for m in player.champion_masteries.values() if m.mastery_level >= 6)
                print(f"   • Total Mastery Points: {total_mastery:,}")
                print(f"   • High-Level Champions: {high_level_champs}")
                print(f"   • Champion Pool Size: {len(player.champion_masteries)}")
            else:
                print("   • No mastery data available")
                
        elif trend_choice == 2:  # Champion-specific
            print("🏆 Champion-Specific Trends:")
            if player.champion_masteries:
                top_champs = sorted(
                    player.champion_masteries.items(),
                    key=lambda x: (x[1].mastery_level, x[1].mastery_points),
                    reverse=True
                )[:5]
                
                for champ_id, mastery in top_champs:
                    level_indicator = "🥇" if mastery.mastery_level == 7 else "🥈" if mastery.mastery_level >= 5 else "🥉"
                    print(f"   {level_indicator} {mastery.champion_name}: Level {mastery.mastery_level} ({mastery.mastery_points:,} pts)")
            else:
                print("   • No champion mastery data available")
                
        elif trend_choice == 3:  # Role-specific
            print("🎭 Role-Specific Trends:")
            if player.role_preferences:
                sorted_roles = sorted(player.role_preferences.items(), key=lambda x: x[1], reverse=True)
                for role, pref in sorted_roles:
                    pref_indicator = "⭐" if pref >= 4 else "👍" if pref >= 3 else "👎"
                    print(f"   {pref_indicator} {role.capitalize()}: Preference {pref}/5")
            else:
                print("   • No role preference data available")
                
        elif trend_choice == 4:  # Recent form
            print("📈 Recent Form Analysis:")
            print("   • Based on available data indicators")
            if hasattr(player, 'performance_cache') and player.performance_cache:
                print("   • Performance data available")
            else:
                print("   • No recent performance data")
            
            if player.champion_masteries:
                recent_activity = any(m.mastery_level >= 6 for m in player.champion_masteries.values())
                print(f"   • Recent Champion Activity: {'High' if recent_activity else 'Moderate'}")
            
        print("\n💡 Tip: Add more match data for detailed trend analysis.")
    
    def _display_performance_trends(self, player, trend_analysis) -> None:
        """Display performance trends in text format."""
        print(f"\n" + "=" * 60)
        print(f"📈 PERFORMANCE TRENDS - {player.name}")
        print("=" * 60)
        
        if not trend_analysis:
            print("No trend data available.")
            return
        
        # Handle different trend analysis formats
        if hasattr(trend_analysis, 'time_series_data') and trend_analysis.time_series_data:
            # Full trend analysis with time series
            print(f"📊 Trend Analysis Summary:")
            print(f"   Overall Trend: {'📈 Improving' if trend_analysis.overall_trend > 0 else '📉 Declining' if trend_analysis.overall_trend < 0 else '➡️ Stable'}")
            print(f"   Trend Strength: {abs(trend_analysis.overall_trend):.3f}")
            print(f"   Data Points: {len(trend_analysis.time_series_data)}")
            
            # Simple ASCII chart
            print(f"\n📈 Win Rate Trend (last {len(trend_analysis.time_series_data)} periods):")
            print("-" * 50)
            
            # Normalize data for display
            values = [point.value for point in trend_analysis.time_series_data]
            if values:
                min_val = min(values)
                max_val = max(values)
                range_val = max_val - min_val if max_val > min_val else 1
                
                for i, point in enumerate(trend_analysis.time_series_data):
                    # Create simple bar chart
                    normalized = (point.value - min_val) / range_val if range_val > 0 else 0.5
                    bar_length = int(normalized * 30)
                    bar = "█" * bar_length + "░" * (30 - bar_length)
                    
                    # Format timestamp if available
                    timestamp_str = ""
                    if hasattr(point, 'timestamp'):
                        timestamp_str = f" ({point.timestamp.strftime('%m/%d')})"
                    
                    print(f"Period {i+1:2}: {bar} {point.value:.1%}{timestamp_str}")
                    
            # Show additional metrics if available
            if hasattr(trend_analysis, 'metric_trends') and trend_analysis.metric_trends:
                print(f"\n📊 Additional Metric Trends:")
                for metric, trend_value in trend_analysis.metric_trends.items():
                    trend_indicator = "📈" if trend_value > 0 else "📉" if trend_value < 0 else "➡️"
                    print(f"   {metric}: {trend_indicator} {trend_value:+.3f}")
                    
        elif hasattr(trend_analysis, 'overall_trend'):
            # Simple trend analysis
            print(f"📊 Trend Summary:")
            print(f"   Overall Trend: {'📈 Improving' if trend_analysis.overall_trend > 0 else '📉 Declining' if trend_analysis.overall_trend < 0 else '➡️ Stable'}")
            print(f"   Trend Strength: {abs(trend_analysis.overall_trend):.3f}")
            
        else:
            # Basic trend display
            print("📊 Basic trend analysis completed.")
            print("   Detailed time series data not available.")
            
        # Show confidence information if available
        if hasattr(trend_analysis, 'confidence_interval'):
            ci = trend_analysis.confidence_interval
            print(f"\n📊 Statistical Confidence:")
            print(f"   Confidence Interval: [{ci.lower_bound:.3f}, {ci.upper_bound:.3f}]")
            print(f"   Confidence Level: {ci.confidence_level:.1%}")
            
        # Show recommendations if available
        if hasattr(trend_analysis, 'recommendations') and trend_analysis.recommendations:
            print(f"\n💡 Recommendations:")
            for rec in trend_analysis.recommendations[:3]:  # Top 3 recommendations
                print(f"   • {rec}")
                
        print(f"\n💡 Tip: Trends are based on available match data. More data improves accuracy.")
    
    def _refresh_analytics_data(self) -> None:
        """Refresh analytics data and caches."""
        print("\n🔄 Refreshing Analytics Data")
        print("-" * 30)
        
        try:
            # Clear analytics caches
            if hasattr(self.engine, 'analytics_cache_manager') and self.engine.analytics_cache_manager:
                self.engine.analytics_cache_manager.clear_all_caches()
                print("✅ Analytics caches cleared")
            
            # Refresh baseline calculations
            if hasattr(self.engine, 'baseline_manager') and self.engine.baseline_manager:
                players = self.engine.data_manager.load_player_data()
                for player in players:
                    self.engine.baseline_manager.invalidate_player_baselines(player.puuid)
                print("✅ Player baselines invalidated for recalculation")
            
            print("✅ Analytics data refresh complete")
            
        except Exception as e:
            self.logger.error(f"Analytics refresh failed: {e}")
            print(f"❌ Refresh failed: {e}")
    
    def _team_synergy_analysis(self) -> None:
        """Comprehensive team synergy analysis with visual representation."""
        print("\n" + "=" * 70)
        print("🤝 TEAM SYNERGY ANALYSIS - Champion & Player Synergies")
        print("=" * 70)
        
        players = self.engine.data_manager.load_player_data()
        
        if len(players) < 2:
            print("Need at least 2 players for synergy analysis.")
            return
        
        # Get synergy database
        try:
            synergy_db = self.engine.synergy_manager.get_synergy_database()
            if not synergy_db:
                print("⚠️ No synergy data available. Some features may be limited.")
                synergy_db = {}
        except Exception as e:
            print(f"⚠️ Error loading synergy data: {e}")
            synergy_db = {}
        
        print("🔄 Analyzing team synergies...")
        
        # Synergy analysis options
        print("\n🔍 Synergy Analysis Options:")
        print("1. 🎯 Current Team Synergy Overview")
        print("2. 🏆 Champion Synergy Matrix")
        print("3. 👥 Player Compatibility Analysis")
        print("4. 🔮 Optimal Synergy Predictions")
        print("5. 📊 Role Synergy Breakdown")
        print("6. 🏠 Back to analysis menu")
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == "1":
            self._current_team_synergy_overview(players, synergy_db)
        elif choice == "2":
            self._champion_synergy_matrix(players, synergy_db)
        elif choice == "3":
            self._player_compatibility_analysis(players)
        elif choice == "4":
            self._optimal_synergy_predictions(players, synergy_db)
        elif choice == "5":
            self._role_synergy_breakdown(players, synergy_db)
        elif choice == "6":
            return
        else:
            print("Invalid choice. Please enter a number between 1-6.")
    
    def _current_team_synergy_overview(self, players: List, synergy_db: Dict[str, Any]) -> None:
        """Display current team synergy overview."""
        print("\n" + "=" * 60)
        print("🎯 CURRENT TEAM SYNERGY OVERVIEW")
        print("=" * 60)
        
        if len(players) < 5:
            print(f"⚠️ Partial team analysis ({len(players)}/5 players)")
            print("Add more players for complete synergy analysis.\n")
        
        # Run a quick optimization to get current best team composition
        success, message, result = self.engine.optimize_team_smart(None, True)
        
        if not success or not result:
            print(f"❌ Could not analyze current team: {message}")
            return
        
        best_assignment = result.best_assignment
        
        print("🏆 Current Best Team Composition:")
        print("-" * 40)
        
        # Display role assignments
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            player_name = best_assignment.assignments.get(role, "Unassigned")
            individual_score = best_assignment.individual_scores.get(player_name, 0)
            
            score_indicator = "🟢" if individual_score >= 4 else "🟡" if individual_score >= 2 else "🔴"
            print(f"   {role.capitalize():8}: {player_name:15} {score_indicator} {individual_score:.2f}")
        
        # Synergy analysis
        print(f"\n🤝 Team Synergy Analysis:")
        print(f"   Overall Team Score: {best_assignment.total_score:.2f}")
        
        if best_assignment.synergy_scores:
            print(f"   Synergy Contributions:")
            
            # Sort synergies by strength
            sorted_synergies = sorted(best_assignment.synergy_scores.items(), 
                                    key=lambda x: x[1], reverse=True)
            
            positive_synergies = [(pair, score) for pair, score in sorted_synergies if score > 0]
            negative_synergies = [(pair, score) for pair, score in sorted_synergies if score < 0]
            
            if positive_synergies:
                print(f"\n   ✅ Positive Synergies:")
                for (player1, player2), score in positive_synergies[:5]:  # Top 5
                    strength = "Strong" if score >= 0.5 else "Moderate" if score >= 0.2 else "Weak"
                    print(f"      {player1} ↔ {player2}: +{score:.2f} ({strength})")
            
            if negative_synergies:
                print(f"\n   ⚠️ Potential Conflicts:")
                for (player1, player2), score in negative_synergies[-3:]:  # Bottom 3
                    print(f"      {player1} ↔ {player2}: {score:.2f}")
        
        # Champion synergy analysis
        if best_assignment.champion_recommendations:
            print(f"\n🏆 Champion Synergy Potential:")
            
            # Analyze potential champion synergies
            role_champions = {}
            for role, recommendations in best_assignment.champion_recommendations.items():
                if recommendations:
                    role_champions[role] = recommendations[0].champion_name  # Top recommendation
            
            if len(role_champions) >= 2:
                print("   Top Champion Combinations:")
                
                # Check synergies between top recommended champions
                synergy_pairs = []
                role_list = list(role_champions.keys())
                
                for i in range(len(role_list)):
                    for j in range(i + 1, len(role_list)):
                        role1, role2 = role_list[i], role_list[j]
                        champ1, champ2 = role_champions[role1], role_champions[role2]
                        
                        # Look up champion synergy in database
                        synergy_key = f"{champ1}_{champ2}"
                        reverse_key = f"{champ2}_{champ1}"
                        
                        synergy_score = synergy_db.get(synergy_key, synergy_db.get(reverse_key, 0))
                        
                        if synergy_score != 0:
                            synergy_pairs.append((champ1, champ2, synergy_score))
                
                if synergy_pairs:
                    synergy_pairs.sort(key=lambda x: x[2], reverse=True)
                    for champ1, champ2, score in synergy_pairs[:3]:
                        indicator = "🟢" if score > 0.3 else "🟡" if score > 0 else "🔴"
                        print(f"      {champ1} + {champ2}: {indicator} {score:.2f}")
                else:
                    print("      No specific champion synergy data available")
        
        # Visual representation
        self._display_synergy_visual(best_assignment.assignments, best_assignment.synergy_scores)
        
        # Improvement suggestions
        print(f"\n💡 Synergy Improvement Suggestions:")
        
        if best_assignment.total_score < 15:  # Assuming max ~20
            print("   • Consider adjusting role assignments for better synergy")
            print("   • Look for players with complementary playstyles")
        
        if not positive_synergies:
            print("   • Focus on building player chemistry through practice")
            print("   • Consider role preference adjustments")
        
        if len(players) < 5:
            print("   • Add more players to complete the team")
    
    def _display_synergy_visual(self, assignments: Dict[str, str], synergy_scores: Dict[tuple, float]) -> None:
        """Display a visual representation of team synergies."""
        print(f"\n📊 Team Synergy Visual:")
        print("=" * 50)
        
        if not synergy_scores:
            print("   No synergy data available for visualization")
            return
        
        # Create a simple text-based visualization
        players = list(assignments.values())
        if len(players) < 2:
            print("   Need at least 2 players for synergy visualization")
            return
        
        print("   Player Synergy Matrix:")
        print("   " + "─" * 40)
        
        # Header
        print(f"   {'':12}", end="")
        for player in players[:4]:  # Limit to 4 for readability
            print(f"{player[:8]:>8}", end="")
        print()
        
        # Matrix rows
        for i, player1 in enumerate(players[:4]):
            print(f"   {player1[:10]:10}  ", end="")
            
            for j, player2 in enumerate(players[:4]):
                if i == j:
                    print("   --   ", end="")
                else:
                    # Look for synergy score
                    synergy = synergy_scores.get((player1, player2), 
                                               synergy_scores.get((player2, player1), 0))
                    
                    if synergy > 0.3:
                        indicator = " 🟢 "
                    elif synergy > 0:
                        indicator = " 🟡 "
                    elif synergy < -0.2:
                        indicator = " 🔴 "
                    else:
                        indicator = " ⚫ "
                    
                    print(f"{indicator:>8}", end="")
            print()
        
        print("\n   Legend: 🟢 Strong Synergy | 🟡 Good Synergy | ⚫ Neutral | 🔴 Conflict")
    
    def _champion_synergy_matrix(self, players: List, synergy_db: Dict[str, Any]) -> None:
        """Display champion synergy matrix analysis."""
        print("\n" + "=" * 60)
        print("🏆 CHAMPION SYNERGY MATRIX")
        print("=" * 60)
        
        # Collect top champions for each player by role
        player_champions = {}
        
        for player in players:
            if not player.champion_masteries:
                continue
                
            player_champions[player.name] = {}
            
            # Get top champions for each role
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                role_champs = []
                
                for champ_id, mastery in player.champion_masteries.items():
                    if role in mastery.primary_roles:
                        role_champs.append((mastery.champion_name, mastery.mastery_level, mastery.mastery_points))
                
                # Sort by mastery level and points
                role_champs.sort(key=lambda x: (x[1], x[2]), reverse=True)
                
                if role_champs:
                    player_champions[player.name][role] = role_champs[:3]  # Top 3 champions
        
        if not player_champions:
            print("No champion data available for synergy analysis.")
            print("Refresh player data to see champion synergies.")
            return
        
        print("🎯 Champion Pool Overview by Role:")
        print("-" * 40)
        
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            print(f"\n{role.upper()} Champions:")
            
            role_champions = []
            for player_name, roles_data in player_champions.items():
                if role in roles_data:
                    for champ_name, level, points in roles_data[role]:
                        role_champions.append((player_name, champ_name, level, points))
            
            if role_champions:
                # Sort by mastery level and points
                role_champions.sort(key=lambda x: (x[2], x[3]), reverse=True)
                
                for player_name, champ_name, level, points in role_champions[:5]:
                    mastery_indicator = "🥇" if level == 7 else "🥈" if level >= 5 else "🥉"
                    print(f"   {champ_name:15} {mastery_indicator} L{level} ({player_name})")
            else:
                print(f"   No champion data available")
        
        # Champion synergy analysis
        print(f"\n🤝 Champion Synergy Analysis:")
        print("-" * 40)
        
        if not synergy_db:
            print("No synergy database available.")
            print("Champion synergy analysis requires synergy data.")
            return
        
        # Find best synergy combinations
        best_synergies = []
        
        # Check synergies between champions from different roles
        roles = ["top", "jungle", "middle", "support", "bottom"]
        
        for i in range(len(roles)):
            for j in range(i + 1, len(roles)):
                role1, role2 = roles[i], roles[j]
                
                # Get champions for each role
                role1_champs = []
                role2_champs = []
                
                for player_name, roles_data in player_champions.items():
                    if role1 in roles_data:
                        role1_champs.extend([champ[0] for champ in roles_data[role1]])
                    if role2 in roles_data:
                        role2_champs.extend([champ[0] for champ in roles_data[role2]])
                
                # Check synergies between role champions
                for champ1 in role1_champs:
                    for champ2 in role2_champs:
                        synergy_key = f"{champ1}_{champ2}"
                        reverse_key = f"{champ2}_{champ1}"
                        
                        synergy_score = synergy_db.get(synergy_key, synergy_db.get(reverse_key, 0))
                        
                        if synergy_score > 0.2:  # Only show meaningful synergies
                            best_synergies.append((champ1, champ2, role1, role2, synergy_score))
        
        if best_synergies:
            # Sort by synergy strength
            best_synergies.sort(key=lambda x: x[4], reverse=True)
            
            print("Top Champion Synergies:")
            for champ1, champ2, role1, role2, score in best_synergies[:10]:
                strength = "Excellent" if score >= 0.5 else "Good" if score >= 0.3 else "Moderate"
                print(f"   {champ1} ({role1}) + {champ2} ({role2}): {score:.2f} ({strength})")
        else:
            print("No significant champion synergies found in current champion pools.")
        
        # Interactive champion lookup
        print(f"\n🔍 Interactive Champion Synergy Lookup:")
        champion_name = input("Enter champion name to see synergies (or press Enter to continue): ").strip()
        
        if champion_name:
            self._champion_synergy_lookup(champion_name, synergy_db, player_champions)
    
    def _champion_synergy_lookup(self, champion_name: str, synergy_db: Dict[str, Any], 
                                player_champions: Dict[str, Dict[str, List]]) -> None:
        """Look up synergies for a specific champion."""
        print(f"\n🏆 {champion_name.upper()} - Synergy Analysis")
        print("-" * 40)
        
        # Find all champions in team pools
        all_team_champions = set()
        for player_data in player_champions.values():
            for role_data in player_data.values():
                for champ_name, _, _ in role_data:
                    all_team_champions.add(champ_name)
        
        # Check synergies with team champions
        champion_synergies = []
        
        for team_champ in all_team_champions:
            if team_champ.lower() == champion_name.lower():
                continue
                
            synergy_key = f"{champion_name}_{team_champ}"
            reverse_key = f"{team_champ}_{champion_name}"
            
            synergy_score = synergy_db.get(synergy_key, synergy_db.get(reverse_key, 0))
            
            if synergy_score != 0:
                champion_synergies.append((team_champ, synergy_score))
        
        if champion_synergies:
            champion_synergies.sort(key=lambda x: x[1], reverse=True)
            
            print(f"Synergies with team champions:")
            for team_champ, score in champion_synergies:
                if score > 0:
                    indicator = "🟢" if score >= 0.3 else "🟡"
                    strength = "Strong" if score >= 0.3 else "Good"
                else:
                    indicator = "🔴"
                    strength = "Conflict"
                
                print(f"   {team_champ:15} {indicator} {score:+.2f} ({strength})")
        else:
            print(f"No synergy data found for {champion_name} with current team champions.")
    
    def _player_compatibility_analysis(self, players: List) -> None:
        """Analyze player compatibility based on preferences and playstyles."""
        print("\n" + "=" * 60)
        print("👥 PLAYER COMPATIBILITY ANALYSIS")
        print("=" * 60)
        
        if len(players) < 2:
            print("Need at least 2 players for compatibility analysis.")
            return
        
        print("🔄 Analyzing player compatibility...")
        
        # Analyze role preference compatibility
        print("\n🎯 Role Preference Compatibility:")
        print("-" * 40)
        
        # Check for role conflicts and complementarity
        role_conflicts = {}
        role_gaps = {}
        
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            high_pref_players = []
            low_pref_players = []
            
            for player in players:
                pref = player.role_preferences.get(role, 3)
                if pref >= 4:
                    high_pref_players.append(player.name)
                elif pref <= 2:
                    low_pref_players.append(player.name)
            
            if len(high_pref_players) > 1:
                role_conflicts[role] = high_pref_players
            elif len(high_pref_players) == 0:
                role_gaps[role] = [p.name for p in players if p.role_preferences.get(role, 3) == 3]
        
        # Display conflicts
        if role_conflicts:
            print("⚠️ Role Preference Conflicts:")
            for role, conflicted_players in role_conflicts.items():
                print(f"   {role.capitalize()}: {', '.join(conflicted_players)} (all prefer this role)")
        
        # Display gaps
        if role_gaps:
            print("❓ Role Coverage Gaps:")
            for role, neutral_players in role_gaps.items():
                print(f"   {role.capitalize()}: No strong preferences (neutral: {', '.join(neutral_players)})")
        
        if not role_conflicts and not role_gaps:
            print("✅ Good role preference distribution - minimal conflicts")
        
        # Player flexibility analysis
        print(f"\n🔄 Player Flexibility Analysis:")
        print("-" * 40)
        
        flexibility_scores = {}
        for player in players:
            # Calculate flexibility based on role preferences
            preferences = list(player.role_preferences.values())
            
            # Flexibility = how many roles they're comfortable with (pref >= 3)
            comfortable_roles = sum(1 for pref in preferences if pref >= 3)
            
            # Specialization = how focused they are (high pref in few roles)
            high_pref_roles = sum(1 for pref in preferences if pref >= 4)
            
            flexibility_scores[player.name] = {
                'comfortable_roles': comfortable_roles,
                'specialized_roles': high_pref_roles,
                'flexibility_type': 'Specialist' if high_pref_roles <= 1 else 'Flexible' if comfortable_roles >= 4 else 'Balanced'
            }
        
        for player_name, scores in flexibility_scores.items():
            flex_type = scores['flexibility_type']
            comfortable = scores['comfortable_roles']
            specialized = scores['specialized_roles']
            
            type_indicator = "🎯" if flex_type == 'Specialist' else "🌐" if flex_type == 'Flexible' else "⚖️"
            
            print(f"   {player_name:15} {type_indicator} {flex_type:10} ({comfortable}/5 comfortable, {specialized} specialized)")
        
        # Team balance analysis
        print(f"\n⚖️ Team Balance Analysis:")
        print("-" * 40)
        
        specialists = sum(1 for scores in flexibility_scores.values() if scores['flexibility_type'] == 'Specialist')
        flexible_players = sum(1 for scores in flexibility_scores.values() if scores['flexibility_type'] == 'Flexible')
        balanced_players = len(players) - specialists - flexible_players
        
        print(f"   Specialists: {specialists} players")
        print(f"   Flexible: {flexible_players} players") 
        print(f"   Balanced: {balanced_players} players")
        
        # Team composition recommendations
        print(f"\n💡 Team Composition Recommendations:")
        
        if specialists > 3:
            print("   ⚠️ Many specialists - may have role assignment conflicts")
            print("   • Consider encouraging some players to expand role comfort")
        elif specialists == 0:
            print("   ⚠️ No clear specialists - may lack role expertise")
            print("   • Consider having players focus on preferred roles")
        else:
            print("   ✅ Good balance of specialists and flexible players")
        
        if flexible_players >= 2:
            print("   ✅ Good flexibility for role adjustments")
        else:
            print("   • More flexible players would improve team adaptability")
        
        # Compatibility matrix
        print(f"\n🤝 Player Compatibility Matrix:")
        print("-" * 40)
        
        if len(players) <= 6:  # Only show matrix for reasonable number of players
            self._display_compatibility_matrix(players, flexibility_scores)
        else:
            print("   Too many players for matrix display - showing summary instead")
            
            # Show most and least compatible pairs
            compatibility_pairs = []
            
            for i in range(len(players)):
                for j in range(i + 1, len(players)):
                    player1, player2 = players[i], players[j]
                    compatibility = self._calculate_player_compatibility(player1, player2)
                    compatibility_pairs.append((player1.name, player2.name, compatibility))
            
            compatibility_pairs.sort(key=lambda x: x[2], reverse=True)
            
            print("   Most Compatible Pairs:")
            for p1, p2, score in compatibility_pairs[:3]:
                indicator = "🟢" if score >= 0.7 else "🟡"
                print(f"      {p1} ↔ {p2}: {indicator} {score:.2f}")
            
            if len(compatibility_pairs) > 3:
                print("   Least Compatible Pairs:")
                for p1, p2, score in compatibility_pairs[-2:]:
                    indicator = "🔴" if score <= 0.3 else "🟡"
                    print(f"      {p1} ↔ {p2}: {indicator} {score:.2f}")
    
    def _display_compatibility_matrix(self, players: List, flexibility_scores: Dict[str, Dict]) -> None:
        """Display a compatibility matrix for players."""
        print("   Compatibility Matrix:")
        print("   " + "─" * 50)
        
        # Header
        print(f"   {'':12}", end="")
        for player in players:
            print(f"{player.name[:8]:>8}", end="")
        print()
        
        # Matrix rows
        for i, player1 in enumerate(players):
            print(f"   {player1.name[:10]:10}  ", end="")
            
            for j, player2 in enumerate(players):
                if i == j:
                    print("   --   ", end="")
                else:
                    compatibility = self._calculate_player_compatibility(player1, player2)
                    
                    if compatibility >= 0.7:
                        indicator = " 🟢 "
                    elif compatibility >= 0.5:
                        indicator = " 🟡 "
                    elif compatibility >= 0.3:
                        indicator = " 🟠 "
                    else:
                        indicator = " 🔴 "
                    
                    print(f"{indicator:>8}", end="")
            print()
        
        print("\n   Legend: 🟢 High Compatibility | 🟡 Good | 🟠 Fair | 🔴 Low")
    
    def _calculate_player_compatibility(self, player1, player2) -> float:
        """Calculate compatibility score between two players."""
        compatibility = 0.0
        
        # Role preference complementarity (avoid conflicts)
        role_conflicts = 0
        role_complements = 0
        
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            pref1 = player1.role_preferences.get(role, 3)
            pref2 = player2.role_preferences.get(role, 3)
            
            # Conflict if both strongly prefer same role
            if pref1 >= 4 and pref2 >= 4:
                role_conflicts += 1
            # Complement if one likes what other dislikes
            elif (pref1 >= 4 and pref2 <= 2) or (pref1 <= 2 and pref2 >= 4):
                role_complements += 1
        
        # Base compatibility from role preferences
        if role_conflicts == 0:
            compatibility += 0.4
        else:
            compatibility -= role_conflicts * 0.2
        
        compatibility += role_complements * 0.1
        
        # Data completeness similarity (players with similar data quality work well together)
        data1 = 1 if (player1.performance_cache or player1.champion_masteries) else 0
        data2 = 1 if (player2.performance_cache or player2.champion_masteries) else 0
        
        if data1 == data2:
            compatibility += 0.2
        
        # Champion pool overlap (some overlap is good, too much may be redundant)
        if player1.champion_masteries and player2.champion_masteries:
            champs1 = set(mastery.champion_name for mastery in player1.champion_masteries.values())
            champs2 = set(mastery.champion_name for mastery in player2.champion_masteries.values())
            
            overlap = len(champs1.intersection(champs2))
            total_unique = len(champs1.union(champs2))
            
            if total_unique > 0:
                overlap_ratio = overlap / total_unique
                # Optimal overlap is around 20-40%
                if 0.2 <= overlap_ratio <= 0.4:
                    compatibility += 0.3
                elif overlap_ratio < 0.1:
                    compatibility += 0.1  # Some diversity is good
                elif overlap_ratio > 0.6:
                    compatibility -= 0.1  # Too much overlap
        
        # Ensure score is between 0 and 1
        return max(0.0, min(1.0, compatibility))
    
    def _optimal_synergy_predictions(self, players: List, synergy_db: Dict[str, Any]) -> None:
        """Predict optimal synergy combinations."""
        print("\n" + "=" * 60)
        print("🔮 OPTIMAL SYNERGY PREDICTIONS")
        print("=" * 60)
        
        if len(players) < 5:
            print(f"⚠️ Partial analysis with {len(players)} players")
            print("Add more players for complete synergy predictions.\n")
        
        print("🔄 Calculating optimal synergy combinations...")
        
        # Run multiple optimization scenarios to find best synergies
        try:
            # Get the best current assignment
            success, message, result = self.engine.optimize_team_smart(None, True)
            
            if not success or not result:
                print(f"❌ Could not generate predictions: {message}")
                return
            
            print("🎯 Predicted Optimal Team Composition:")
            print("-" * 50)
            
            best_assignment = result.best_assignment
            
            # Display predicted assignments with synergy scores
            total_synergy = sum(best_assignment.synergy_scores.values()) if best_assignment.synergy_scores else 0
            
            print(f"   Predicted Team Score: {best_assignment.total_score:.2f}")
            print(f"   Total Synergy Contribution: {total_synergy:+.2f}")
            print()
            
            for role in ["top", "jungle", "middle", "support", "bottom"]:
                player_name = best_assignment.assignments.get(role, "Unassigned")
                individual_score = best_assignment.individual_scores.get(player_name, 0)
                
                # Get champion recommendation
                champion_rec = ""
                if role in best_assignment.champion_recommendations and best_assignment.champion_recommendations[role]:
                    top_champ = best_assignment.champion_recommendations[role][0]
                    champion_rec = f" → {top_champ.champion_name}"
                
                score_indicator = "🟢" if individual_score >= 4 else "🟡" if individual_score >= 2 else "🔴"
                print(f"   {role.capitalize():8}: {player_name:15} {score_indicator} {individual_score:.2f}{champion_rec}")
            
            # Synergy breakdown
            if best_assignment.synergy_scores:
                print(f"\n🤝 Predicted Synergy Breakdown:")
                
                sorted_synergies = sorted(best_assignment.synergy_scores.items(), 
                                        key=lambda x: x[1], reverse=True)
                
                positive_synergies = [(pair, score) for pair, score in sorted_synergies if score > 0.1]
                
                if positive_synergies:
                    print("   Key Synergy Pairs:")
                    for (player1, player2), score in positive_synergies:
                        strength = "Excellent" if score >= 0.5 else "Good" if score >= 0.3 else "Moderate"
                        print(f"      {player1} ↔ {player2}: +{score:.2f} ({strength})")
                
                # Identify potential improvements
                weak_synergies = [(pair, score) for pair, score in sorted_synergies if score < 0]
                if weak_synergies:
                    print(f"\n   ⚠️ Areas for Improvement:")
                    for (player1, player2), score in weak_synergies[-2:]:
                        print(f"      {player1} ↔ {player2}: {score:.2f} (consider role adjustments)")
            
            # Alternative compositions
            if len(result.assignments) > 1:
                print(f"\n🔄 Alternative High-Synergy Compositions:")
                print("-" * 50)
                
                for i, assignment in enumerate(result.assignments[1:3], 2):  # Show 2nd and 3rd best
                    alt_synergy = sum(assignment.synergy_scores.values()) if assignment.synergy_scores else 0
                    
                    print(f"\n   Alternative #{i} (Score: {assignment.total_score:.2f}, Synergy: {alt_synergy:+.2f}):")
                    
                    for role in ["top", "jungle", "middle", "support", "bottom"]:
                        player_name = assignment.assignments.get(role, "Unassigned")
                        individual_score = assignment.individual_scores.get(player_name, 0)
                        print(f"      {role.capitalize():8}: {player_name:15} ({individual_score:.2f})")
            
            # Synergy optimization recommendations
            print(f"\n💡 Synergy Optimization Recommendations:")
            
            if total_synergy < 1.0:
                print("   • Focus on building player chemistry through practice")
                print("   • Consider adjusting role assignments for better synergy")
            
            if len(positive_synergies) < 3:
                print("   • Look for players with complementary champion pools")
                print("   • Practice team compositions with high champion synergy")
            
            if best_assignment.total_score < 15:
                print("   • Consider role preference adjustments")
                print("   • Focus on individual skill development")
            
            # Long-term predictions
            print(f"\n🔮 Long-term Synergy Potential:")
            
            if len(players) >= 7:
                print("   ✅ Good player pool for synergy experimentation")
                print("   • Try different role combinations in practice")
            elif len(players) == 5:
                print("   ⚠️ Limited flexibility with exactly 5 players")
                print("   • Consider adding substitute players")
            else:
                print("   🔴 Need more players for optimal synergy development")
            
        except Exception as e:
            print(f"❌ Error generating synergy predictions: {e}")
            self.logger.error(f"Synergy prediction error: {e}", exc_info=True)
    
    def _role_synergy_breakdown(self, players: List, synergy_db: Dict[str, Any]) -> None:
        """Analyze synergies by role combinations."""
        print("\n" + "=" * 60)
        print("📊 ROLE SYNERGY BREAKDOWN")
        print("=" * 60)
        
        print("🔄 Analyzing role-specific synergies...")
        
        # Define key role synergy pairs
        key_synergy_pairs = [
            ("jungle", "middle", "Jungle-Mid Synergy"),
            ("support", "bottom", "Bot Lane Synergy"),
            ("jungle", "top", "Jungle-Top Synergy"),
            ("middle", "support", "Mid-Support Roaming"),
            ("top", "jungle", "Top-Jungle Coordination")
        ]
        
        print("\n🎯 Key Role Synergy Analysis:")
        print("-" * 40)
        
        for role1, role2, synergy_name in key_synergy_pairs:
            print(f"\n{synergy_name}:")
            
            # Find players suitable for each role
            role1_players = []
            role2_players = []
            
            for player in players:
                role1_pref = player.role_preferences.get(role1, 3)
                role2_pref = player.role_preferences.get(role2, 3)
                
                if role1_pref >= 3:  # At least neutral
                    role1_players.append((player.name, role1_pref))
                if role2_pref >= 3:
                    role2_players.append((player.name, role2_pref))
            
            if not role1_players or not role2_players:
                print(f"   ⚠️ Insufficient players for {role1}-{role2} synergy analysis")
                continue
            
            # Sort by preference
            role1_players.sort(key=lambda x: x[1], reverse=True)
            role2_players.sort(key=lambda x: x[1], reverse=True)
            
            print(f"   {role1.capitalize()} candidates: {', '.join(p[0] for p in role1_players[:3])}")
            print(f"   {role2.capitalize()} candidates: {', '.join(p[0] for p in role2_players[:3])}")
            
            # Analyze potential synergies between top candidates
            best_synergy = 0
            best_pair = None
            
            for p1_name, p1_pref in role1_players[:2]:
                for p2_name, p2_pref in role2_players[:2]:
                    if p1_name == p2_name:
                        continue
                    
                    # Calculate compatibility
                    player1 = next(p for p in players if p.name == p1_name)
                    player2 = next(p for p in players if p.name == p2_name)
                    
                    compatibility = self._calculate_player_compatibility(player1, player2)
                    
                    if compatibility > best_synergy:
                        best_synergy = compatibility
                        best_pair = (p1_name, p2_name)
            
            if best_pair:
                synergy_indicator = "🟢" if best_synergy >= 0.7 else "🟡" if best_synergy >= 0.5 else "🟠"
                print(f"   Best Pairing: {best_pair[0]} + {best_pair[1]} {synergy_indicator} {best_synergy:.2f}")
            else:
                print(f"   No clear synergy pairing identified")
        
        # Overall role coverage synergy
        print(f"\n🌐 Overall Role Coverage Synergy:")
        print("-" * 40)
        
        # Calculate how well roles are covered
        role_coverage_quality = {}
        
        for role in ["top", "jungle", "middle", "support", "bottom"]:
            suitable_players = sum(1 for p in players if p.role_preferences.get(role, 3) >= 3)
            preferred_players = sum(1 for p in players if p.role_preferences.get(role, 3) >= 4)
            
            if preferred_players >= 2:
                quality = "Excellent"
                indicator = "🟢"
            elif preferred_players >= 1 or suitable_players >= 2:
                quality = "Good"
                indicator = "🟡"
            elif suitable_players >= 1:
                quality = "Fair"
                indicator = "🟠"
            else:
                quality = "Poor"
                indicator = "🔴"
            
            role_coverage_quality[role] = (quality, indicator, suitable_players, preferred_players)
            
            print(f"   {role.capitalize():8}: {indicator} {quality:10} ({suitable_players} suitable, {preferred_players} preferred)")
        
        # Synergy recommendations by role
        print(f"\n💡 Role-Specific Synergy Recommendations:")
        print("-" * 50)
        
        # Bot lane synergy
        bot_adc_players = [p.name for p in players if p.role_preferences.get("bottom", 3) >= 3]
        bot_sup_players = [p.name for p in players if p.role_preferences.get("support", 3) >= 3]
        
        if bot_adc_players and bot_sup_players:
            print("   🎯 Bot Lane Focus:")
            print("      • Practice laning phase coordination")
            print("      • Work on vision control synergy")
            print(f"      • Available pairs: {len(bot_adc_players)} ADC × {len(bot_sup_players)} Support")
        
        # Jungle synergy
        jungle_players = [p.name for p in players if p.role_preferences.get("jungle", 3) >= 3]
        if jungle_players:
            print("   🌲 Jungle Synergy:")
            print("      • Focus on jungle-lane coordination")
            print("      • Practice objective control with team")
            if len(jungle_players) > 1:
                print(f"      • Multiple jungle options available: {', '.join(jungle_players)}")
        
        # Mid-game synergy
        mid_players = [p.name for p in players if p.role_preferences.get("middle", 3) >= 3]
        if mid_players:
            print("   ⚡ Mid-Game Synergy:")
            print("      • Practice roaming coordination")
            print("      • Work on team fight positioning")
        
        # Overall team synergy potential
        excellent_roles = sum(1 for quality, _, _, _ in role_coverage_quality.values() if quality == "Excellent")
        good_roles = sum(1 for quality, _, _, _ in role_coverage_quality.values() if quality in ["Excellent", "Good"])
        
        print(f"\n🏆 Overall Team Synergy Potential:")
        
        if excellent_roles >= 3:
            print("   🟢 High synergy potential - excellent role coverage")
        elif good_roles >= 4:
            print("   🟡 Good synergy potential - solid role coverage")
        elif good_roles >= 3:
            print("   🟠 Moderate synergy potential - some gaps to address")
        else:
            print("   🔴 Limited synergy potential - significant role coverage issues")
        
        print(f"   Roles with excellent coverage: {excellent_roles}/5")
        print(f"   Roles with good+ coverage: {good_roles}/5")
    
    def _settings(self) -> None:
        """Settings and maintenance menu."""
        while True:
            print("\n" + "=" * 50)
            print("⚙️ SETTINGS & MAINTENANCE")
            print("=" * 50)
            
            print("1. System diagnostics")
            print("2. Clear cache")
            print("3. Update champion data")
            print("4. API configuration")
            print("5. Export system logs")
            print("6. Back to main menu")
            
            choice = input("\nEnter your choice (1-6): ").strip()
            
            if choice == "1":
                self._system_diagnostics()
            elif choice == "2":
                self._clear_cache()
            elif choice == "3":
                self._update_champion_data()
            elif choice == "4":
                self._api_configuration()
            elif choice == "5":
                self._export_logs()
            elif choice == "6":
                break
            else:
                print("Invalid choice. Please enter a number between 1-6.")
    
    def _system_diagnostics(self) -> None:
        """Display comprehensive system diagnostics."""
        print("\n🔍 Running system diagnostics...")
        
        diagnostics = self.engine.get_system_diagnostics()
        
        print("\n" + "=" * 50)
        print("🔧 SYSTEM DIAGNOSTICS")
        print("=" * 50)
        
        print(f"Timestamp: {diagnostics['timestamp']}")
        
        # Component status
        print("\nComponent Status:")
        for component, status in diagnostics['component_status'].items():
            status_icon = "✅" if status == "OK" else "⚠️" if status == "OFFLINE" else "❌"
            print(f"  {component.capitalize():15}: {status_icon} {status}")
        
        # Performance metrics
        if 'performance_metrics' in diagnostics:
            metrics = diagnostics['performance_metrics']
            if 'error' not in metrics:
                print("\nPerformance Metrics:")
                print(f"  Total Players: {metrics.get('total_players', 0)}")
                print(f"  Complete Data: {metrics.get('players_with_complete_data', 0)}")
                print(f"  Cache Size: {metrics.get('cache_size', 0)} champions")
        
        # Recommendations
        if diagnostics.get('recommendations'):
            print("\nRecommendations:")
            for rec in diagnostics['recommendations']:
                print(f"  💡 {rec}")
        
        # System status
        status = diagnostics['system_status']
        print(f"\nSystem Ready: {'✅ Yes' if status.get('ready_for_optimization') else '❌ No'}")
        if not status.get('ready_for_optimization'):
            print(f"  Need {5 - status.get('player_count', 0)} more players")
    
    def _clear_cache(self) -> None:
        """Clear system cache."""
        print("\n🗑️ Cache Management")
        print("-" * 20)
        
        print("1. Clear API response cache")
        print("2. Clear champion data cache")
        print("3. Clear all caches")
        print("4. Cancel")
        
        choice = input("\nEnter your choice (1-4): ").strip()
        
        if choice == "1":
            # Clear API cache
            try:
                import shutil
                from pathlib import Path
                
                cache_dir = Path(self.engine.config.cache_directory)
                api_cache = cache_dir / "api_responses"
                if api_cache.exists():
                    shutil.rmtree(api_cache)
                    api_cache.mkdir(parents=True)
                    print("✅ API response cache cleared.")
                else:
                    print("ℹ️ API cache directory not found.")
            except Exception as e:
                print(f"❌ Failed to clear API cache: {e}")
        
        elif choice == "2":
            # Clear champion data cache
            try:
                self.engine.champion_data_manager.champions = {}
                print("✅ Champion data cache cleared.")
                print("ℹ️ Champion data will be reloaded on next use.")
            except Exception as e:
                print(f"❌ Failed to clear champion cache: {e}")
        
        elif choice == "3":
            # Clear all caches
            confirm = input("Clear all caches? This will remove all cached data. (y/n): ").strip().lower()
            if confirm == 'y':
                try:
                    import shutil
                    from pathlib import Path
                    
                    cache_dir = Path(self.engine.config.cache_directory)
                    if cache_dir.exists():
                        shutil.rmtree(cache_dir)
                        cache_dir.mkdir(parents=True)
                    
                    self.engine.champion_data_manager.champions = {}
                    print("✅ All caches cleared.")
                except Exception as e:
                    print(f"❌ Failed to clear caches: {e}")
    
    def _update_champion_data(self) -> None:
        """Update champion data from Riot API."""
        print("\n🔄 Updating champion data...")
        
        try:
            # Force refresh champion data
            self.engine.champion_data_manager.fetch_champion_list()
            champion_count = len(self.engine.champion_data_manager.champions)
            print(f"✅ Champion data updated successfully! ({champion_count} champions loaded)")
        except Exception as e:
            print(f"❌ Failed to update champion data: {e}")
    
    def _api_configuration(self) -> None:
        """API configuration and testing."""
        print("\n🔑 API Configuration")
        print("-" * 20)
        
        print(f"Current API Status: {'✅ Available' if self.engine.api_available else '❌ Offline'}")
        
        if self.engine.api_available:
            print("API client is working correctly.")
            
            # Test API connection
            test_api = input("Test API connection? (y/n): ").strip().lower()
            if test_api == 'y':
                print("🔄 Testing API connection...")
                try:
                    # Try to fetch champion data as a test
                    test_result = self.engine.champion_data_manager.fetch_champion_list()
                    if test_result:
                        print("✅ API connection test successful!")
                    else:
                        print("⚠️ API connection test failed.")
                except Exception as e:
                    print(f"❌ API test failed: {e}")
        else:
            print("API client is offline. Check your RIOT_API_KEY environment variable.")
            print("\nTo set up API access:")
            print("1. Get an API key from https://developer.riotgames.com/")
            print("2. Set the RIOT_API_KEY environment variable")
            print("3. Restart the application")
    
    def _export_logs(self) -> None:
        """Export system logs."""
        try:
            import shutil
            from pathlib import Path
            from datetime import datetime
            
            log_dir = Path(self.engine.config.data_directory) / "logs"
            if not log_dir.exists():
                print("No logs found.")
                return
            
            # Create export archive
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_name = f"system_logs_{timestamp}"
            
            shutil.make_archive(export_name, 'zip', log_dir)
            print(f"✅ Logs exported to {export_name}.zip")
            
        except Exception as e:
            print(f"❌ Failed to export logs: {e}")
    
    def _historical_match_extraction(self) -> None:
        """Historical match extraction with tracking."""
        try:
            print("\n" + "="*60)
            print("📜 HISTORICAL MATCH EXTRACTION")
            print("="*60)
            
            # Show current extraction status first
            print("\n📊 Current Extraction Status:")
            status = self.engine.get_extraction_status()
            
            if status['summary']['total_players'] > 0:
                print(f"   Players with extraction data: {status['summary']['total_players']}")
                print(f"   Completed extractions: {status['summary']['completed_players']}")
                print(f"   Total matches extracted: {status['summary']['total_matches_extracted']}")
                print(f"   Completion rate: {status['summary']['completion_rate']:.1%}")
            else:
                print("   No extraction data yet - this will be the first extraction")
            
            print("\n🎯 Extraction Options:")
            print("1. 📜 Extract for All Players")
            print("2. 👤 Extract for Specific Player")
            print("3. 🔄 Continue Incomplete Extractions")
            print("4. 🔙 Back to Player Management")
            
            choice = input("\n🎯 Choose extraction option (1-4): ").strip()
            
            if choice == '1':
                self._extract_all_players()
            elif choice == '2':
                self._extract_specific_player()
            elif choice == '3':
                self._continue_incomplete_extractions()
            elif choice == '4':
                return
            else:
                print("\n❌ Invalid choice. Please try again.")
                
        except Exception as e:
            print(f"\n❌ Error in historical extraction: {e}")
    
    def _extract_all_players(self) -> None:
        """Extract historical matches for all players."""
        try:
            print("\n📜 Historical Extraction for All Players")
            print("-" * 50)
            
            # Get extraction parameters
            max_matches = input("\n🎯 Maximum matches per player (default 300): ").strip()
            if not max_matches:
                max_matches = 300
            else:
                try:
                    max_matches = int(max_matches)
                    if max_matches <= 0:
                        max_matches = 300
                except ValueError:
                    max_matches = 300
            
            force_restart = input("\n🔄 Force restart extraction from beginning? (y/N): ").strip().lower() == 'y'
            
            print(f"\n🚀 Starting historical extraction (max {max_matches} matches per player)...")
            if force_restart:
                print("⚠️  Restarting extraction from the beginning")
            
            # Perform extraction
            results = self.engine.historical_match_scraping(
                player_names=None,  # All players
                max_matches_per_player=max_matches,
                force_restart=force_restart
            )
            
            # Display results
            self._display_extraction_results(results)
            
        except Exception as e:
            print(f"\n❌ Error in all-player extraction: {e}")
    
    def _extract_specific_player(self) -> None:
        """Extract historical matches for a specific player."""
        try:
            players = self.engine.data_manager.load_player_data()
            if not players:
                print("\n❌ No players found. Add players first.")
                return
            
            print("\n👥 Available Players:")
            for i, player in enumerate(players, 1):
                puuid_status = "✅ Has PUUID" if player.puuid else "❌ No PUUID"
                print(f"   {i}. {player.name} ({puuid_status})")
            
            choice = input(f"\n🎯 Select player (1-{len(players)}): ").strip()
            
            try:
                player_index = int(choice) - 1
                if 0 <= player_index < len(players):
                    selected_player = players[player_index]
                    
                    if not selected_player.puuid:
                        print(f"\n❌ {selected_player.name} has no PUUID. Cannot extract matches.")
                        return
                    
                    # Get extraction parameters
                    max_matches = input("\n🎯 Maximum matches to extract (default 300): ").strip()
                    if not max_matches:
                        max_matches = 300
                    else:
                        try:
                            max_matches = int(max_matches)
                            if max_matches <= 0:
                                max_matches = 300
                        except ValueError:
                            max_matches = 300
                    
                    force_restart = input("\n🔄 Force restart extraction from beginning? (y/N): ").strip().lower() == 'y'
                    
                    print(f"\n🚀 Starting historical extraction for {selected_player.name}...")
                    
                    # Perform extraction
                    results = self.engine.historical_match_scraping(
                        player_names=[selected_player.name],
                        max_matches_per_player=max_matches,
                        force_restart=force_restart
                    )
                    
                    # Display results
                    self._display_extraction_results(results)
                    
                else:
                    print("\n❌ Invalid player selection.")
            except ValueError:
                print("\n❌ Invalid input. Please enter a number.")
                
        except Exception as e:
            print(f"\n❌ Error in specific player extraction: {e}")
    
    def _continue_incomplete_extractions(self) -> None:
        """Continue incomplete extractions."""
        try:
            print("\n🔄 Continue Incomplete Extractions")
            print("-" * 40)
            
            # Get extraction status
            status = self.engine.get_extraction_status()
            
            if status['summary']['total_players'] == 0:
                print("\n❌ No extraction data found. Start a new extraction first.")
                return
            
            # Find incomplete extractions
            incomplete_players = []
            for player_name, details in status['players'].items():
                if not details['extraction_complete']:
                    incomplete_players.append(player_name)
            
            if not incomplete_players:
                print("\n✅ All extractions are complete!")
                return
            
            print(f"\n📊 Found {len(incomplete_players)} incomplete extractions:")
            for player_name in incomplete_players:
                details = status['players'][player_name]
                print(f"   • {player_name}: {details['matches_extracted']} matches ({details['extraction_progress']:.1%} complete)")
            
            # Get continuation parameters
            max_matches = input("\n🎯 Maximum additional matches per player (default 100): ").strip()
            if not max_matches:
                max_matches = 100
            else:
                try:
                    max_matches = int(max_matches)
                    if max_matches <= 0:
                        max_matches = 100
                except ValueError:
                    max_matches = 100
            
            print(f"\n🚀 Continuing extraction for {len(incomplete_players)} players...")
            
            # Continue extraction
            results = self.engine.historical_match_scraping(
                player_names=incomplete_players,
                max_matches_per_player=max_matches,
                force_restart=False
            )
            
            # Display results
            self._display_extraction_results(results)
            
        except Exception as e:
            print(f"\n❌ Error continuing extractions: {e}")
    
    def _display_extraction_results(self, results: dict) -> None:
        """Display extraction results in a formatted way."""
        print("\n" + "="*60)
        print("📊 EXTRACTION RESULTS")
        print("="*60)
        
        print(f"\n📈 Summary:")
        print(f"   Players processed: {results.get('players_processed', 0)}")
        print(f"   New matches stored: {results.get('total_new_matches', 0)}")
        print(f"   Duplicates skipped: {results.get('total_duplicates_skipped', 0)}")
        print(f"   Duration: {results.get('duration_minutes', 0):.2f} minutes")
        
        if 'player_results' in results and results['player_results']:
            print(f"\n👥 Player Results:")
            for player, stats in results['player_results'].items():
                status_icon = "✅" if stats.get('extraction_complete', False) else "🔄"
                print(f"   {status_icon} {player}:")
                print(f"      Status: {stats.get('status', 'unknown')}")
                print(f"      New matches: {stats.get('new_matches_stored', 0)}")
                print(f"      Total extracted: {stats.get('total_matches_extracted', 0)}")
                if 'extraction_progress' in stats:
                    print(f"      Progress: {stats['extraction_progress']:.1%}")
        
        if 'errors' in results and results['errors']:
            print(f"\n❌ Errors:")
            for error in results['errors']:
                print(f"   • {error}")
        
        input("\n📋 Press Enter to continue...")
    
    def _view_extraction_status(self) -> None:
        """View detailed extraction status for all players."""
        try:
            print("\n" + "="*60)
            print("📊 EXTRACTION STATUS")
            print("="*60)
            
            status = self.engine.get_extraction_status()
            
            # Summary
            print(f"\n📈 Summary:")
            print(f"   Total players: {status['summary']['total_players']}")
            print(f"   Completed players: {status['summary']['completed_players']}")
            print(f"   Completion rate: {status['summary']['completion_rate']:.1%}")
            print(f"   Total matches extracted: {status['summary']['total_matches_extracted']}")
            
            if status['summary']['total_players'] > 0:
                print(f"\n👥 Player Details:")
                for player_name, details in status['players'].items():
                    status_icon = "✅" if details['extraction_complete'] else "🔄"
                    print(f"\n   {status_icon} {player_name}:")
                    print(f"      Matches extracted: {details['matches_extracted']}")
                    print(f"      Extraction range: {details['start_index']} - {details['end_index']}")
                    print(f"      Next extraction start: {details['next_extraction_start']}")
                    print(f"      Extraction complete: {details['extraction_complete']}")
                    if details['last_extraction']:
                        print(f"      Last extraction: {details['last_extraction']}")
                    if details['total_matches_available']:
                        print(f"      Total available: {details['total_matches_available']}")
                        print(f"      Progress: {details['extraction_progress']:.1%}")
            else:
                print("\n❌ No extraction data found. Start an extraction to see status.")
            
            # Options
            print("\n🎯 Options:")
            print("1. 🔄 Reset Player Extraction")
            print("2. 🔙 Back to Player Management")
            
            choice = input("\n🎯 Choose option (1-2): ").strip()
            
            if choice == '1':
                self._reset_player_extraction()
            elif choice == '2':
                return
            else:
                print("\n❌ Invalid choice.")
                
        except Exception as e:
            print(f"\n❌ Error viewing extraction status: {e}")
    
    def _reset_player_extraction(self) -> None:
        """Reset extraction progress for a specific player."""
        try:
            status = self.engine.get_extraction_status()
            
            if status['summary']['total_players'] == 0:
                print("\n❌ No extraction data found to reset.")
                return
            
            print("\n🔄 Reset Player Extraction")
            print("-" * 30)
            
            players_list = list(status['players'].keys())
            print("\n👥 Players with extraction data:")
            for i, player_name in enumerate(players_list, 1):
                details = status['players'][player_name]
                status_icon = "✅" if details['extraction_complete'] else "🔄"
                print(f"   {i}. {status_icon} {player_name} ({details['matches_extracted']} matches)")
            
            choice = input(f"\n🎯 Select player to reset (1-{len(players_list)}): ").strip()
            
            try:
                player_index = int(choice) - 1
                if 0 <= player_index < len(players_list):
                    selected_player = players_list[player_index]
                    
                    confirm = input(f"\n⚠️  Reset extraction for {selected_player}? This will clear all extraction progress. (y/N): ").strip().lower()
                    
                    if confirm == 'y':
                        result = self.engine.reset_player_extraction(selected_player)
                        
                        if result.get('success'):
                            print(f"\n✅ Extraction reset successful for {selected_player}")
                        else:
                            print(f"\n❌ Reset failed: {result.get('error', 'Unknown error')}")
                    else:
                        print("\n❌ Reset cancelled.")
                else:
                    print("\n❌ Invalid player selection.")
            except ValueError:
                print("\n❌ Invalid input. Please enter a number.")
                
        except Exception as e:
            print(f"\n❌ Error resetting extraction: {e}")

    def _historical_match_browser(self) -> None:
        """Historical match viewing and filtering interface."""
        print("\n" + "=" * 60)
        print("📜 HISTORICAL MATCH BROWSER")
        print("=" * 60)
        
        # Check if we have any match data
        try:
            match_manager = self.engine.match_manager
            total_matches = len(match_manager._matches_cache)
            
            if total_matches == 0:
                print("\n❌ No historical match data found.")
                print("💡 Use the historical match extraction feature first to collect match data.")
                return
            
            print(f"\n📊 Database Status: {total_matches} matches available")
            
            # Initialize filters
            current_filters = {
                'date_range': None,
                'champions': None,
                'roles': None,
                'outcome': None,
                'players': None
            }
            
            while True:
                self._display_match_browser_menu(current_filters, total_matches)
                choice = input("\nEnter your choice (1-8): ").strip()
                
                if choice == "1":
                    self._view_match_list(current_filters)
                elif choice == "2":
                    current_filters = self._set_match_filters(current_filters)
                elif choice == "3":
                    self._search_matches(current_filters)
                elif choice == "4":
                    self._view_match_details()
                elif choice == "5":
                    self._export_matches(current_filters)
                elif choice == "6":
                    self._match_statistics(current_filters)
                elif choice == "7":
                    current_filters = self._clear_filters()
                elif choice == "8":
                    break
                else:
                    print("\n❌ Invalid choice. Please enter a number between 1-8.")
                
                if choice != "8":
                    input("\nPress Enter to continue...")
                    
        except Exception as e:
            self.logger.error(f"Error in historical match browser: {e}", exc_info=True)
            print(f"\n❌ Error accessing match data: {e}")

    def _display_match_browser_menu(self, filters: dict, total_matches: int) -> None:
        """Display the match browser menu with current filter status."""
        print("\n" + "=" * 50)
        print("MATCH BROWSER MENU")
        print("=" * 50)
        
        # Show active filters
        active_filters = []
        if filters['date_range']:
            active_filters.append(f"Date: {filters['date_range']['start']} to {filters['date_range']['end']}")
        if filters['champions']:
            active_filters.append(f"Champions: {len(filters['champions'])} selected")
        if filters['roles']:
            active_filters.append(f"Roles: {', '.join(filters['roles'])}")
        if filters['outcome'] is not None:
            active_filters.append(f"Outcome: {'Wins' if filters['outcome'] else 'Losses'}")
        if filters['players']:
            active_filters.append(f"Players: {len(filters['players'])} selected")
        
        if active_filters:
            print(f"\n🔍 Active Filters: {' | '.join(active_filters)}")
        else:
            print(f"\n🔍 No filters active - showing all {total_matches} matches")
        
        print("\n1. 📋 View Match List")
        print("2. 🔍 Set Filters")
        print("3. 🔎 Search Matches")
        print("4. 📄 View Match Details")
        print("5. 📤 Export Matches")
        print("6. 📊 Match Statistics")
        print("7. 🗑️ Clear All Filters")
        print("8. ⬅️ Back to Main Menu")

    def _view_match_list(self, filters: dict) -> None:
        """Display a paginated list of matches with current filters."""
        print("\n" + "=" * 70)
        print("📋 MATCH LIST")
        print("=" * 70)
        
        try:
            # Get filtered matches
            matches = self._get_filtered_matches(filters)
            
            if not matches:
                print("\n❌ No matches found with current filters.")
                return
            
            # Sort matches by date (newest first)
            matches.sort(key=lambda m: m.game_creation, reverse=True)
            
            # Pagination
            page_size = 20
            total_pages = (len(matches) + page_size - 1) // page_size
            current_page = 1
            
            while True:
                start_idx = (current_page - 1) * page_size
                end_idx = min(start_idx + page_size, len(matches))
                page_matches = matches[start_idx:end_idx]
                
                print(f"\n📄 Page {current_page}/{total_pages} - Showing {start_idx + 1}-{end_idx} of {len(matches)} matches")
                print("-" * 70)
                print(f"{'#':<3} {'Date':<12} {'Duration':<8} {'Queue':<12} {'Result':<6} {'Players':<15}")
                print("-" * 70)
                
                for i, match in enumerate(page_matches, start_idx + 1):
                    # Get match info
                    date_str = match.game_creation_datetime.strftime("%Y-%m-%d")
                    duration_str = f"{match.game_duration // 60}:{match.game_duration % 60:02d}"
                    queue_name = self._get_queue_name(match.queue_id)
                    
                    # Get known players in this match
                    known_players = self._get_known_players_in_match(match)
                    players_str = f"{len(known_players)} known"
                    
                    # Determine result for known players
                    result_str = self._get_match_result_summary(match, known_players)
                    
                    print(f"{i:<3} {date_str:<12} {duration_str:<8} {queue_name:<12} {result_str:<6} {players_str:<15}")
                
                # Navigation options
                print("\n" + "-" * 70)
                nav_options = []
                if current_page > 1:
                    nav_options.append("p=Previous")
                if current_page < total_pages:
                    nav_options.append("n=Next")
                nav_options.extend(["d=Details", "b=Back"])
                
                print(f"Navigation: {' | '.join(nav_options)}")
                
                nav_choice = input("\nEnter choice: ").strip().lower()
                
                if nav_choice == 'p' and current_page > 1:
                    current_page -= 1
                elif nav_choice == 'n' and current_page < total_pages:
                    current_page += 1
                elif nav_choice == 'd':
                    match_num = input("Enter match number to view details: ").strip()
                    try:
                        match_idx = int(match_num) - 1
                        if 0 <= match_idx < len(matches):
                            self._display_single_match_details(matches[match_idx])
                        else:
                            print("❌ Invalid match number.")
                    except ValueError:
                        print("❌ Please enter a valid number.")
                elif nav_choice == 'b':
                    break
                else:
                    print("❌ Invalid choice.")
                    
        except Exception as e:
            self.logger.error(f"Error viewing match list: {e}", exc_info=True)
            print(f"\n❌ Error displaying matches: {e}")

    def _set_match_filters(self, current_filters: dict) -> dict:
        """Set filters for match browsing."""
        print("\n" + "=" * 50)
        print("🔍 SET MATCH FILTERS")
        print("=" * 50)
        
        new_filters = current_filters.copy()
        
        while True:
            print("\nFilter Options:")
            print("1. 📅 Date Range")
            print("2. 🏆 Champions")
            print("3. 🎭 Roles")
            print("4. 🎯 Match Outcome")
            print("5. 👥 Players")
            print("6. ✅ Apply Filters")
            print("7. ❌ Cancel")
            
            # Show current filter status
            print("\nCurrent Filters:")
            if new_filters['date_range']:
                print(f"  📅 Date: {new_filters['date_range']['start']} to {new_filters['date_range']['end']}")
            if new_filters['champions']:
                print(f"  🏆 Champions: {len(new_filters['champions'])} selected")
            if new_filters['roles']:
                print(f"  🎭 Roles: {', '.join(new_filters['roles'])}")
            if new_filters['outcome'] is not None:
                print(f"  🎯 Outcome: {'Wins only' if new_filters['outcome'] else 'Losses only'}")
            if new_filters['players']:
                print(f"  👥 Players: {len(new_filters['players'])} selected")
            
            choice = input("\nEnter your choice (1-7): ").strip()
            
            if choice == "1":
                new_filters['date_range'] = self._set_date_filter()
            elif choice == "2":
                new_filters['champions'] = self._set_champion_filter()
            elif choice == "3":
                new_filters['roles'] = self._set_role_filter()
            elif choice == "4":
                new_filters['outcome'] = self._set_outcome_filter()
            elif choice == "5":
                new_filters['players'] = self._set_player_filter()
            elif choice == "6":
                print("\n✅ Filters applied!")
                return new_filters
            elif choice == "7":
                print("\n❌ Filter changes cancelled.")
                return current_filters
            else:
                print("\n❌ Invalid choice.")

    def _set_date_filter(self) -> Optional[dict]:
        """Set date range filter."""
        print("\n📅 Date Range Filter")
        print("-" * 20)
        
        try:
            # Get available date range from matches
            match_manager = self.engine.match_manager
            if not match_manager._matches_cache:
                print("❌ No matches available.")
                return None
            
            matches = list(match_manager._matches_cache.values())
            earliest_date = min(m.game_creation_datetime for m in matches)
            latest_date = max(m.game_creation_datetime for m in matches)
            
            print(f"Available date range: {earliest_date.strftime('%Y-%m-%d')} to {latest_date.strftime('%Y-%m-%d')}")
            
            # Get start date
            start_input = input("Enter start date (YYYY-MM-DD) or press Enter for earliest: ").strip()
            if start_input:
                start_date = datetime.strptime(start_input, "%Y-%m-%d")
            else:
                start_date = earliest_date
            
            # Get end date
            end_input = input("Enter end date (YYYY-MM-DD) or press Enter for latest: ").strip()
            if end_input:
                end_date = datetime.strptime(end_input, "%Y-%m-%d")
            else:
                end_date = latest_date
            
            if start_date > end_date:
                print("❌ Start date cannot be after end date.")
                return None
            
            return {
                'start': start_date.strftime('%Y-%m-%d'),
                'end': end_date.strftime('%Y-%m-%d')
            }
            
        except ValueError:
            print("❌ Invalid date format. Please use YYYY-MM-DD.")
            return None
        except Exception as e:
            print(f"❌ Error setting date filter: {e}")
            return None

    def _set_champion_filter(self) -> Optional[List[str]]:
        """Set champion filter."""
        print("\n🏆 Champion Filter")
        print("-" * 15)
        
        try:
            # Get champions from matches
            match_manager = self.engine.match_manager
            champion_counts = {}
            
            for match in match_manager._matches_cache.values():
                for participant in match.participants:
                    if participant.champion_name:
                        champion_counts[participant.champion_name] = champion_counts.get(participant.champion_name, 0) + 1
            
            if not champion_counts:
                print("❌ No champion data available.")
                return None
            
            # Sort by frequency
            sorted_champions = sorted(champion_counts.items(), key=lambda x: x[1], reverse=True)
            
            print("Available champions (showing top 20):")
            for i, (champion_name, count) in enumerate(sorted_champions[:20], 1):
                print(f"{i:2}. {champion_name} ({count} matches)")
            
            selection = input("\nEnter champion names separated by commas (or 'all' for all): ").strip()
            
            if selection.lower() == 'all':
                return None  # No filter means all champions
            
            if not selection:
                return None
            
            # Parse champion names
            selected_names = [name.strip() for name in selection.split(',')]
            selected_champions = []
            
            for name in selected_names:
                # Find matching champion (case insensitive)
                for champion_name, _ in champion_counts.items():
                    if champion_name.lower() == name.lower():
                        selected_champions.append(champion_name)
                        break
                else:
                    print(f"⚠️ Champion '{name}' not found in match data.")
            
            return selected_champions if selected_champions else None
            
        except Exception as e:
            print(f"❌ Error setting champion filter: {e}")
            return None

    def _set_role_filter(self) -> Optional[List[str]]:
        """Set role filter."""
        print("\n🎭 Role Filter")
        print("-" * 12)
        
        roles = ["TOP", "JUNGLE", "MIDDLE", "BOTTOM", "UTILITY"]
        
        print("Available roles:")
        for i, role in enumerate(roles, 1):
            print(f"{i}. {role}")
        
        selection = input("\nEnter role numbers separated by commas (e.g., 1,3,5): ").strip()
        
        if not selection:
            return None
        
        try:
            selected_indices = [int(x.strip()) - 1 for x in selection.split(',')]
            selected_roles = []
            
            for idx in selected_indices:
                if 0 <= idx < len(roles):
                    selected_roles.append(roles[idx])
                else:
                    print(f"❌ Invalid role number: {idx + 1}")
                    return None
            
            return selected_roles if selected_roles else None
            
        except ValueError:
            print("❌ Invalid input. Please enter numbers separated by commas.")
            return None

    def _set_outcome_filter(self) -> Optional[bool]:
        """Set match outcome filter."""
        print("\n🎯 Match Outcome Filter")
        print("-" * 20)
        
        print("Filter options:")
        print("1. Wins only")
        print("2. Losses only")
        print("3. No filter (show all)")
        
        choice = input("\nEnter your choice (1-3): ").strip()
        
        if choice == "1":
            return True
        elif choice == "2":
            return False
        elif choice == "3":
            return None
        else:
            print("❌ Invalid choice.")
            return None

    def _set_player_filter(self) -> Optional[List[str]]:
        """Set player filter."""
        print("\n👥 Player Filter")
        print("-" * 13)
        
        try:
            # Get players from our database
            players = self.engine.data_manager.load_player_data()
            
            if not players:
                print("❌ No players found in database.")
                return None
            
            print("Available players:")
            for i, player in enumerate(players, 1):
                print(f"{i:2}. {player.name} ({player.summoner_name})")
            
            selection = input("\nEnter player numbers separated by commas: ").strip()
            
            if not selection:
                return None
            
            selected_indices = [int(x.strip()) - 1 for x in selection.split(',')]
            selected_players = []
            
            for idx in selected_indices:
                if 0 <= idx < len(players):
                    selected_players.append(players[idx].puuid)
                else:
                    print(f"❌ Invalid player number: {idx + 1}")
                    return None
            
            return selected_players if selected_players else None
            
        except ValueError:
            print("❌ Invalid input. Please enter numbers separated by commas.")
            return None
        except Exception as e:
            print(f"❌ Error setting player filter: {e}")
            return None

    def _search_matches(self, filters: dict) -> None:
        """Search matches with text-based criteria."""
        print("\n" + "=" * 50)
        print("🔎 SEARCH MATCHES")
        print("=" * 50)
        
        search_term = input("Enter search term (champion name, player name, or match ID): ").strip()
        
        if not search_term:
            print("❌ Please enter a search term.")
            return
        
        try:
            matches = self._get_filtered_matches(filters)
            search_results = []
            
            search_lower = search_term.lower()
            
            for match in matches:
                # Search in match ID
                if search_lower in match.match_id.lower():
                    search_results.append((match, f"Match ID: {match.match_id}"))
                    continue
                
                # Search in participants
                for participant in match.participants:
                    if (search_lower in participant.champion_name.lower() or 
                        search_lower in participant.summoner_name.lower()):
                        reason = f"Player: {participant.summoner_name} ({participant.champion_name})"
                        search_results.append((match, reason))
                        break
            
            if not search_results:
                print(f"\n❌ No matches found for '{search_term}'.")
                return
            
            print(f"\n✅ Found {len(search_results)} matches for '{search_term}':")
            print("-" * 70)
            
            for i, (match, reason) in enumerate(search_results[:20], 1):  # Limit to 20 results
                date_str = match.game_creation_datetime.strftime("%Y-%m-%d %H:%M")
                duration_str = f"{match.game_duration // 60}:{match.game_duration % 60:02d}"
                print(f"{i:2}. {date_str} | {duration_str} | {reason}")
            
            if len(search_results) > 20:
                print(f"\n... and {len(search_results) - 20} more results")
            
            # Option to view details
            choice = input("\nEnter match number to view details (or press Enter to continue): ").strip()
            if choice:
                try:
                    match_idx = int(choice) - 1
                    if 0 <= match_idx < min(len(search_results), 20):
                        self._display_single_match_details(search_results[match_idx][0])
                    else:
                        print("❌ Invalid match number.")
                except ValueError:
                    print("❌ Please enter a valid number.")
                    
        except Exception as e:
            self.logger.error(f"Error searching matches: {e}", exc_info=True)
            print(f"\n❌ Error searching matches: {e}")

    def _view_match_details(self) -> None:
        """View detailed information for a specific match."""
        print("\n" + "=" * 50)
        print("📄 MATCH DETAILS")
        print("=" * 50)
        
        match_id = input("Enter match ID: ").strip()
        
        if not match_id:
            print("❌ Please enter a match ID.")
            return
        
        try:
            match_manager = self.engine.match_manager
            match = match_manager._matches_cache.get(match_id)
            
            if not match:
                print(f"❌ Match '{match_id}' not found.")
                return
            
            self._display_single_match_details(match)
            
        except Exception as e:
            self.logger.error(f"Error viewing match details: {e}", exc_info=True)
            print(f"\n❌ Error viewing match details: {e}")

    def _display_single_match_details(self, match) -> None:
        """Display detailed information for a single match."""
        print("\n" + "=" * 80)
        print(f"📄 MATCH DETAILS - {match.match_id}")
        print("=" * 80)
        
        # Basic match info
        date_str = match.game_creation_datetime.strftime("%Y-%m-%d %H:%M:%S")
        duration_str = f"{match.game_duration // 60}:{match.game_duration % 60:02d}"
        queue_name = self._get_queue_name(match.queue_id)
        
        print(f"📅 Date: {date_str}")
        print(f"⏱️ Duration: {duration_str}")
        print(f"🎮 Queue: {queue_name}")
        print(f"🗺️ Map: {self._get_map_name(match.map_id)}")
        print(f"🏆 Winner: {'Blue Team' if match.winning_team == 100 else 'Red Team'}")
        
        # Team compositions
        blue_team = match.get_participants_by_team(100)
        red_team = match.get_participants_by_team(200)
        
        print(f"\n🔵 BLUE TEAM {'(WINNER)' if match.winning_team == 100 else '(LOSER)'}")
        print("-" * 80)
        print(f"{'Player':<20} {'Champion':<15} {'Role':<10} {'KDA':<8} {'CS':<6} {'Damage':<10} {'Vision':<8}")
        print("-" * 80)
        
        for participant in blue_team:
            kda_str = f"{participant.kills}/{participant.deaths}/{participant.assists}"
            cs_str = str(participant.cs_total)
            damage_str = f"{participant.total_damage_dealt_to_champions:,}"
            vision_str = str(participant.vision_score)
            role_str = self._format_role(participant.individual_position or participant.lane)
            
            # Highlight known players
            player_name = participant.summoner_name
            if self._is_known_player(participant.puuid):
                player_name = f"⭐ {player_name}"
            
            print(f"{player_name:<20} {participant.champion_name:<15} {role_str:<10} {kda_str:<8} {cs_str:<6} {damage_str:<10} {vision_str:<8}")
        
        print(f"\n🔴 RED TEAM {'(WINNER)' if match.winning_team == 200 else '(LOSER)'}")
        print("-" * 80)
        print(f"{'Player':<20} {'Champion':<15} {'Role':<10} {'KDA':<8} {'CS':<6} {'Damage':<10} {'Vision':<8}")
        print("-" * 80)
        
        for participant in red_team:
            kda_str = f"{participant.kills}/{participant.deaths}/{participant.assists}"
            cs_str = str(participant.cs_total)
            damage_str = f"{participant.total_damage_dealt_to_champions:,}"
            vision_str = str(participant.vision_score)
            role_str = self._format_role(participant.individual_position or participant.lane)
            
            # Highlight known players
            player_name = participant.summoner_name
            if self._is_known_player(participant.puuid):
                player_name = f"⭐ {player_name}"
            
            print(f"{player_name:<20} {participant.champion_name:<15} {role_str:<10} {kda_str:<8} {cs_str:<6} {damage_str:<10} {vision_str:<8}")
        
        print("\n⭐ = Players in your database")

    def _export_matches(self, filters: dict) -> None:
        """Export filtered matches to various formats."""
        print("\n" + "=" * 50)
        print("📤 EXPORT MATCHES")
        print("=" * 50)
        
        try:
            matches = self._get_filtered_matches(filters)
            
            if not matches:
                print("❌ No matches to export with current filters.")
                return
            
            print(f"📊 Found {len(matches)} matches to export")
            
            print("\nExport formats:")
            print("1. 📄 CSV (Comma-separated values)")
            print("2. 📋 JSON (JavaScript Object Notation)")
            print("3. 📊 Excel (XLSX)")
            print("4. ❌ Cancel")
            
            format_choice = input("\nSelect export format (1-4): ").strip()
            
            if format_choice == "4":
                return
            
            # Get filename
            default_filename = f"matches_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            filename = input(f"Enter filename (default: {default_filename}): ").strip()
            
            if not filename:
                filename = default_filename
            
            if format_choice == "1":
                self._export_to_csv(matches, filename)
            elif format_choice == "2":
                self._export_to_json(matches, filename)
            elif format_choice == "3":
                self._export_to_excel(matches, filename)
            else:
                print("❌ Invalid format choice.")
                
        except Exception as e:
            self.logger.error(f"Error exporting matches: {e}", exc_info=True)
            print(f"\n❌ Error exporting matches: {e}")

    def _export_to_csv(self, matches: list, filename: str) -> None:
        """Export matches to CSV format."""
        import csv
        
        try:
            if not filename.endswith('.csv'):
                filename += '.csv'
            
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Header
                writer.writerow([
                    'Match ID', 'Date', 'Duration (min)', 'Queue', 'Winner',
                    'Player Name', 'Champion', 'Role', 'Team', 'Win',
                    'Kills', 'Deaths', 'Assists', 'KDA', 'CS', 'Damage', 'Vision Score'
                ])
                
                # Data rows
                for match in matches:
                    for participant in match.participants:
                        if self._is_known_player(participant.puuid):  # Only export known players
                            writer.writerow([
                                match.match_id,
                                match.game_creation_datetime.strftime('%Y-%m-%d %H:%M:%S'),
                                round(match.game_duration / 60, 1),
                                self._get_queue_name(match.queue_id),
                                'Blue' if match.winning_team == 100 else 'Red',
                                participant.summoner_name,
                                participant.champion_name,
                                self._format_role(participant.individual_position or participant.lane),
                                'Blue' if participant.team_id == 100 else 'Red',
                                'Yes' if participant.win else 'No',
                                participant.kills,
                                participant.deaths,
                                participant.assists,
                                round(participant.kda, 2),
                                participant.cs_total,
                                participant.total_damage_dealt_to_champions,
                                participant.vision_score
                            ])
            
            print(f"✅ Matches exported to {filename}")
            
        except Exception as e:
            print(f"❌ Error exporting to CSV: {e}")

    def _export_to_json(self, matches: list, filename: str) -> None:
        """Export matches to JSON format."""
        import json
        
        try:
            if not filename.endswith('.json'):
                filename += '.json'
            
            export_data = []
            
            for match in matches:
                match_data = {
                    'match_id': match.match_id,
                    'date': match.game_creation_datetime.isoformat(),
                    'duration_seconds': match.game_duration,
                    'queue_id': match.queue_id,
                    'queue_name': self._get_queue_name(match.queue_id),
                    'winning_team': match.winning_team,
                    'participants': []
                }
                
                for participant in match.participants:
                    if self._is_known_player(participant.puuid):  # Only export known players
                        participant_data = {
                            'summoner_name': participant.summoner_name,
                            'champion_name': participant.champion_name,
                            'role': self._format_role(participant.individual_position or participant.lane),
                            'team_id': participant.team_id,
                            'win': participant.win,
                            'kills': participant.kills,
                            'deaths': participant.deaths,
                            'assists': participant.assists,
                            'kda': round(participant.kda, 2),
                            'cs_total': participant.cs_total,
                            'damage_dealt': participant.total_damage_dealt_to_champions,
                            'vision_score': participant.vision_score
                        }
                        match_data['participants'].append(participant_data)
                
                if match_data['participants']:  # Only include matches with known players
                    export_data.append(match_data)
            
            with open(filename, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
            
            print(f"✅ Matches exported to {filename}")
            
        except Exception as e:
            print(f"❌ Error exporting to JSON: {e}")

    def _export_to_excel(self, matches: list, filename: str) -> None:
        """Export matches to Excel format."""
        try:
            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
            except ImportError:
                print("❌ Excel export requires openpyxl. Install with: pip install openpyxl")
                return
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Match Data"
            
            # Header row
            headers = [
                'Match ID', 'Date', 'Duration (min)', 'Queue', 'Winner',
                'Player Name', 'Champion', 'Role', 'Team', 'Win',
                'Kills', 'Deaths', 'Assists', 'KDA', 'CS', 'Damage', 'Vision Score'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Data rows
            row = 2
            for match in matches:
                for participant in match.participants:
                    if self._is_known_player(participant.puuid):  # Only export known players
                        ws.cell(row=row, column=1, value=match.match_id)
                        ws.cell(row=row, column=2, value=match.game_creation_datetime.strftime('%Y-%m-%d %H:%M:%S'))
                        ws.cell(row=row, column=3, value=round(match.game_duration / 60, 1))
                        ws.cell(row=row, column=4, value=self._get_queue_name(match.queue_id))
                        ws.cell(row=row, column=5, value='Blue' if match.winning_team == 100 else 'Red')
                        ws.cell(row=row, column=6, value=participant.summoner_name)
                        ws.cell(row=row, column=7, value=participant.champion_name)
                        ws.cell(row=row, column=8, value=self._format_role(participant.individual_position or participant.lane))
                        ws.cell(row=row, column=9, value='Blue' if participant.team_id == 100 else 'Red')
                        ws.cell(row=row, column=10, value='Yes' if participant.win else 'No')
                        ws.cell(row=row, column=11, value=participant.kills)
                        ws.cell(row=row, column=12, value=participant.deaths)
                        ws.cell(row=row, column=13, value=participant.assists)
                        ws.cell(row=row, column=14, value=round(participant.kda, 2))
                        ws.cell(row=row, column=15, value=participant.cs_total)
                        ws.cell(row=row, column=16, value=participant.total_damage_dealt_to_champions)
                        ws.cell(row=row, column=17, value=participant.vision_score)
                        row += 1
            
            wb.save(filename)
            print(f"✅ Matches exported to {filename}")
            
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")

    def _match_statistics(self, filters: dict) -> None:
        """Display statistics for filtered matches."""
        print("\n" + "=" * 60)
        print("📊 MATCH STATISTICS")
        print("=" * 60)
        
        try:
            matches = self._get_filtered_matches(filters)
            
            if not matches:
                print("❌ No matches found with current filters.")
                return
            
            # Basic statistics
            total_matches = len(matches)
            total_duration = sum(match.game_duration for match in matches)
            avg_duration = total_duration / total_matches if total_matches > 0 else 0
            
            # Date range
            if matches:
                earliest = min(match.game_creation_datetime for match in matches)
                latest = max(match.game_creation_datetime for match in matches)
                date_range = (latest - earliest).days
            else:
                earliest = latest = None
                date_range = 0
            
            print(f"📈 Overview:")
            print(f"   Total Matches: {total_matches:,}")
            print(f"   Average Duration: {int(avg_duration // 60)}:{int(avg_duration % 60):02d}")
            print(f"   Date Range: {earliest.strftime('%Y-%m-%d') if earliest else 'N/A'} to {latest.strftime('%Y-%m-%d') if latest else 'N/A'} ({date_range} days)")
            
            # Queue type distribution
            queue_counts = {}
            for match in matches:
                queue_name = self._get_queue_name(match.queue_id)
                queue_counts[queue_name] = queue_counts.get(queue_name, 0) + 1
            
            print(f"\n🎮 Queue Distribution:")
            for queue_name, count in sorted(queue_counts.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_matches) * 100
                print(f"   {queue_name}: {count} ({percentage:.1f}%)")
            
            # Champion statistics (for known players only)
            champion_stats = {}
            known_player_matches = 0
            
            for match in matches:
                for participant in match.participants:
                    if self._is_known_player(participant.puuid):
                        known_player_matches += 1
                        champion = participant.champion_name
                        if champion not in champion_stats:
                            champion_stats[champion] = {'games': 0, 'wins': 0}
                        
                        champion_stats[champion]['games'] += 1
                        if participant.win:
                            champion_stats[champion]['wins'] += 1
            
            if champion_stats:
                print(f"\n🏆 Champion Performance (Known Players, {known_player_matches} total games):")
                # Sort by games played
                sorted_champions = sorted(champion_stats.items(), key=lambda x: x[1]['games'], reverse=True)
                
                for champion, stats in sorted_champions[:10]:  # Top 10
                    win_rate = (stats['wins'] / stats['games']) * 100 if stats['games'] > 0 else 0
                    print(f"   {champion}: {stats['games']} games, {win_rate:.1f}% win rate")
            
            # Win rate by team side
            blue_wins = sum(1 for match in matches if match.winning_team == 100)
            red_wins = total_matches - blue_wins
            blue_win_rate = (blue_wins / total_matches) * 100 if total_matches > 0 else 0
            
            print(f"\n⚖️ Team Side Balance:")
            print(f"   Blue Team Wins: {blue_wins} ({blue_win_rate:.1f}%)")
            print(f"   Red Team Wins: {red_wins} ({100 - blue_win_rate:.1f}%)")
            
        except Exception as e:
            self.logger.error(f"Error calculating match statistics: {e}", exc_info=True)
            print(f"\n❌ Error calculating statistics: {e}")

    def _clear_filters(self) -> dict:
        """Clear all active filters."""
        print("\n🗑️ All filters cleared!")
        return {
            'date_range': None,
            'champions': None,
            'roles': None,
            'outcome': None,
            'players': None
        }

    def _get_filtered_matches(self, filters: dict) -> list:
        """Get matches that match the current filters."""
        match_manager = self.engine.match_manager
        matches = list(match_manager._matches_cache.values())
        
        # Apply date filter
        if filters['date_range']:
            start_date = datetime.strptime(filters['date_range']['start'], '%Y-%m-%d')
            end_date = datetime.strptime(filters['date_range']['end'], '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)  # Include full end date
            
            matches = [m for m in matches if start_date <= m.game_creation_datetime <= end_date]
        
        # Apply champion filter
        if filters['champions']:
            matches = [m for m in matches if any(
                p.champion_name in filters['champions'] and self._is_known_player(p.puuid)
                for p in m.participants
            )]
        
        # Apply role filter
        if filters['roles']:
            matches = [m for m in matches if any(
                (p.individual_position in filters['roles'] or p.lane in filters['roles']) and self._is_known_player(p.puuid)
                for p in m.participants
            )]
        
        # Apply outcome filter
        if filters['outcome'] is not None:
            filtered_matches = []
            for match in matches:
                known_players = [p for p in match.participants if self._is_known_player(p.puuid)]
                if known_players:
                    # Check if any known player has the desired outcome
                    if any(p.win == filters['outcome'] for p in known_players):
                        filtered_matches.append(match)
            matches = filtered_matches
        
        # Apply player filter
        if filters['players']:
            matches = [m for m in matches if any(
                p.puuid in filters['players'] for p in m.participants
            )]
        
        return matches

    def _get_queue_name(self, queue_id: int) -> str:
        """Get human-readable queue name."""
        queue_names = {
            420: "Ranked Solo",
            440: "Ranked Flex",
            400: "Normal Draft",
            430: "Normal Blind",
            450: "ARAM",
            900: "URF",
            1020: "One for All",
            1300: "Nexus Blitz",
            1400: "Ultimate Spellbook"
        }
        return queue_names.get(queue_id, f"Queue {queue_id}")

    def _get_map_name(self, map_id: int) -> str:
        """Get human-readable map name."""
        map_names = {
            11: "Summoner's Rift",
            12: "Howling Abyss",
            21: "Nexus Blitz"
        }
        return map_names.get(map_id, f"Map {map_id}")

    def _format_role(self, role: str) -> str:
        """Format role for display."""
        if not role:
            return "Unknown"
        
        role_mapping = {
            "TOP": "Top",
            "JUNGLE": "Jungle", 
            "MIDDLE": "Mid",
            "BOTTOM": "ADC",
            "UTILITY": "Support"
        }
        
        return role_mapping.get(role.upper(), role.title())

    def _is_known_player(self, puuid: str) -> bool:
        """Check if a player is in our database."""
        try:
            players = self.engine.data_manager.load_player_data()
            return any(player.puuid == puuid for player in players)
        except:
            return False

    def _get_known_players_in_match(self, match) -> list:
        """Get list of known players in a match."""
        known_players = []
        try:
            players = self.engine.data_manager.load_player_data()
            known_puuids = {player.puuid for player in players}
            
            for participant in match.participants:
                if participant.puuid in known_puuids:
                    known_players.append(participant)
        except:
            pass
        
        return known_players

    def _get_match_result_summary(self, match, known_players: list) -> str:
        """Get a summary of match results for known players."""
        if not known_players:
            return "N/A"
        
        wins = sum(1 for p in known_players if p.win)
        losses = len(known_players) - wins
        
        if wins > 0 and losses > 0:
            return f"W{wins}L{losses}"
        elif wins > 0:
            return f"W{wins}"
        else:
            return f"L{losses}"

    def _interactive_analytics_dashboard(self) -> None:
        """Launch the interactive analytics dashboard."""
        print("\n📊 Launching Interactive Analytics Dashboard...")
        
        try:
            # Check if required analytics engines are available
            if not hasattr(self.engine, 'historical_analytics_engine') or not self.engine.historical_analytics_engine:
                print("❌ Historical analytics engine not available.")
                print("Please ensure the system is properly configured with analytics capabilities.")
                return
            
            # Import and initialize the dashboard
            from .interactive_analytics_dashboard import InteractiveAnalyticsDashboard
            
            # Get required components from the engine
            analytics_engine = self.engine.historical_analytics_engine
            recommendation_engine = getattr(self.engine, 'champion_recommendation_engine', None)
            composition_analyzer = getattr(self.engine, 'team_composition_analyzer', None)
            comparative_analyzer = getattr(self.engine, 'comparative_analyzer', None)
            data_manager = self.engine.data_manager
            
            # Check for missing components
            missing_components = []
            if not recommendation_engine:
                missing_components.append("Champion Recommendation Engine")
            if not composition_analyzer:
                missing_components.append("Team Composition Analyzer")
            if not comparative_analyzer:
                missing_components.append("Comparative Analyzer")
            
            if missing_components:
                print(f"⚠️ Some components are not available: {', '.join(missing_components)}")
                print("Dashboard will run with limited functionality.")
                
                # Create placeholder components if needed
                if not recommendation_engine:
                    recommendation_engine = None
                if not composition_analyzer:
                    composition_analyzer = None
                if not comparative_analyzer:
                    comparative_analyzer = None
            
            # Initialize and launch dashboard
            dashboard = InteractiveAnalyticsDashboard(
                analytics_engine=analytics_engine,
                recommendation_engine=recommendation_engine,
                composition_analyzer=composition_analyzer,
                comparative_analyzer=comparative_analyzer,
                data_manager=data_manager
            )
            
            print("✅ Dashboard initialized successfully")
            dashboard.show_dashboard()
            
        except ImportError as e:
            self.logger.error(f"Failed to import dashboard: {e}")
            print(f"❌ Failed to load interactive dashboard: {e}")
            print("Please ensure all required components are installed.")
        except Exception as e:
            self.logger.error(f"Dashboard error: {e}", exc_info=True)
            print(f"❌ Dashboard error: {e}")
            print("Please check the logs for more details.")


def main():
    """Main entry point for the streamlined CLI."""
    cli = StreamlinedCLI()
    cli.main()


if __name__ == "__main__":
    main()